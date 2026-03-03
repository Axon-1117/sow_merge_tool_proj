"""GUI self-test: bottom bar remains aligned when A has no save button.

Run:
  .venv\\Scripts\\python.exe _gui_self_test_bottom_bar_aligned.py
"""

import os
import tempfile
from openpyxl import Workbook


def _make_xlsx(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.cell(row=1, column=1).value = "h1"
    ws.cell(row=2, column=1).value = "v1"
    wb.save(path)


def main():
    td = tempfile.mkdtemp(prefix="sow_merge_gui_test_bottom_align_")
    fb_dir = os.path.join(td, "normal")
    os.makedirs(fb_dir, exist_ok=True)

    # file_a under temp dir: SaveA button should be hidden by design.
    fa = os.path.join(td, "a.xlsx")
    fb = os.path.join(fb_dir, "b.xlsx")
    _make_xlsx(fa)
    _make_xlsx(fb)

    import sow_merge_tool as mod

    app = mod.SowMergeApp(fa, fb)
    sheet = app.common_sheets[0]
    view = app.sheet_views.get(sheet)
    if view is None:
        app.nb.select(app._sheet_containers[sheet])
        app.root.update_idletasks()
        app.root.update()
        view = app.sheet_views[sheet]

    app.root.update_idletasks()
    app.root.update()

    y_left = view.hsb_left.winfo_rooty()
    y_right = view.hsb_right.winfo_rooty()
    assert y_left == y_right, f"Expected aligned horizontal bars; left_y={y_left} right_y={y_right}"

    try:
        app.root.destroy()
    except Exception:
        pass

    print("GUI_SELF_TEST_BOTTOM_BAR_ALIGNED_OK")


if __name__ == "__main__":
    main()
