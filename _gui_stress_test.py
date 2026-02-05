"""Stress test for sow_merge_tool (no desktop automation).

Creates two large XLSX files (same basename in different dirs) with multiple sheets,
then starts SowMergeApp and measures:
- XLSX generation time
- App init time (workbook load)
- Time to get sample diff states populated
- Time for background confirmation to complete (all sheets reach state 0 or 2)

Run:
  .venv\Scripts\python.exe _gui_stress_test.py

You can override defaults via env vars:
  SOW_STRESS_ROWS=10000
  SOW_STRESS_COLS=30
  SOW_STRESS_SHEETS=6
  SOW_STRESS_DIFF_EVERY=50   (every N rows introduce diffs)
  SOW_STRESS_DIFF_COL_EVERY=7 (every N cols introduce diffs)
"""

import os
import tempfile
import time
from openpyxl import Workbook


def env_int(name: str, default: int) -> int:
    try:
        return int(os.environ.get(name, default))
    except Exception:
        return default


def make_big_xlsx(path: str, sheets: int, rows: int, cols: int, diff: bool, diff_every: int, diff_col_every: int):
    wb = Workbook()
    # remove default
    wb.remove(wb.active)

    header = [f"H{c}" for c in range(1, cols + 1)]

    for si in range(sheets):
        name = f"S{si+1}"
        ws = wb.create_sheet(title=name)
        # header
        for c, v in enumerate(header, start=1):
            ws.cell(row=1, column=c).value = v

        for r in range(2, rows + 2):
            base_row = [f"R{r-1}C{c}" for c in range(1, cols + 1)]
            # introduce diffs on B file only
            if diff:
                if (r % diff_every) == 0:
                    for c in range(1, cols + 1):
                        if (c % diff_col_every) == 0:
                            base_row[c - 1] = f"DIFF_R{r-1}C{c}"
            for c, v in enumerate(base_row, start=1):
                ws.cell(row=r, column=c).value = v

    wb.save(path)


def pump(root, seconds: float):
    end = time.time() + seconds
    while time.time() < end:
        try:
            root.update_idletasks(); root.update()
        except Exception:
            pass
        time.sleep(0.01)


def main():
    rows = env_int("SOW_STRESS_ROWS", 10000)
    cols = env_int("SOW_STRESS_COLS", 30)
    sheets = env_int("SOW_STRESS_SHEETS", 6)
    diff_every = env_int("SOW_STRESS_DIFF_EVERY", 50)
    diff_col_every = env_int("SOW_STRESS_DIFF_COL_EVERY", 7)

    td1 = tempfile.mkdtemp(prefix="sow_stress_a_")
    td2 = tempfile.mkdtemp(prefix="sow_stress_b_")
    fa = os.path.join(td1, "same.xlsx")
    fb = os.path.join(td2, "same.xlsx")

    t0 = time.perf_counter()
    make_big_xlsx(fa, sheets, rows, cols, diff=False, diff_every=diff_every, diff_col_every=diff_col_every)
    t1 = time.perf_counter()
    make_big_xlsx(fb, sheets, rows, cols, diff=True, diff_every=diff_every, diff_col_every=diff_col_every)
    t2 = time.perf_counter()

    import sys
    sys.path.insert(0, r"D:\Tools\sow_merge_tool")
    import sow_merge_tool as mod

    t3 = time.perf_counter()
    app = mod.SowMergeApp(fa, fb)
    t4 = time.perf_counter()

    # Wait a bit for sample scan to run
    pump(app.root, 2.0)
    t5 = time.perf_counter()
    maybe_count = sum(1 for v in app.sheet_diff_state.values() if v == 1)
    confirmed_count = sum(1 for v in app.sheet_diff_state.values() if v == 2)

    # Wait for background confirmation (up to 60s)
    start_wait = time.perf_counter()
    while True:
        pump(app.root, 0.3)
        states = list(app.sheet_diff_state.values())
        if all(v in (0, 2) for v in states):
            break
        if (time.perf_counter() - start_wait) > 60:
            break

    t6 = time.perf_counter()

    done_states = app.sheet_diff_state.copy()

    try:
        app.root.destroy()
    except Exception:
        pass

    gen_a_s = (t1 - t0)
    gen_b_s = (t2 - t1)
    app_init_s = (t4 - t3)
    sample_s = (t5 - t4)
    confirm_s = (t6 - t4)

    print("STRESS_TEST_PARAMS", {"rows": rows, "cols": cols, "sheets": sheets, "diff_every": diff_every, "diff_col_every": diff_col_every})
    print(f"GEN_A_S={gen_a_s:.2f}")
    print(f"GEN_B_S={gen_b_s:.2f}")
    print(f"APP_INIT_S={app_init_s:.2f}")
    print(f"AFTER_2S_SAMPLE maybe={maybe_count} confirmed={confirmed_count} states={done_states}")
    print(f"CONFIRM_TOTAL_S={confirm_s:.2f}")


if __name__ == "__main__":
    main()
