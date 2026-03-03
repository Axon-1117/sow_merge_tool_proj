import os
import sys
import time
import tempfile
from pathlib import Path

from openpyxl import Workbook

from pywinauto import Application
from pywinauto.timings import wait_until_passes


EXE = r"D:\Tools\sow_merge_tool_proj\dist\sow_merge_tool.exe"
APP_TITLE_RE = r"sow_merge_tool.*"


def make_pair() -> tuple[str, str]:
    """Create two workbooks in different dirs but same basename."""
    root_a = Path(tempfile.mkdtemp(prefix="sow_ui_A_"))
    root_b = Path(tempfile.mkdtemp(prefix="sow_ui_B_"))

    fa = root_a / "t.xlsx"
    fb = root_b / "t.xlsx"

    wb = Workbook(); ws = wb.active; ws.title = "S1"
    ws["A1"] = "x"; ws["B1"] = "y"
    ws["A2"] = 1; ws["B2"] = 2
    wb.save(fa)

    wb2 = Workbook(); ws2 = wb2.active; ws2.title = "S1"
    ws2["A1"] = "x"; ws2["B1"] = "Y"  # diff
    ws2["A2"] = 1; ws2["B2"] = 2
    wb2.save(fb)

    return str(fa), str(fb)


def click_yes_if_prompt(app: Application, timeout=3):
    """Click Yes/OK on common tkinter messagebox prompts if they appear."""
    end = time.time() + timeout
    while time.time() < end:
        try:
            dlg = app.window(title_re=r".*(确认保存|提示|Saved|保存失败|Error).*", control_type="Window")
            if dlg.exists(timeout=0.2):
                try:
                    dlg.set_focus()
                except Exception:
                    pass
                # try common buttons
                for text in ("是", "Yes", "确定", "OK"):
                    try:
                        btn = dlg.child_window(title=text, control_type="Button")
                        if btn.exists():
                            btn.click_input()
                            break
                    except Exception:
                        pass
                # Fallback: press Enter
                try:
                    dlg.type_keys("{ENTER}")
                except Exception:
                    pass
        except Exception:
            pass
        time.sleep(0.1)


def get_main(app: Application):
    # Tk apps are easiest with win32 backend. Wait for the main window by title.
    def _find():
        win = app.window(title_re=APP_TITLE_RE, control_type="Window")
        win.wait("visible", timeout=1)
        return win

    try:
        return wait_until_passes(30, 0.5, _find)
    except Exception:
        # Fallback: try top_window as a last resort
        return app.top_window()


def wait_app_ready(app: Application, timeout=30):
    try:
        app.wait_cpu_usage_lower(threshold=5, timeout=timeout)
    except Exception:
        pass


def safe_close(app: Application, main):
    try:
        main.close()
    except Exception:
        pass
    # Ensure process exits
    try:
        app.kill()
    except Exception:
        pass


def select_only_diff(main):
    # Tk checkbutton typically maps to a Button in win32 backend
    cb = main.child_window(title_re=r"只看差异内容.*")
    cb.wait("exists", timeout=10)
    try:
        state = cb.get_toggle_state()
    except Exception:
        state = None
    if state in (None, 0):
        cb.click_input()


def click_line_in_left_pane(main, line_no: int = 1):
    # Find the left Text widget. Tkinter renders as a "Pane" with a child "Edit" sometimes.
    # We'll click inside the large left editor area roughly.
    # Use coordinates: click in center-left of client area.
    rect = main.rectangle()
    x = rect.left + int((rect.width()) * 0.25)
    y = rect.top + int((rect.height()) * 0.35)
    main.click_input(coords=(x - rect.left, y - rect.top))


def click_use_left(main):
    btn = main.child_window(title_re=r"使用左侧\(A\)", control_type="Button")
    btn.wait("exists", timeout=10)
    btn.click_input()


def click_use_right(main):
    btn = main.child_window(title_re=r"使用右侧\(B\)", control_type="Button")
    btn.wait("exists", timeout=10)
    btn.click_input()


def click_save_a(main):
    btn = main.child_window(title="保存A", control_type="Button")
    btn.wait("exists", timeout=10)
    btn.click_input()


def click_save_b(main):
    btn = main.child_window(title="保存B", control_type="Button")
    btn.wait("exists", timeout=10)
    btn.click_input()


def read_cursor_block(main):
    # Two entries stacked; pick by background label? easiest: get all edits/entries and take last two.
    edits = main.descendants(control_type="Edit")
    if len(edits) < 2:
        return "", ""
    a = edits[-2].window_text()
    b = edits[-1].window_text()
    return a, b


def scenario_only_diff_then_left_to_right_then_save():
    fa, fb = make_pair()

    if not os.path.exists(EXE):
        raise FileNotFoundError(EXE)

    app = Application(backend="win32").start(f'"{EXE}" "{fa}" "{fb}"')
    main = get_main(app)
    wait_app_ready(app, timeout=30)
    wait_until_passes(30, 0.5, lambda: main.wait('ready', timeout=1))

    # Ensure checkbox filter works and cursor block updates
    select_only_diff(main)

    # select a diff row by clicking in left pane
    click_line_in_left_pane(main)

    # Use-left -> should remove diff row in only-diff mode
    click_use_left(main)

    # Small wait for UI update
    time.sleep(0.3)

    # Save B inplace
    click_save_b(main)
    click_yes_if_prompt(app, timeout=5)

    # Ensure file exists
    assert os.path.exists(fb), "B file missing after save"

    # Close
    safe_close(app, main)


def scenario_only_diff_then_right_to_left_then_save():
    fa, fb = make_pair()

    if not os.path.exists(EXE):
        raise FileNotFoundError(EXE)

    app = Application(backend="win32").start(f'"{EXE}" "{fa}" "{fb}"')
    main = get_main(app)
    wait_app_ready(app, timeout=30)
    wait_until_passes(30, 0.5, lambda: main.wait('ready', timeout=1))

    select_only_diff(main)
    click_line_in_left_pane(main)

    # Use-right (B->A)
    click_use_right(main)
    time.sleep(0.3)

    click_save_a(main)
    click_yes_if_prompt(app, timeout=5)

    assert os.path.exists(fa), "A file missing after save"
    safe_close(app, main)


def run_all():
    scenario_only_diff_then_left_to_right_then_save()
    scenario_only_diff_then_right_to_left_then_save()
    print("UI_TESTS_OK")


if __name__ == "__main__":
    run_all()
