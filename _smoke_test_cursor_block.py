"""Smoke test for the per-sheet cursor compare block.

Validates:
- cursor block always shows 2 lines
- line1 mirrors left pane cursor row text
- line2 mirrors right pane cursor row text

Run with: D:\\Tools\\sow_merge_tool\\.venv\\Scripts\\python.exe _smoke_test_cursor_block.py
"""

from openpyxl import Workbook
import os, tempfile
import importlib.util

root_a = tempfile.mkdtemp(prefix='sow_cursor_A_')
root_b = tempfile.mkdtemp(prefix='sow_cursor_B_')
fa = os.path.join(root_a, 't.xlsx')
fb = os.path.join(root_b, 't.xlsx')

wb = Workbook(); ws = wb.active; ws.title = 'S1'
ws['A1'] = 'x'; ws['B1'] = 'y'
ws['A2'] = 'left2'; ws['B2'] = 'left2b'
wb.save(fa)

wb2 = Workbook(); ws2 = wb2.active; ws2.title = 'S1'
ws2['A1'] = 'x'; ws2['B1'] = 'Y'
ws2['A2'] = 'right2'; ws2['B2'] = 'right2b'
wb2.save(fb)

spec = importlib.util.spec_from_file_location('sow', r'D:\Tools\sow_merge_tool\sow_merge_tool.py')
mod = importlib.util.module_from_spec(spec)
spec.loader.exec_module(mod)

app = mod.SowMergeApp(fa, fb)

# Lazy SheetView creation
sheet = 'S1'
view = app.sheet_views.get(sheet)
if view is None:
    app.nb.select(app._sheet_containers[sheet])
    try:
        app.root.update_idletasks(); app.root.update()
    except Exception:
        pass
    view = app.sheet_views[sheet]

view.refresh(row_only=None, rescan=True)

# Put left cursor on line 2, right cursor on line 1
view.left.mark_set('insert', '2.0')
view.right.mark_set('insert', '1.0')
view._update_cursor_lines()

line1 = view.cursor_cmp.get('1.0', '1.end')
line2 = view.cursor_cmp.get('2.0', '2.end')

assert line1.startswith('2\t') or line1.startswith('2\t') or line1.startswith('2\t') or line1.startswith('2\t')
assert line1.startswith('2\t') or line1.startswith('2\t') or line1.startswith('2\t')
# easier: check it contains left2
assert 'left2' in line1, (line1, line2)
assert 'x' in line2 and 'Y' in line2, (line1, line2)

# Ensure correct order: A on top, B on bottom
assert 'left2' in line1 and 'Y' in line2

app.root.destroy()
print('SMOKE_CURSOR_BLOCK_OK')
