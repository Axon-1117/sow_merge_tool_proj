from openpyxl import Workbook
import os, tempfile
import importlib.util

# create two small workbooks in separate dirs but same basename
root_a = tempfile.mkdtemp(prefix='sow_merge_testA_')
root_b = tempfile.mkdtemp(prefix='sow_merge_testB_')

fa = os.path.join(root_a, 't.xlsx')
fb = os.path.join(root_b, 't.xlsx')

wb = Workbook(); ws = wb.active; ws.title = 'S1'
ws['A1'] = 'x'; ws['B1'] = 'y'
ws['A2'] = 1; ws['B2'] = 2
wb.save(fa)

wb2 = Workbook(); ws2 = wb2.active; ws2.title = 'S1'
ws2['A1'] = 'x'; ws2['B1'] = 'Y'
ws2['A2'] = 1; ws2['B2'] = 2
wb2.save(fb)

spec = importlib.util.spec_from_file_location('sow', r'D:\Tools\sow_merge_tool\sow_merge_tool.py')
mod = importlib.util.module_from_spec(spec)
spec.loader.exec_module(mod)

app = mod.SowMergeApp(fa, fb)

# Lazy SheetView creation: select the tab to instantiate it
sheet = 'S1'
view = app.sheet_views.get(sheet)
if view is None:
    app.nb.select(app._sheet_containers[sheet])
    try:
        app.root.update_idletasks(); app.root.update()
    except Exception:
        pass
    view = app.sheet_views[sheet]

# Force synchronous rescan (rescan=True bypasses the background-wait guard)
view.refresh(row_only=None, rescan=True)
assert view._data_ready, 'refresh(rescan=True) should set _data_ready'

# Row 1 of A differs from row 1 of B (B1: 'y' vs 'Y'); look up via pair index
pair_idx_r1 = view.row_a_to_pair_idx.get(1)
assert pair_idx_r1 is not None, f'row 1 not found in row_a_to_pair_idx: {view.row_a_to_pair_idx}'
assert view.pair_diff_cols.get(pair_idx_r1), f'Row1 should have diff; pair_diff_cols={view.pair_diff_cols}'

view.only_diff_var.set(True)
view._toggle_only_diff()
assert pair_idx_r1 in view.display_rows, f'only-diff failed: {view.display_rows}'

# Select pair (row 1) and merge A -> B
view.selected_pair_idx = pair_idx_r1
view._copy_selected_row('A2B')
# Snapshot mode keeps row visible; diff cols should be cleared after overwrite
assert not view.pair_diff_cols.get(pair_idx_r1), \
    f'merge did not clear diff; pair_diff_cols[{pair_idx_r1}]={view.pair_diff_cols.get(pair_idx_r1)}'

app.root.destroy()
print('SMOKE_TEST_OK')
