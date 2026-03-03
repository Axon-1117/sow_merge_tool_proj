"""GUI self-test: verifies sampled (maybe) and confirmed sheet diff state updates.

Run:
  .venv\\Scripts\\python.exe _gui_self_test_sheet_diff_state.py

No desktop automation required.
"""

import os
import tempfile
import time

from openpyxl import Workbook


def _make_xlsx(path: str, sheets: dict[str, list[list[object]]]):
    wb = Workbook()
    # remove default sheet
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, v in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx).value = v
    wb.save(path)


def _pump(root, seconds=1.0):
    end = time.time() + seconds
    while time.time() < end:
        try:
            root.update_idletasks(); root.update()
        except Exception:
            pass
        time.sleep(0.01)


def main():
    td1 = tempfile.mkdtemp(prefix="sow_state_a_")
    td2 = tempfile.mkdtemp(prefix="sow_state_b_")
    fa = os.path.join(td1, "same.xlsx")
    fb = os.path.join(td2, "same.xlsx")

    base = [["h1", "h2"], [1, 2], [3, 4]]
    sheets_a = {"S_ok": base, "S_diff": base, "S_diff2": base}
    sheets_b = {"S_ok": base, "S_diff": [["h1", "h2"], [1, 999], [3, 4]], "S_diff2": [["h1", "h2"], [1, 2], [3, 888]]}

    _make_xlsx(fa, sheets_a)
    _make_xlsx(fb, sheets_b)

    import sow_merge_tool as mod

    app = mod.SowMergeApp(fa, fb)

    # Give sample scan + background compute some time
    _pump(app.root, seconds=2.5)

    # After background confirmation, diff sheets should be confirmed (2) and ok sheet should be 0
    st = app.sheet_diff_state
    assert st.get("S_ok") == 0, f"Expected S_ok=0, got {st.get('S_ok')}"
    assert st.get("S_diff") == 2, f"Expected S_diff=2, got {st.get('S_diff')}"
    assert st.get("S_diff2") == 2, f"Expected S_diff2=2, got {st.get('S_diff2')}"

    try:
        app.root.destroy()
    except Exception:
        pass

    print("GUI_SELF_TEST_SHEET_DIFF_STATE_OK")


if __name__ == "__main__":
    main()
