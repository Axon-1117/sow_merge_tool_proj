import os
import sys
import argparse
import re
import difflib
import tempfile
import subprocess
import traceback
from datetime import datetime
import time
import stat
import shutil
import zipfile
import platform
import xml.etree.ElementTree as ET

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import json
import threading

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.formula import ArrayFormula
# Note: formulas will be treated as cached values only (data_only), with fallback when cache is missing.
from openpyxl.utils import get_column_letter


APP_NAME = "sow_merge_tool"
APP_VERSION = "2026-03-07.update17"
APP_BUILD_TAG = "new94-rowheader-bg-restore"

# Debug logging (writes to %TEMP%\sow_merge_tool_debug.log)
_DEBUG_LOG_PATH = os.path.join(tempfile.gettempdir(), f"{APP_NAME}_debug.log")
_DEBUG_ENABLED = False
_LAUNCH_TRACE_PATH = os.path.join(tempfile.gettempdir(), f"{APP_NAME}_launch_trace.log")

# Save performance: fast mode writes directly to target (faster, less safe than atomic replace)
_FAST_SAVE_ENABLED = True
# Save correctness: keep workbook fidelity (styles/formulas/metadata).
# values-only fast save can make unrelated sheets look modified in SVN diff.
_FAST_SAVE_VALUES_ONLY = False
# Open performance: skip background preloads and global scans (loads on demand)
_FAST_OPEN_ENABLED = True
# Global mode: compare and save cached values only (ignore formulas)
_USE_CACHED_VALUES_ONLY = True
# When cached values are missing for formulas, try to recalc via Excel (if available)
_AUTO_RECALC_MISSING_CACHE = False
_AUTO_RECALC_FORMULAS_ALWAYS = False
_AUTO_RECALC_ON_OPEN = True
_EXCEL_NATIVE_SAVE_ON_MERGE = True
_CACHE_CHECK_MAX_CELLS = 3000
# Render performance: limit initial rows rendered (user can load full)
_FAST_RENDER_ROW_LIMIT = 800
_FAST_RENDER_BATCH = 500
_LARGE_SHEET_ROW_THRESHOLD = 1000
_LARGE_SHEET_INITIAL_ROWS = 200
_LARGE_SHEET_BLOCK_ROWS = 1000
_LARGE_SHEET_DIRECT_PAIR_THRESHOLD = 5000
_ROW_ALIGN_MAX_ROWS = 1000
_TABMARK_QUICK_TAIL_ROWS = 2000
# Fast tab-mark pre-scan can duplicate workbook open cost on huge files.
# Skip it above this size and rely on the normal background compute path.
_FAST_TABMARK_SCAN_SKIP_MB = 25
_FAST_TABMARK_PHASE2_ENABLED = False
# Grid display: max chars shown per cell before truncation, and column separator
_COL_MAX_DISPLAY_WIDTH = 30
_COL_SEP = " \u2502 "    # 3-char separator between columns (U+2502 BOX DRAWINGS LIGHT VERTICAL)
_COL_SEP_LEN = 3

# Unified pane colors (main 3-way panes and C-area rows)
_MINE_BG = "#F6C16B"
_BASE_BG = "#E3E3FF"
_THEIRS_BG = "#FFF176"
_DIFF_CELL_BG = "#FF2D2D"

# Unified row-header hover/action arrows (keep one visual family).
_ROW_ARROW_RIGHT = "➡"
_ROW_ARROW_LEFT = "⬅"

# Settings (persist UI prefs)
_SETTINGS_PATH = os.path.join(os.environ.get("LOCALAPPDATA", tempfile.gettempdir()), APP_NAME, "settings.json")


def _dlog(msg: str):
    if not _DEBUG_ENABLED:
        return
    try:
        ts = datetime.now().strftime("%H:%M:%S.%f")[:-3]
        with open(_DEBUG_LOG_PATH, "a", encoding="utf-8") as f:
            f.write(f"[{ts}] {msg}\n")
    except Exception:
        pass


def _val_to_str(v):
    """Render a cell value as single-line text for the Text widget.

    IMPORTANT: We must keep each Excel row rendered as exactly ONE line in tk.Text.
    So we sanitize embedded newlines/tabs that would otherwise break line alignment
    and cause diff highlights to drift.
    """
    if v is None:
        return ""
    s = str(v)
    # Normalize line breaks and tabs to keep one-row-per-line invariant
    s = s.replace("\r\n", " ").replace("\r", " ").replace("\n", " ")
    s = s.replace("\t", "    ")
    return s


def _format_cell(val_str: str, width: int) -> str:
    """Left-justify val_str in exactly `width` chars; truncate with ellipsis if too long."""
    if len(val_str) > width:
        return val_str[:width - 1] + "\u2026"  # … (single char)
    return val_str.ljust(width)


def _effective_bounds(ws):
    """Return (max_row, max_col) based on actual non-empty cells.

    Some workbooks have an inaccurate ws.max_row/ws.max_column (e.g. only first
    N rows reported). We scan ws._cells (when available) to derive a safer bound.
    """
    max_r = ws.max_row or 1
    max_c = ws.max_column or 1
    try:
        cells = getattr(ws, "_cells", None)
        if cells:
            last_r = 1
            last_c = 1
            found = False
            for cell in cells.values():
                v = cell.value
                if v not in (None, ""):
                    found = True
                    if cell.row > last_r:
                        last_r = cell.row
                    if cell.column > last_c:
                        last_c = cell.column
            if found:
                max_r = max(max_r, last_r)
                max_c = max(max_c, last_c)
    except Exception:
        pass
    return max(1, max_r), max(1, max_c)


def _save_values_only_from_wb(src_wb, target_path: str):
    """Fast save: values only, no styles. Drops formatting."""
    def _trim_bounds_ws(ws):
        # Find last non-empty row/col (scan ws._cells then fallback)
        max_r = ws.max_row or 1
        max_c = ws.max_column or 1
        last_r = 1
        last_c = 1
        found = False
        try:
            cells = getattr(ws, "_cells", None)
            if cells:
                for cell in cells.values():
                    v = cell.value
                    if v not in (None, ""):
                        found = True
                        if cell.row > last_r:
                            last_r = cell.row
                        if cell.column > last_c:
                            last_c = cell.column
        except Exception:
            pass
        if not found:
            for r in range(max_r, max(1, max_r - 5000), -1):
                row = next(ws.iter_rows(min_row=r, max_row=r, min_col=1, max_col=max_c, values_only=True), ())
                if any(v not in (None, "") for v in row):
                    found = True
                    last_r = r
                    for ci in range(len(row), 0, -1):
                        v = row[ci - 1]
                        if v not in (None, ""):
                            last_c = ci
                            break
                    break
        if not found:
            return max_r, max_c
        use_r = min(max_r, last_r + 50)
        use_c = min(max_c, last_c + 50)
        return max(1, use_r), max(1, use_c)

    dst = Workbook(write_only=True)
    # Remove default sheet
    try:
        if dst.sheetnames:
            dst.remove(dst[dst.sheetnames[0]])
    except Exception:
        pass
    for name in src_wb.sheetnames:
        ws_src = src_wb[name]
        ws_dst = dst.create_sheet(title=name)
        max_row, max_col = _trim_bounds_ws(ws_src)
        if max_row <= 0 or max_col <= 0:
            continue
        for row in ws_src.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True):
            ws_dst.append(list(row))
    try:
        dst.save(target_path)
    finally:
        try:
            dst.close()
        except Exception:
            pass


def _capture_external_link_parts(xlsx_path: str):
    """Capture fragile package parts to preserve them across saves."""
    if not xlsx_path or (not os.path.isfile(xlsx_path)):
        return None
    try:
        with zipfile.ZipFile(xlsx_path, "r") as zf:
            names = zf.namelist()
            target_names = set(
                n for n in names
                if n.startswith("xl/externalLinks/")
            )
            # workbook rels may carry externalLink relationship targets
            if "xl/_rels/workbook.xml.rels" in names:
                target_names.add("xl/_rels/workbook.xml.rels")
            # Preserve workbook/content-types metadata as-is to avoid Excel repair
            # on heavily customized workbooks with external links.
            if "xl/workbook.xml" in names:
                target_names.add("xl/workbook.xml")
            if "[Content_Types].xml" in names:
                target_names.add("[Content_Types].xml")
            if not target_names:
                return None
            parts = {n: zf.read(n) for n in target_names}
            _dlog(f"capture fragile parts: path={xlsx_path} count={len(parts)}")
            return parts
    except Exception:
        return None


def _restore_external_link_parts(xlsx_path: str, parts) -> int:
    """Restore captured fragile package parts into saved workbook."""
    if not parts:
        return 0
    if not xlsx_path or (not os.path.isfile(xlsx_path)):
        return 0
    try:
        with zipfile.ZipFile(xlsx_path, "r") as zf:
            names = zf.namelist()
            original_infos = {n: zf.getinfo(n) for n in names}
            all_bytes = {n: zf.read(n) for n in names}
    except Exception:
        return 0

    # Merge: keep all saved parts, overwrite/add captured external-link parts.
    for n, b in parts.items():
        all_bytes[n] = b

    changed = len(parts)

    tmp = f"{xlsx_path}.extlinkrestore.{os.getpid()}.tmp"
    try:
        with zipfile.ZipFile(tmp, "w") as zf:
            for name in all_bytes.keys():
                info = original_infos.get(name)
                comp = info.compress_type if info else zipfile.ZIP_DEFLATED
                zf.writestr(name, all_bytes[name], compress_type=comp)
        os.replace(tmp, xlsx_path)
        _dlog(f"fragile parts restored: {xlsx_path} changed={changed}")
        return changed
    except Exception as e:
        _dlog(f"fragile parts restore failed: {xlsx_path} err={e}")
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except Exception:
            pass
        return 0


def _apply_external_link_parts_on_file(path: str, external_link_parts=None):
    try:
        if path:
            _restore_external_link_parts(path, external_link_parts)
    except Exception as e:
        _dlog(f"apply fragile parts failed: path={path} err={e}")


def _cell_display_and_equal(ws_a_val, ws_b_val, ws_a_edit, ws_b_edit, r: int, c: int):
    va_val = ws_a_val.cell(row=r, column=c).value
    vb_val = ws_b_val.cell(row=r, column=c).value

    if _USE_CACHED_VALUES_ONLY:
        # If cache missing but edit has a literal value, use it for display/compare.
        if ws_a_edit is not None and ws_b_edit is not None:
            try:
                if va_val is None:
                    va_edit = ws_a_edit.cell(row=r, column=c).value
                    if va_edit is not None and not _formula_text(va_edit):
                        va_val = va_edit
                if vb_val is None:
                    vb_edit = ws_b_edit.cell(row=r, column=c).value
                    if vb_edit is not None and not _formula_text(vb_edit):
                        vb_val = vb_edit
            except Exception:
                pass
        # If cache missing on one side but both formulas are the same, treat as equal and display the available value.
        if ws_a_edit is not None and ws_b_edit is not None and ((va_val is None) != (vb_val is None)):
            va_edit = ws_a_edit.cell(row=r, column=c).value
            vb_edit = ws_b_edit.cell(row=r, column=c).value
            if _same_formula(va_edit, vb_edit):
                v = va_val if va_val is not None else vb_val
                return v, v, True
        # Compare with numeric/string normalization to avoid false diffs
        eq = (_merge_cmp_value(va_val) == _merge_cmp_value(vb_val))
        return va_val, vb_val, eq

    eq = (_merge_cmp_value(va_val) == _merge_cmp_value(vb_val))
    return va_val, vb_val, eq


def _cell_display_and_equal_by_row(ws_a_val, ws_b_val, ws_a_edit, ws_b_edit, ra: int | None, rb: int | None, c: int):
    va_val = ws_a_val.cell(row=ra, column=c).value if ra is not None else None
    vb_val = ws_b_val.cell(row=rb, column=c).value if rb is not None else None

    if _USE_CACHED_VALUES_ONLY:
        if ws_a_edit is not None and ws_b_edit is not None:
            try:
                if va_val is None and ra is not None:
                    va_edit = ws_a_edit.cell(row=ra, column=c).value
                    if va_edit is not None and not _formula_text(va_edit):
                        va_val = va_edit
                if vb_val is None and rb is not None:
                    vb_edit = ws_b_edit.cell(row=rb, column=c).value
                    if vb_edit is not None and not _formula_text(vb_edit):
                        vb_val = vb_edit
            except Exception:
                pass
            if (va_val is None) != (vb_val is None):
                try:
                    va_edit = ws_a_edit.cell(row=ra, column=c).value if ra is not None else None
                    vb_edit = ws_b_edit.cell(row=rb, column=c).value if rb is not None else None
                    if _same_formula(va_edit, vb_edit):
                        v = va_val if va_val is not None else vb_val
                        return v, v, True
                except Exception:
                    pass
        eq = (_merge_cmp_value(va_val) == _merge_cmp_value(vb_val))
        return va_val, vb_val, eq

    eq = (_merge_cmp_value(va_val) == _merge_cmp_value(vb_val))
    return va_val, vb_val, eq


def _formula_text(v):
    if isinstance(v, str) and v.startswith("="):
        return v
    if isinstance(v, ArrayFormula):
        return getattr(v, "text", None)
    t = getattr(v, "text", None)
    if isinstance(t, str) and t.startswith("="):
        return t
    return None


def _norm_formula_text(v):
    f = _formula_text(v)
    if not f:
        return None
    s = str(f).strip()
    if s.startswith("="):
        s = s[1:]
    # normalize cosmetic differences
    s = re.sub(r"\s+", "", s).upper()
    return s


def _same_formula(a, b):
    na = _norm_formula_text(a)
    nb = _norm_formula_text(b)
    return bool(na and nb and na == nb)


def _choose_edit_value(v_val, v_edit):
    """Preserve formula semantics on overwrite, even in cached-value mode."""
    f = _formula_text(v_edit)
    if f:
        return f
    return v_val if _USE_CACHED_VALUES_ONLY else v_edit


def _col_index_from_ref(ref: str) -> int:
    if not ref:
        return 0
    m = re.match(r"([A-Za-z]+)", str(ref))
    if not m:
        return 0
    letters = m.group(1).upper()
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _extract_ns_map(xml_bytes: bytes) -> dict:
    """Extract XML namespace prefix→URI map from the opening bytes of an XML document.
    Used to register namespaces with ET before re-serialization to preserve original prefixes."""
    text = xml_bytes[:4096].decode("utf-8", errors="ignore")
    ns_map = {}
    for m in re.finditer(r'xmlns(?::(\w+))?="([^"]+)"', text):
        prefix = m.group(1) if m.group(1) is not None else ""
        uri = m.group(2)
        ns_map[prefix] = uri
    return ns_map


def _sheet_xml_set_cell(ws_root, row_idx: int, col_idx: int, value):
    ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    q = lambda t: f"{{{ns}}}{t}"
    xml_space = "{http://www.w3.org/XML/1998/namespace}space"

    sheet_data = ws_root.find(q("sheetData"))
    if sheet_data is None:
        sheet_data = ET.SubElement(ws_root, q("sheetData"))

    target_row = None
    insert_row_at = None
    rows = list(sheet_data.findall(q("row")))
    for i, r in enumerate(rows):
        try:
            rr = int(r.attrib.get("r", "0") or 0)
        except (ValueError, TypeError):
            rr = 0
        if rr == row_idx:
            target_row = r
            break
        if rr > row_idx and insert_row_at is None:
            insert_row_at = i
    if target_row is None:
        target_row = ET.Element(q("row"), {"r": str(row_idx)})
        if insert_row_at is None:
            sheet_data.append(target_row)
        else:
            sheet_data.insert(insert_row_at, target_row)

    cell_ref = f"{get_column_letter(col_idx)}{row_idx}"
    target_cell = None
    insert_cell_at = None
    cells = list(target_row.findall(q("c")))
    for i, c in enumerate(cells):
        ref = c.attrib.get("r", "")
        cc = _col_index_from_ref(ref)
        if ref == cell_ref:
            target_cell = c
            break
        if cc > col_idx and insert_cell_at is None:
            insert_cell_at = i
    if target_cell is None:
        target_cell = ET.Element(q("c"), {"r": cell_ref})
        if insert_cell_at is None:
            target_row.append(target_cell)
        else:
            target_row.insert(insert_cell_at, target_cell)

    # Keep row/cell attrs (like style index), replace only value payload.
    for child in list(target_cell):
        target_cell.remove(child)

    formula = _formula_text(value)
    if formula:
        target_cell.attrib.pop("t", None)
        f = ET.SubElement(target_cell, q("f"))
        f.text = formula[1:] if str(formula).startswith("=") else str(formula)
        return

    if value is None:
        target_cell.attrib.pop("t", None)
        return

    if isinstance(value, bool):
        target_cell.attrib["t"] = "b"
        v = ET.SubElement(target_cell, q("v"))
        v.text = "1" if value else "0"
        return

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        target_cell.attrib.pop("t", None)
        v = ET.SubElement(target_cell, q("v"))
        v.text = str(value)
        return

    # Keep text independent from sharedStrings to avoid global reindex churn.
    s = str(value)
    target_cell.attrib["t"] = "inlineStr"
    is_node = ET.SubElement(target_cell, q("is"))
    t_node = ET.SubElement(is_node, q("t"))
    if s and (s[0].isspace() or s[-1].isspace()):
        t_node.attrib[xml_space] = "preserve"
    t_node.text = s


def _build_manual_merge_xlsx_via_zip(src_xlsx: str, out_xlsx: str, manual_ops: dict):
    """Patch only selected cells at XML level; keep untouched parts byte-identical."""
    with zipfile.ZipFile(src_xlsx, "r") as zf:
        names = zf.namelist()
        infos = {n: zf.getinfo(n) for n in names}
        files = {n: zf.read(n) for n in names}

    wb_name = "xl/workbook.xml"
    rels_name = "xl/_rels/workbook.xml.rels"
    if wb_name not in files or rels_name not in files:
        shutil.copy2(src_xlsx, out_xlsx)
        return

    wb_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    doc_rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

    wb_root = ET.fromstring(files[wb_name])
    rel_root = ET.fromstring(files[rels_name])
    rid_to_target = {}
    for rel in rel_root.findall(f"{{{rel_ns}}}Relationship"):
        rid = rel.attrib.get("Id")
        tgt = rel.attrib.get("Target")
        if not rid or not tgt:
            continue
        if tgt.startswith("/"):
            norm = tgt.lstrip("/")
        elif tgt.startswith("xl/"):
            norm = tgt
        else:
            norm = f"xl/{tgt}"
        rid_to_target[rid] = norm.replace("\\", "/")

    sheet_to_part = {}
    sheets = wb_root.find(f"{{{wb_ns}}}sheets")
    if sheets is not None:
        for sh in sheets.findall(f"{{{wb_ns}}}sheet"):
            name = sh.attrib.get("name")
            rid = sh.attrib.get(f"{{{doc_rel_ns}}}id")
            if not name or not rid:
                continue
            part = rid_to_target.get(rid)
            if part and part in files:
                sheet_to_part[name] = part

    ops_by_sheet = {}
    for (sheet, r, c), v in manual_ops.items():
        ops_by_sheet.setdefault(sheet, []).append((int(r), int(c), v))

    for sheet, ops in ops_by_sheet.items():
        part = sheet_to_part.get(sheet)
        if not part:
            continue
        # Register original namespace prefixes before parsing to preserve them in output
        ns_map = _extract_ns_map(files[part])
        for pfx, uri in ns_map.items():
            ET.register_namespace(pfx, uri)
        ws_root = ET.fromstring(files[part])
        for r, c, v in sorted(ops, key=lambda x: (x[0], x[1])):
            _sheet_xml_set_cell(ws_root, r, c, v)
        xml_bytes = ET.tostring(ws_root, encoding="utf-8", xml_declaration=True)
        xml_bytes = re.sub(
            rb'<\?xml[^?]*\?>',
            b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
            xml_bytes, count=1
        )
        # ET silently drops namespace declarations for prefixes not used in any element
        # tag (e.g. x14ac/xr/xr2 referenced only as string tokens in mc:Ignorable).
        # Re-inject any such declarations so mc:Ignorable remains valid.
        missing_decls = [
            f'xmlns:{pfx}="{uri}"'
            for pfx, uri in ns_map.items()
            if pfx and pfx != "xml" and f'xmlns:{pfx}="'.encode() not in xml_bytes
        ]
        if missing_decls:
            inject = (" " + " ".join(missing_decls)).encode()
            xml_bytes = re.sub(
                rb'(<\?xml[^?]*\?>\s*)(<[A-Za-z][\w:.-]*)',
                lambda m: m.group(1) + m.group(2) + inject,
                xml_bytes, count=1
            )
        files[part] = xml_bytes

    # Remove calcChain.xml to avoid Excel repair caused by stale calculation chain
    calc_chain = "xl/calcChain.xml"
    if calc_chain in names:
        names = [n for n in names if n != calc_chain]
        files.pop(calc_chain, None)

        # Remove calcChain Override from [Content_Types].xml
        ct_name = "[Content_Types].xml"
        if ct_name in files:
            ct_ns_map = _extract_ns_map(files[ct_name])
            for pfx, uri in ct_ns_map.items():
                ET.register_namespace(pfx, uri)
            ct_root = ET.fromstring(files[ct_name])
            ct_ns = "http://schemas.openxmlformats.org/package/2006/content-types"
            for override in ct_root.findall(f"{{{ct_ns}}}Override"):
                if "/calcChain" in override.attrib.get("PartName", ""):
                    ct_root.remove(override)
            ct_bytes = ET.tostring(ct_root, encoding="utf-8", xml_declaration=True)
            ct_bytes = re.sub(
                rb'<\?xml[^?]*\?>',
                b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
                ct_bytes, count=1
            )
            files[ct_name] = ct_bytes

        # Remove calcChain Relationship from xl/_rels/workbook.xml.rels
        if rels_name in files:
            wbr_ns_map = _extract_ns_map(files[rels_name])
            for pfx, uri in wbr_ns_map.items():
                ET.register_namespace(pfx, uri)
            wbr_root = ET.fromstring(files[rels_name])
            for rel in wbr_root.findall(f"{{{rel_ns}}}Relationship"):
                if "calcChain" in rel.attrib.get("Target", ""):
                    wbr_root.remove(rel)
            wbr_bytes = ET.tostring(wbr_root, encoding="utf-8", xml_declaration=True)
            wbr_bytes = re.sub(
                rb'<\?xml[^?]*\?>',
                b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
                wbr_bytes, count=1
            )
            files[rels_name] = wbr_bytes

    with zipfile.ZipFile(out_xlsx, "w") as zf:
        for n in names:
            info = infos.get(n)
            comp = info.compress_type if info else zipfile.ZIP_DEFLATED
            zf.writestr(n, files[n], compress_type=comp)


def _build_manual_merge_output_with_excel(src_xlsx: str, out_xlsx: str, manual_ops: dict) -> bool:
    """Apply manual ops through Excel COM to preserve workbook fidelity."""
    if not manual_ops:
        try:
            shutil.copy2(src_xlsx, out_xlsx)
            return True
        except Exception:
            return False
    ops = []
    for (sheet, r, c), v in manual_ops.items():
        fv = _formula_text(v)
        if fv:
            ops.append({"sheet": sheet, "r": int(r), "c": int(c), "formula": fv})
        else:
            ops.append({"sheet": sheet, "r": int(r), "c": int(c), "value": v})
    try:
        ops_json = os.path.join(tempfile.gettempdir(), f"{APP_NAME}_merge_ops_{os.getpid()}_{int(time.time())}.json")
        with open(ops_json, "w", encoding="utf-8") as f:
            json.dump(ops, f, ensure_ascii=False)
    except Exception as e:
        _dlog(f"excel native save: write ops json failed: {e}")
        return False
    try:
        ps = (
            "$ErrorActionPreference='Stop';"
            "$src='" + src_xlsx.replace("'", "''") + "';"
            "$out='" + out_xlsx.replace("'", "''") + "';"
            "$opsPath='" + ops_json.replace("'", "''") + "';"
            "$ops=Get-Content -Raw -LiteralPath $opsPath | ConvertFrom-Json;"
            "$xl=New-Object -ComObject Excel.Application;"
            "$xl.Visible=$false;"
            "$xl.DisplayAlerts=$false;"
            "$xl.AskToUpdateLinks=$false;"
            "$xl.EnableEvents=$false;"
            "$wb=$xl.Workbooks.Open($src,0,$false);"
            "foreach($op in $ops){"
            "  $ws=$wb.Worksheets.Item($op.sheet);"
            "  $cell=$ws.Cells.Item([int]$op.r,[int]$op.c);"
            "  if($null -ne $op.formula){$cell.Formula=$op.formula}else{$cell.Value2=$op.value}"
            "};"
            "$wb.SaveCopyAs($out);"
            "$wb.Close($false);"
            "$xl.Quit();"
            "[System.Runtime.Interopservices.Marshal]::ReleaseComObject($wb) | Out-Null;"
            "[System.Runtime.Interopservices.Marshal]::ReleaseComObject($xl) | Out-Null;"
        )
        r = subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps],
            capture_output=True,
            text=True,
            timeout=180,
        )
        if r.returncode != 0:
            _dlog(f"excel native save failed rc={r.returncode} err={r.stderr.strip()}")
            return False
        _dlog(f"excel native save ok: {out_xlsx} ops={len(ops)}")
        return True
    except Exception as e:
        _dlog(f"excel native save exception: {e}")
        return False
    finally:
        try:
            if os.path.exists(ops_json):
                os.remove(ops_json)
        except Exception:
            pass


def _merge_cmp_value(v):
    """Normalize values for merge conflict comparison to match UI display."""
    try:
        if v is None:
            return ""
        s = _val_to_str(v)
        if isinstance(s, str):
            # Normalize line endings and trim trailing whitespace to avoid false conflicts
            s = s.replace("\r\n", "\n").rstrip()
            # Normalize numeric strings
            try:
                num = float(s)
                if num.is_integer():
                    return str(int(num))
                return str(num)
            except Exception:
                return s
        return s
    except Exception:
        return v


def _scan_formula_cache(path: str):
    """Return (has_formula, missing_cache) based on a sample scan."""
    wb_val = None
    wb_edit = None
    try:
        wb_val = load_workbook(path, data_only=True, read_only=True)
        wb_edit = load_workbook(path, data_only=False, read_only=True)
    except Exception as e:
        _dlog(f"cache check open failed: {e}")
        _wbs_close(wb_val, wb_edit)
        return False, True

    checked = 0
    has_formula = False
    missing_cache = False
    try:
        for sheet in wb_edit.sheetnames:
            ws_e = wb_edit[sheet]
            if sheet not in wb_val.sheetnames:
                continue
            ws_v = wb_val[sheet]
            for row in ws_e.iter_rows(values_only=False):
                for cell in row:
                    if checked >= _CACHE_CHECK_MAX_CELLS:
                        return has_formula, missing_cache
                    f = _formula_text(cell.value)
                    if not f:
                        continue
                    has_formula = True
                    checked += 1
                    try:
                        v = ws_v.cell(row=cell.row, column=cell.column).value
                    except Exception:
                        v = None
                    if v is None:
                        missing_cache = True
                        return has_formula, missing_cache
    finally:
        try:
            wb_val.close()
            wb_edit.close()
        except Exception:
            pass
    return has_formula, missing_cache


def _recalc_with_excel(path: str) -> str | None:
    """Use Excel COM to recalc formulas and update cached values in a temp copy."""
    try:
        base = os.path.basename(path)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        tmp = os.path.join(tempfile.gettempdir(), f"{APP_NAME}_recalc_{os.getpid()}_{ts}_{base}")
        shutil.copy2(path, tmp)
    except Exception as e:
        _dlog(f"recalc copy failed: {e}")
        return None

    try:
        ps = (
            "$ErrorActionPreference='Stop';"
            "$p='" + tmp.replace("'", "''") + "';"
            "$xl=New-Object -ComObject Excel.Application;"
            "$xl.Visible=$false;"
            "$xl.DisplayAlerts=$false;"
            "$xl.AskToUpdateLinks=$false;"
            "$xl.EnableEvents=$false;"
            "$wb=$xl.Workbooks.Open($p,$false,$false);"
            "try{$xl.Calculation=-4105}catch{};"
            "try{$xl.CalculateFullRebuild()}catch{};"
            "try{$wb.RefreshAll();$xl.CalculateFullRebuild()}catch{};"
            "$wb.Save();"
            "$wb.Close($true);"
            "$xl.Quit();"
        )
        no_window = getattr(subprocess, "CREATE_NO_WINDOW", 0)
        r = subprocess.run(
            ["powershell", "-NoProfile", "-Command", ps],
            capture_output=True,
            text=True,
            timeout=120,
            creationflags=no_window,
        )
        if r.returncode != 0:
            _dlog(f"excel recalc ps failed: {r.stderr.strip()}")
            return None
        return tmp
    except Exception as e:
        _dlog(f"excel recalc failed: {e}")
        return None


def _prepare_val_path(path: str) -> str:
    if not _USE_CACHED_VALUES_ONLY or not _AUTO_RECALC_MISSING_CACHE:
        return path
    try:
        has_formula, missing_cache = _scan_formula_cache(path)
        if has_formula and (_AUTO_RECALC_FORMULAS_ALWAYS or missing_cache):
            _dlog(f"formula cache recalc: has_formula={has_formula} missing_cache={missing_cache} path={path}")
            tmp = _recalc_with_excel(path)
            if tmp:
                _dlog(f"recalc cache OK: {tmp}")
                return tmp
    except Exception as e:
        _dlog(f"prepare val path failed: {e}")
    return path


def _recalc_and_prepare_val_path(path: str) -> str | None:
    """Force Excel recalc to refresh cached values and return temp path."""
    try:
        tmp = _recalc_with_excel(path)
        return tmp
    except Exception:
        return None


def _launch_deferred_copy(src: str, dst: str, retries: int = 60, delay_ms: int = 500):
    """Launch a background copy that retries for a while (to avoid lock issues)."""
    try:
        ps = (
            "$src='" + src.replace("'", "''") + "';"
            "$dst='" + dst.replace("'", "''") + "';"
            f"for($i=0;$i -lt {retries};$i++){{"
            "try{Copy-Item -LiteralPath $src -Destination $dst -Force;"
            "Remove-Item -LiteralPath $src -Force;exit 0}catch{Start-Sleep -Milliseconds "
            f"{delay_ms}}};exit 1"
        )
        creationflags = 0
        try:
            creationflags = subprocess.CREATE_NO_WINDOW
        except Exception:
            creationflags = 0
        subprocess.Popen(["powershell", "-NoProfile", "-Command", ps], creationflags=creationflags)
    except Exception as e:
        _dlog(f"deferred copy launch failed: {e}")


def _find_tortoise_merge_exe():
    candidates = [
        os.path.join(os.environ.get("ProgramFiles", r"C:\Program Files"), "TortoiseSVN", "bin", "TortoiseMerge.exe"),
        os.path.join(os.environ.get("ProgramFiles(x86)", r"C:\Program Files (x86)"), "TortoiseSVN", "bin", "TortoiseMerge.exe"),
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    return "TortoiseMerge.exe"


def _find_tortoise_proc_exe():
    candidates = [
        os.path.join(os.environ.get("ProgramFiles", r"C:\Program Files"), "TortoiseSVN", "bin", "TortoiseProc.exe"),
        os.path.join(os.environ.get("ProgramFiles(x86)", r"C:\Program Files (x86)"), "TortoiseSVN", "bin", "TortoiseProc.exe"),
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    return "TortoiseProc.exe"


def _find_svn_cli_exe() -> str | None:
    """Resolve svn.exe from PATH first, then common TortoiseSVN install paths."""
    p = shutil.which("svn")
    if p:
        return p
    candidates = [
        os.path.join(os.environ.get("ProgramFiles", r"C:\Program Files"), "TortoiseSVN", "bin", "svn.exe"),
        os.path.join(os.environ.get("ProgramFiles(x86)", r"C:\Program Files (x86)"), "TortoiseSVN", "bin", "svn.exe"),
    ]
    for c in candidates:
        try:
            if os.path.exists(c):
                return c
        except Exception:
            pass
    return None


def _query_svn_version(svn_exe: str) -> str:
    try:
        no_window = getattr(subprocess, "CREATE_NO_WINDOW", 0)
        out = subprocess.run(
            [svn_exe, "--version", "--quiet"],
            capture_output=True,
            text=True,
            timeout=8,
            creationflags=no_window,
        )
        if out.returncode == 0:
            return (out.stdout or "").strip()
    except Exception:
        pass
    return ""


def _try_export_svn_revision_from_merge_temp(path: str) -> str:
    """If path looks like *.merge-left.r#### or *.merge-right.r####, export that revision from WC.

    Returns replacement path if export succeeded; otherwise returns original path.
    """
    try:
        if not path:
            return path
        p = os.path.abspath(path)
        m = re.match(r"^(?P<base>.+)\.merge-(left|right)\.r(?P<rev>\d+)$", p, flags=re.IGNORECASE)
        if not m:
            return path
        base_path = m.group("base")
        rev = m.group("rev")
        if not os.path.exists(base_path):
            # Try same dir + original base filename
            base_path = os.path.join(os.path.dirname(p), os.path.basename(m.group("base")))
        if not os.path.exists(base_path):
            _dlog(f"svn export skip: base not found for {path}")
            return path

        proc_exe = _find_tortoise_proc_exe()
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = os.path.basename(base_path)
        save_path = os.path.join(tempfile.gettempdir(), f"{APP_NAME}_svncat_r{rev}_{ts}_{base_name}")
        if not save_path.lower().endswith(".xlsx"):
            save_path += ".xlsx"

        try:
            _dlog(f"svn export: {base_path} r{rev} -> {save_path}")
        except Exception:
            pass

        # TortoiseProc may show UI; run and wait briefly for file to appear.
        try:
            subprocess.Popen([
                proc_exe,
                "/command:cat",
                f"/path:{base_path}",
                f"/revision:{rev}",
                f"/savepath:{save_path}",
                "/closeonend:1",
            ])
        except Exception as e:
            _dlog(f"svn export failed launch: {e}")
            return path

        # Wait for file to be created (best-effort)
        for _ in range(50):
            try:
                if os.path.exists(save_path) and os.path.getsize(save_path) > 0:
                    return save_path
            except Exception:
                pass
            time.sleep(0.1)

        _dlog(f"svn export timeout: {save_path}")
        return path
    except Exception as e:
        _dlog(f"svn export error: {e}")
        return path



def _try_export_svn_base_from_working_copy(path: str) -> str | None:
    """Export BASE revision for a working-copy file path.

    Returns exported temp .xlsx path when successful, otherwise None.
    """
    try:
        if not path:
            return None
        p = os.path.abspath(path)
        if not os.path.exists(p):
            return None

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = os.path.basename(p)
        save_path = os.path.join(tempfile.gettempdir(), f"{APP_NAME}_svncat_BASE_{ts}_{base_name}")
        if not save_path.lower().endswith(".xlsx"):
            save_path += ".xlsx"

        # Prefer svn CLI (usually uses WC metadata for BASE).
        svn_exe = shutil.which("svn")
        if svn_exe:
            try:
                no_window = getattr(subprocess, "CREATE_NO_WINDOW", 0)
                with open(save_path, "wb") as f:
                    r = subprocess.run(
                        [svn_exe, "cat", "-r", "BASE", p],
                        stdout=f,
                        stderr=subprocess.PIPE,
                        timeout=30,
                        creationflags=no_window,
                    )
                if r.returncode == 0 and os.path.exists(save_path) and os.path.getsize(save_path) > 0:
                    _dlog(f"svn base export(cli): {p} -> {save_path}")
                    return save_path
                try:
                    if os.path.exists(save_path):
                        os.remove(save_path)
                except Exception:
                    pass
                _dlog(
                    f"svn base export(cli) failed: rc={r.returncode} "
                    f"err={(r.stderr or b'').decode('utf-8', errors='ignore')}"
                )
            except Exception as e:
                _dlog(f"svn base export(cli) exception: {e}")

        # Fallback: TortoiseProc cat BASE
        try:
            proc_exe = _find_tortoise_proc_exe()
            subprocess.Popen([
                proc_exe,
                "/command:cat",
                f"/path:{p}",
                "/revision:BASE",
                f"/savepath:{save_path}",
                "/closeonend:1",
            ])
            for _ in range(50):
                if os.path.exists(save_path) and os.path.getsize(save_path) > 0:
                    _dlog(f"svn base export(tortoise): {p} -> {save_path}")
                    return save_path
                time.sleep(0.1)
        except Exception as e:
            _dlog(f"svn base export(tortoise) exception: {e}")
            return None
    except Exception as e:
        _dlog(f"svn base export error: {e}")
        return None
    return None


def _find_handle_exe():
    candidates = [
        os.path.join(os.environ.get("SystemRoot", r"C:\Windows"), "System32", "handle.exe"),
        r"C:\Sysinternals\handle.exe",
        r"C:\Tools\Sysinternals\handle.exe",
        r"D:\Tools\Sysinternals\handle.exe",
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    return shutil.which("handle.exe")


def _log_lock_holders(path: str) -> bool:
    """Return True if Excel is detected holding the file."""
    excel_found = False
    try:
        handle_exe = _find_handle_exe()
        if not handle_exe:
            _dlog(f"lock holders: handle.exe not found for {path}")
            return False
        no_window = getattr(subprocess, "CREATE_NO_WINDOW", 0)
        r = subprocess.run(
            [handle_exe, "-accepteula", path],
            capture_output=True,
            text=True,
            timeout=10,
            creationflags=no_window,
        )
        out = (r.stdout or "") + (r.stderr or "")
        if out.strip():
            for line in out.splitlines():
                if path.lower() in line.lower():
                    _dlog(f"lock holders: {line.strip()}")
                    if "excel.exe" in line.lower():
                        excel_found = True
            return excel_found
        _dlog(f"lock holders: no output for {path}")
    except Exception as e:
        _dlog(f"lock holders: failed {e}")
    return excel_found


def _try_svn_resolve(path: str) -> bool:
    """Attempt to mark conflict as resolved in SVN."""
    try:
        svn_exe = shutil.which("svn")
        if svn_exe:
            subprocess.run(
                [svn_exe, "resolve", "--accept", "working", path],
                check=False,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
            )
            return True
    except Exception:
        pass
    # Fallback to TortoiseProc (may show UI)
    try:
        proc_exe = _find_tortoise_proc_exe()
        subprocess.Popen([proc_exe, "/command:resolve", f"/path:{path}", "/closeonend:1"])
        return True
    except Exception:
        return False


def open_tortoise_merge(left_txt: str, right_txt: str, title: str):
    exe = _find_tortoise_merge_exe()
    args = [exe, "/base", left_txt, "/mine", right_txt, "/title", title]
    subprocess.Popen(args)


def _show_conflict_popup(conflicts):
    try:
        root = tk.Tk()
        root.withdraw()
        win = tk.Toplevel(root)
        win.title("发现冲突")
        win.resizable(False, False)
        win.geometry("+{}+{}".format(root.winfo_screenwidth() // 2 - 220, root.winfo_screenheight() // 2 - 180))
        msg = "与其他同学冲突，请联系确认后再修改保存！！！"
        lbl = tk.Label(win, text=msg, fg="red", font=("Microsoft YaHei", 12, "bold"), padx=16, pady=10)
        lbl.pack()

        detail_lines = []
        for sheet, r, c, _vm, _vt in conflicts[:3]:
            col = get_column_letter(c)
            detail_lines.append(f"{sheet}!{col}{r}")
        if len(conflicts) > 3:
            detail_lines.append("...")
        detail_text = "\n".join(detail_lines) if detail_lines else "（无）"
        txt = tk.Text(win, height=12, width=60)
        txt.insert("1.0", detail_text)
        txt.configure(state="disabled")
        txt.pack(padx=12, pady=(0, 10))

        tk.Button(win, text="确定", command=win.destroy).pack(pady=(0, 10))
        win.grab_set()
        win.wait_window()
        root.destroy()
    except Exception:
        pass


def excel_to_text(path: str, out_path: str, thick_sep_char: str = "="):
    val_path = _prepare_val_path(path)
    wb = load_workbook(val_path, data_only=True)
    try:
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(f"{APP_NAME} text export\n")
            f.write(f"Source: {path}\n")
            f.write(f"Time: {datetime.now().isoformat(sep=' ', timespec='seconds')}\n\n")

            for idx, name in enumerate(wb.sheetnames):
                ws = wb[name]
                max_row = ws.max_row or 1
                max_col = ws.max_column or 1

                if idx != 0:
                    f.write("\n" + (thick_sep_char * 120) + "\n")

                title = f"SHEET: {name}"
                pad = max(0, 120 - len(title) - 2)
                left = thick_sep_char * (pad // 2)
                right = thick_sep_char * (pad - (pad // 2))
                f.write(f"{left} {title} {right}\n")

                cols = [ws.cell(row=1, column=c).coordinate[:-1] for c in range(1, max_col + 1)]
                f.write("ROW\t" + "\t".join(cols) + "\n")

                for r in range(1, max_row + 1):
                    vals = []
                    for c in range(1, max_col + 1):
                        vals.append(_val_to_str(ws.cell(row=r, column=c).value))
                    f.write(str(r) + "\t" + "\t".join(vals) + "\n")
    finally:
        wb.close()
        if val_path != path:
            try:
                os.remove(val_path)
            except Exception:
                pass


def pick_two_files_same_name():
    root = tk.Tk()
    root.withdraw()

    a = filedialog.askopenfilename(title="Select first .xlsx file", filetypes=[("Excel Workbook", "*.xlsx")])
    if not a:
        return None, None
    b = filedialog.askopenfilename(title="Select second .xlsx file (same filename)", filetypes=[("Excel Workbook", "*.xlsx")])
    if not b:
        return None, None

    if os.path.basename(a).lower() != os.path.basename(b).lower():
        messagebox.showerror(
            "Filename mismatch",
            f"The two files must have the same filename.\n\nA: {os.path.basename(a)}\nB: {os.path.basename(b)}",
        )
        return None, None

    return a, b


def _detect_svn_conflict_files(target_path: str):
    # If user selected a conflict artifact directly, map back to merged target first.
    try:
        p = os.path.abspath(target_path)
        m = re.match(r"^(?P<base>.+)\.merge-(left|right)\.r\d+$", p, flags=re.IGNORECASE)
        if m:
            target_path = m.group("base")
    except Exception:
        pass
    folder = os.path.dirname(target_path)
    base_name = os.path.basename(target_path)
    # SVN conflict artifacts:
    # - file.merge-left.r#### / file.merge-right.r#### (newer SVN)
    # - file.r<rev> (older SVN), possibly .mine
    merge_left = []
    merge_right = []
    for name in os.listdir(folder):
        if name.startswith(base_name + ".merge-left.r"):
            suffix = name[len(base_name) + len(".merge-left.r"):]
            if suffix.isdigit():
                merge_left.append((int(suffix), os.path.join(folder, name)))
        elif name.startswith(base_name + ".merge-right.r"):
            suffix = name[len(base_name) + len(".merge-right.r"):]
            if suffix.isdigit():
                merge_right.append((int(suffix), os.path.join(folder, name)))
        elif name == base_name + ".merge-left":
            merge_left.append((0, os.path.join(folder, name)))
        elif name == base_name + ".merge-right":
            merge_right.append((0, os.path.join(folder, name)))
    if merge_left and merge_right:
        merge_left.sort(key=lambda x: x[0])
        merge_right.sort(key=lambda x: x[0])
        base_path = merge_left[-1][1]
        theirs_path = merge_right[-1][1]
        mine_path = target_path
        merged_path = target_path
        return base_path, mine_path, theirs_path, merged_path

    # Older SVN conflict artifacts: file.r<rev> (numeric), possibly .mine
    r_files = []
    for name in os.listdir(folder):
        if not name.startswith(base_name + ".r"):
            continue
        suffix = name[len(base_name) + 2:]
        if suffix.isdigit():
            r_files.append((int(suffix), os.path.join(folder, name)))
    if len(r_files) >= 2:
        r_files.sort(key=lambda x: x[0])
        base_path = r_files[0][1]
        theirs_path = r_files[-1][1]
        mine_path = target_path
        merged_path = target_path
        return base_path, mine_path, theirs_path, merged_path
    # Fallback for rOLD/rNEW naming
    r_old = os.path.join(folder, base_name + ".rOLD")
    r_new = os.path.join(folder, base_name + ".rNEW")
    if os.path.exists(r_old) and os.path.exists(r_new):
        return r_old, target_path, r_new, target_path
    # Fallback: fuzzy match for temp-stable names that still contain "<base>.merge-left/right.r####"
    # e.g. sow_merge_tool_stable_..._<base>.merge-right.r27548_...
    try:
        merge_left_fuzzy = []
        merge_right_fuzzy = []
        key = (base_name + ".merge-").lower()
        for name in os.listdir(folder):
            low = name.lower()
            if key not in low:
                continue
            i_left = low.find((base_name + ".merge-left.r").lower())
            if i_left >= 0:
                j = i_left + len((base_name + ".merge-left.r").lower())
                rev = []
                while j < len(low) and low[j].isdigit():
                    rev.append(low[j])
                    j += 1
                if rev:
                    merge_left_fuzzy.append((int("".join(rev)), os.path.join(folder, name)))
            i_right = low.find((base_name + ".merge-right.r").lower())
            if i_right >= 0:
                j = i_right + len((base_name + ".merge-right.r").lower())
                rev = []
                while j < len(low) and low[j].isdigit():
                    rev.append(low[j])
                    j += 1
                if rev:
                    merge_right_fuzzy.append((int("".join(rev)), os.path.join(folder, name)))
        if merge_left_fuzzy and merge_right_fuzzy:
            merge_left_fuzzy.sort(key=lambda x: x[0])
            merge_right_fuzzy.sort(key=lambda x: x[0])
            return merge_left_fuzzy[-1][1], target_path, merge_right_fuzzy[-1][1], target_path
    except Exception:
        pass
    return None


def _trace_launch(msg: str):
    try:
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
        with open(_LAUNCH_TRACE_PATH, "a", encoding="utf-8") as f:
            f.write(f"[{ts}] {msg}\n")
    except Exception:
        pass


def _has_svn_conflict_artifacts(target_path: str) -> bool:
    try:
        folder = os.path.dirname(target_path)
        base_name = os.path.basename(target_path)
        mine = os.path.join(folder, base_name + ".mine")
        if os.path.exists(mine):
            return True
        for name in os.listdir(folder):
            if name.startswith(base_name + ".merge-left") or name.startswith(base_name + ".merge-right"):
                return True
        r_old = os.path.join(folder, base_name + ".rOLD")
        r_new = os.path.join(folder, base_name + ".rNEW")
        if os.path.exists(r_old) or os.path.exists(r_new):
            return True
        for name in os.listdir(folder):
            if name.startswith(base_name + ".r"):
                suffix = name[len(base_name) + 2:]
                if suffix.isdigit():
                    return True
    except Exception:
        pass
    return False


def _find_conflict_in_dir(folder: str):
    try:
        # If there is exactly one conflicted file in folder, return it.
        base_names = set()
        for name in os.listdir(folder):
            if ".merge-left" in name:
                base = name.split(".merge-left")[0]
                base_names.add(base)
                continue
            if ".merge-right" in name:
                base = name.split(".merge-right")[0]
                base_names.add(base)
                continue
            if ".r" in name:
                base = name.split(".r")[0]
                base_names.add(base)
        candidates = []
        for base in base_names:
            target = os.path.join(folder, base)
            if os.path.exists(target) and _has_svn_conflict_artifacts(target):
                candidates.append(target)
        if len(candidates) == 1:
            return candidates[0]
    except Exception:
        pass
    return None


def _auto_pick_conflict_file():
    # Best-effort: try current working directory
    try:
        cwd = os.getcwd()
        p = _find_conflict_in_dir(cwd)
        if p:
            return p
        try:
            svn_exe = shutil.which("svn")
            if svn_exe:
                no_window = getattr(subprocess, "CREATE_NO_WINDOW", 0)
                r = subprocess.run(
                    [svn_exe, "status", cwd],
                    capture_output=True,
                    text=True,
                    timeout=8,
                    creationflags=no_window,
                )
                if r.returncode == 0:
                    conflicted = []
                    for line in (r.stdout or "").splitlines():
                        if not line:
                            continue
                        if line[0] != "C":
                            continue
                        rel = line[8:].strip() if len(line) > 8 else ""
                        if not rel:
                            continue
                        cand = os.path.abspath(os.path.join(cwd, rel))
                        if os.path.exists(cand):
                            conflicted.append(cand)
                    if len(conflicted) == 1:
                        return conflicted[0]
        except Exception:
            pass
        # Walk up to find SVN working copy root (.svn)
        cur = cwd
        wc_root = None
        while True:
            if os.path.isdir(os.path.join(cur, ".svn")):
                wc_root = cur
                break
            parent = os.path.dirname(cur)
            if parent == cur:
                break
            cur = parent
        if wc_root:
            # If exactly one conflicted file exists in the working copy, auto-pick it
            candidates = []
            for root, _dirs, files in os.walk(wc_root):
                base_names = set()
                for name in files:
                    if ".r" in name:
                        base = name.split(".r")[0]
                        base_names.add(base)
                for base in base_names:
                    target = os.path.join(root, base)
                    if os.path.exists(target) and _has_svn_conflict_artifacts(target):
                        candidates.append(target)
                        if len(candidates) > 1:
                            return None
            if len(candidates) == 1:
                return candidates[0]
    except Exception:
        pass
    return None


def pick_files_or_conflict():
    root = tk.Tk()
    root.withdraw()

    auto = _auto_pick_conflict_file()
    if auto:
        conflict = _detect_svn_conflict_files(auto)
        if conflict:
            return ("merge",) + conflict + (True,)

    a = filedialog.askopenfilename(title="Select .xlsx file", filetypes=[("Excel Workbook", "*.xlsx")])
    if not a:
        return None

    conflict = _detect_svn_conflict_files(a)
    if conflict:
        return ("merge",) + conflict + (True,)

    b = filedialog.askopenfilename(title="Select second .xlsx file (same filename)", filetypes=[("Excel Workbook", "*.xlsx")])
    if not b:
        return None

    if os.path.basename(a).lower() != os.path.basename(b).lower():
        messagebox.showerror(
            "Filename mismatch",
            f"The two files must have the same filename.\n\nA: {os.path.basename(a)}\nB: {os.path.basename(b)}",
        )
        return None

    return ("diff", a, b)


def _atomic_save_wb(wb, target_path: str):
    """Safely overwrite a workbook."""
    folder = os.path.dirname(target_path)
    if folder:
        os.makedirs(folder, exist_ok=True)
    base = os.path.basename(target_path)
    tmp_path = os.path.join(folder, f"~{base}.{os.getpid()}.tmp")
    try:
        if _FAST_SAVE_VALUES_ONLY and _USE_CACHED_VALUES_ONLY:
            _save_values_only_from_wb(wb, tmp_path)
        else:
            wb.save(tmp_path)
        os.replace(tmp_path, target_path)
    except Exception:
        try:
            os.remove(tmp_path)
        except Exception:
            pass
        raise


def _ensure_xlsx_copy(path: str) -> str:
    """If path is not .xlsx, copy to a temp .xlsx and return new path."""
    if not path:
        return path
    if os.path.splitext(path)[1].lower() == ".xlsx":
        return _ensure_stable_copy(path)
    try:
        base = os.path.basename(path)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        tmp = os.path.join(tempfile.gettempdir(), f"{APP_NAME}_svn_{base}_{ts}.xlsx")
        shutil.copy2(path, tmp)
        return tmp
    except Exception:
        return path


def _ensure_stable_copy(path: str) -> str:
    """If path looks like a temp/svn artifact, copy to a stable temp file."""
    if not path:
        return path
    try:
        temp_root = os.path.abspath(tempfile.gettempdir()).lower()
        p_abs = os.path.abspath(path)
        p_low = p_abs.lower()
        base = os.path.basename(path)
        looks_temp = p_low.startswith(temp_root)
        looks_svn = ".svn" in p_low or ".r" in base or "revbase" in base.lower() or "rev" in base.lower()
        if looks_temp or looks_svn:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            tmp = os.path.join(tempfile.gettempdir(), f"{APP_NAME}_stable_{ts}_{base}")
            if not tmp.lower().endswith(".xlsx"):
                tmp += ".xlsx"
            shutil.copy2(path, tmp)
            return tmp
    except Exception:
        pass
    return path


def _is_temp_base_path(path: str) -> bool:
    if not path:
        return False
    p = os.path.abspath(path).lower()
    base = os.path.basename(path).lower()
    if p.startswith(os.path.abspath(tempfile.gettempdir()).lower()):
        return True
    if "revbase" in base or ".svn" in p or base.endswith(".tmp.xlsx") or ".r" in base:
        return True
    return False


def _wbs_close(*wbs):
    """Safely close one or more openpyxl workbooks, ignoring errors."""
    for wb in wbs:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def _merge_three_way(base_path: str, mine_path: str, theirs_path: str, merged_path: str, save_merged: bool = True):
    """3-way merge: apply theirs onto mine when no conflict.

    Conflict: mine and theirs both changed a cell differently vs base.
    Returns (conflicts, merged_preview_path, conflict_cells_by_sheet).
    """
    # XLSM not supported in current scope
    for p in (base_path, mine_path, theirs_path, merged_path):
        if p and os.path.splitext(p)[1].lower() == ".xlsm":
            raise ValueError("当前版本暂不支持 .xlsm 文件合并")

    base_path = _ensure_xlsx_copy(base_path)
    mine_path = _ensure_xlsx_copy(mine_path)
    theirs_path = _ensure_xlsx_copy(theirs_path)

    base_val_path = _prepare_val_path(base_path)
    mine_val_path = _prepare_val_path(mine_path)
    theirs_val_path = _prepare_val_path(theirs_path)

    wb_base_val = load_workbook(base_val_path, data_only=True)
    wb_mine_val = load_workbook(mine_val_path, data_only=True)
    wb_theirs_val = load_workbook(theirs_val_path, data_only=True)
    wb_mine = load_workbook(mine_path, data_only=False)
    wb_base_edit = load_workbook(base_path, data_only=False)
    wb_theirs_edit = load_workbook(theirs_path, data_only=False)

    # Start merged as mine (copy in-memory by reusing workbook; then save to merged_path)
    wb_merged = wb_mine

    set_base = set(wb_base_val.sheetnames)
    set_mine = set(wb_mine_val.sheetnames)
    set_theirs = set(wb_theirs_val.sheetnames)

    # If a sheet exists only in theirs, copy into merged preserving formulas.
    only_theirs = sorted(set_theirs - set_mine)
    for name in only_theirs:
        ws_t = wb_theirs_edit[name]  # data_only=False preserves formulas
        ws_m = wb_merged.create_sheet(title=name)
        max_row = ws_t.max_row or 1
        max_col = ws_t.max_column or 1
        for r in range(1, max_row + 1):
            row = next(ws_t.iter_rows(min_row=r, max_row=r, min_col=1, max_col=max_col, values_only=True), ())
            if len(row) < max_col:
                row = tuple(row) + (None,) * (max_col - len(row))
            for c, v in enumerate(row, start=1):
                ws_m.cell(row=r, column=c).value = v

    conflicts = []
    conflict_cells_by_sheet = {}

    common = sorted(set_mine & set_theirs)
    for name in common:
        ws_b = wb_base_val[name] if name in set_base else None
        ws_m_val = wb_mine_val[name]
        ws_t = wb_theirs_val[name]
        ws_m_edit = wb_mine[name]
        ws_b_edit = wb_base_edit[name] if name in wb_base_edit.sheetnames else None
        ws_t_edit = wb_theirs_edit[name] if name in wb_theirs_edit.sheetnames else None

        max_row = max(ws_m_val.max_row or 1, ws_t.max_row or 1, (ws_b.max_row or 1) if ws_b else 1)
        max_col = max(ws_m_val.max_column or 1, ws_t.max_column or 1, (ws_b.max_column or 1) if ws_b else 1)

        it_m = ws_m_val.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True)
        it_t = ws_t.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True)
        if ws_b:
            it_b = ws_b.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True)
        else:
            it_b = ((None,) * max_col for _ in range(max_row))

        for r, (row_m, row_t, row_b) in enumerate(zip(it_m, it_t, it_b), start=1):
            if len(row_m) < max_col:
                row_m = tuple(row_m) + (None,) * (max_col - len(row_m))
            if len(row_t) < max_col:
                row_t = tuple(row_t) + (None,) * (max_col - len(row_t))
            if len(row_b) < max_col:
                row_b = tuple(row_b) + (None,) * (max_col - len(row_b))

            for c, (vm, vt, vb) in enumerate(zip(row_m, row_t, row_b), start=1):
                # Normalize formula objects to text for comparison
                vm_cmp = vm
                vt_cmp = vt
                vb_cmp = vb
                # If cached value missing but edit cell has a literal (non-formula) value, use it
                try:
                    if vm_cmp is None:
                        vme = ws_m_edit.cell(row=r, column=c).value
                        if vme is not None and not _formula_text(vme):
                            vm_cmp = vme
                    if vt_cmp is None:
                        vte = ws_t_edit.cell(row=r, column=c).value if ws_t_edit is not None else None
                        if vte is not None and not _formula_text(vte):
                            vt_cmp = vte
                    if vb_cmp is None and ws_b_edit is not None:
                        vbe = ws_b_edit.cell(row=r, column=c).value
                        if vbe is not None and not _formula_text(vbe):
                            vb_cmp = vbe
                except Exception:
                    pass
                # If base cache is missing but base/theirs formulas are identical, treat base as theirs
                if vb_cmp is None and vt_cmp is not None and ws_b_edit is not None and ws_t_edit is not None:
                    try:
                        if _same_formula(ws_b_edit.cell(row=r, column=c).value, ws_t_edit.cell(row=r, column=c).value):
                            vb_cmp = vt_cmp
                    except Exception:
                        pass

                vm_key = _merge_cmp_value(vm_cmp)
                vt_key = _merge_cmp_value(vt_cmp)
                vb_key = _merge_cmp_value(vb_cmp)

                mine_changed = (vm_key != vb_key)
                theirs_changed = (vt_key != vb_key)
                if mine_changed and theirs_changed:
                    if vm_key != vt_key:
                        conflicts.append((name, r, c, vm_cmp, vt_cmp))
                        conflict_cells_by_sheet.setdefault(name, {}).setdefault(r, set()).add(c)
                    else:
                        # same change; keep as is
                        continue
                elif (not mine_changed) and theirs_changed:
                    # safe to apply theirs: use edit value to preserve formulas
                    _t_edit_v = ws_t_edit.cell(row=r, column=c).value if ws_t_edit is not None else None
                    ws_m_edit.cell(row=r, column=c).value = _t_edit_v if _t_edit_v is not None else vt

    _merge_result = None
    try:
        # Always save a preview for UI if needed
        if conflicts or (not save_merged):
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            preview = os.path.join(tempfile.gettempdir(), f"{APP_NAME}_merged_preview_{os.getpid()}_{ts}.xlsx")
            _atomic_save_wb(wb_merged, preview)
            _merge_result = (conflicts, preview, conflict_cells_by_sheet)
        else:
            # No conflicts: save directly to merged path
            _atomic_save_wb(wb_merged, merged_path)
            _merge_result = ([], None, {})
    finally:
        _wbs_close(wb_base_val, wb_mine_val, wb_theirs_val, wb_mine, wb_base_edit, wb_theirs_edit)
        for _orig, _vp in ((base_path, base_val_path), (mine_path, mine_val_path), (theirs_path, theirs_val_path)):
            if _vp != _orig:
                try:
                    os.remove(_vp)
                except Exception:
                    pass
    return _merge_result


def _scan_three_way_conflicts(base_path: str, mine_path: str, theirs_path: str):
    """Detect 3-way conflicts only; do NOT auto-apply theirs before UI."""
    base_path = _ensure_xlsx_copy(base_path)
    mine_path = _ensure_xlsx_copy(mine_path)
    theirs_path = _ensure_xlsx_copy(theirs_path)

    base_val_path = _prepare_val_path(base_path)
    mine_val_path = _prepare_val_path(mine_path)
    theirs_val_path = _prepare_val_path(theirs_path)

    wb_base_val = None
    wb_mine_val = None
    wb_theirs_val = None
    wb_mine_edit = None
    wb_base_edit = None
    wb_theirs_edit = None
    try:
        wb_base_val = load_workbook(base_val_path, data_only=True)
        wb_mine_val = load_workbook(mine_val_path, data_only=True)
        wb_theirs_val = load_workbook(theirs_val_path, data_only=True)
        wb_mine_edit = load_workbook(mine_path, data_only=False)
        wb_base_edit = load_workbook(base_path, data_only=False)
        wb_theirs_edit = load_workbook(theirs_path, data_only=False)

        conflicts = []
        conflict_cells_by_sheet = {}

        set_base = set(wb_base_val.sheetnames)
        set_mine = set(wb_mine_val.sheetnames)
        set_theirs = set(wb_theirs_val.sheetnames)
        common = sorted(set_mine & set_theirs)

        for name in common:
            ws_b = wb_base_val[name] if name in set_base else None
            ws_m = wb_mine_val[name]
            ws_t = wb_theirs_val[name]
            ws_m_e = wb_mine_edit[name]
            ws_b_e = wb_base_edit[name] if name in wb_base_edit.sheetnames else None
            ws_t_e = wb_theirs_edit[name] if name in wb_theirs_edit.sheetnames else None

            max_row = max(ws_m.max_row or 1, ws_t.max_row or 1, (ws_b.max_row or 1) if ws_b else 1)
            max_col = max(ws_m.max_column or 1, ws_t.max_column or 1, (ws_b.max_column or 1) if ws_b else 1)

            it_m = ws_m.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True)
            it_t = ws_t.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True)
            it_b = ws_b.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True) if ws_b else ((None,) * max_col for _ in range(max_row))

            for r, (row_m, row_t, row_b) in enumerate(zip(it_m, it_t, it_b), start=1):
                if len(row_m) < max_col:
                    row_m = tuple(row_m) + (None,) * (max_col - len(row_m))
                if len(row_t) < max_col:
                    row_t = tuple(row_t) + (None,) * (max_col - len(row_t))
                if len(row_b) < max_col:
                    row_b = tuple(row_b) + (None,) * (max_col - len(row_b))

                for c, (vm, vt, vb) in enumerate(zip(row_m, row_t, row_b), start=1):
                    vm_cmp = vm
                    vt_cmp = vt
                    vb_cmp = vb
                    try:
                        if vm_cmp is None:
                            vme = ws_m_e.cell(row=r, column=c).value
                            if vme is not None and not _formula_text(vme):
                                vm_cmp = vme
                        if vt_cmp is None and ws_t_e is not None:
                            vte = ws_t_e.cell(row=r, column=c).value
                            if vte is not None and not _formula_text(vte):
                                vt_cmp = vte
                        if vb_cmp is None and ws_b_e is not None:
                            vbe = ws_b_e.cell(row=r, column=c).value
                            if vbe is not None and not _formula_text(vbe):
                                vb_cmp = vbe
                    except Exception:
                        pass
                    if vb_cmp is None and vt_cmp is not None and ws_b_e is not None and ws_t_e is not None:
                        try:
                            if _same_formula(ws_b_e.cell(row=r, column=c).value, ws_t_e.cell(row=r, column=c).value):
                                vb_cmp = vt_cmp
                        except Exception:
                            pass

                    vm_key = _merge_cmp_value(vm_cmp)
                    vt_key = _merge_cmp_value(vt_cmp)
                    vb_key = _merge_cmp_value(vb_cmp)
                    mine_changed = (vm_key != vb_key)
                    theirs_changed = (vt_key != vb_key)
                    if mine_changed and theirs_changed and vm_key != vt_key:
                        conflicts.append((name, r, c, vm_cmp, vt_cmp))
                        conflict_cells_by_sheet.setdefault(name, {}).setdefault(r, set()).add(c)

        return conflicts, conflict_cells_by_sheet
    finally:
        _wbs_close(wb_base_val, wb_mine_val, wb_theirs_val, wb_mine_edit, wb_base_edit, wb_theirs_edit)
        for _orig, _vp in ((base_path, base_val_path), (mine_path, mine_val_path), (theirs_path, theirs_val_path)):
            if _vp != _orig:
                try:
                    os.remove(_vp)
                except Exception:
                    pass


class SheetView:
    """TortoiseMerge-like side-by-side full-sheet viewer.

    Performance notes (optimized for responsiveness):
    - Avoids O(N) tag_remove across the whole document on every click.
    - Avoids per-cell ws.cell access loops during normal interactions.
    - Keeps per-row cached text and per-row diff columns; row merge refreshes only that row.
    """

    def __init__(self, parent, app, sheet_name: str):
        self.parent = parent
        self.app = app
        self.sheet = sheet_name
        # Support lazy tab containers: if parent is already a tab frame, reuse it.
        if isinstance(parent, ttk.Frame) and not parent.winfo_children():
            self.frame = parent
        else:
            self.frame = ttk.Frame(parent)

        self.max_row = 1
        self.max_col = 1
        self._bounds_checked = False
        # Per-column max display width (chars), computed during diff scan
        self.col_char_widths: dict[int, int] = {}
        self._rownum_display_width: int = 3  # right-justified row number gutter width
        self._row_header_width: int = 4

        # Cached row text and diff cols
        self.row_text_a: dict[int, str] = {}
        self.row_text_b: dict[int, str] = {}
        self.diff_cols_by_row: dict[int, set[int]] = {}
        self._display_diff_row_count: int = 0
        self._sample_scan_started = False
        # Row alignment (pair-wise) caches
        self.row_pairs: list[tuple[int | None, int | None]] = []
        self.pair_text_a: dict[int, str] = {}
        self.pair_text_b: dict[int, str] = {}
        self.pair_diff_cols: dict[int, set[int]] = {}
        self.row_a_to_pair_idx: dict[int, int] = {}
        self.row_b_to_pair_idx: dict[int, int] = {}

        # Render state
        # display_rows stores pair indices (into row_pairs)
        self.display_rows: list[int] = []
        self._full_display_rows: list[int] = []
        self._render_limit: int = _FAST_RENDER_ROW_LIMIT
        self.row_to_line: dict[int, int] = {}
        self._pending_yview: float | None = None
        self._render_cache = {}
        self._data_version = 0
        self.selected_excel_row: int | None = None
        self.selected_excel_row_a: int | None = None
        self.selected_excel_row_b: int | None = None
        self.selected_pair_idx: int | None = None
        self._last_selected_line: int | None = None
        self._main_sel_col: int | None = None
        self._main_sel_line: int | None = None
        self._applied_main_sel_col: int | None = None
        self._applied_main_sel_line: int | None = None
        self._cursor_cmp_sel_col: int | None = None
        self._cursor_cmp_sel_line: int | None = None
        self._trace_click_until: float = 0.0
        self._click_trace_seq: int = 0
        self._is_large_sheet = False
        self._prefer_only_diff_when_ready = False
        self._diff_partial = False
        self._align_rows_enabled = True
        self._force_sequence_align = False
        # After user-triggered rescan/toggle, ignore late background cache apply for this sheet
        # to avoid delayed stale overwrite (rows unexpectedly disappear a few seconds later).
        self._suppress_bg_apply = False
        # Set to True once initial diff data has been computed (background or manual).
        # Prevents refresh(rescan=False) from triggering a full rescan on empty initial state.
        self._data_ready = False

        # Rows that were modified via overwrite in this session.
        # In "只看差异" mode, we keep these rows visible even if diffs are resolved.
        self.touched_rows: set[int] = set()

        # Snapshot mode: build the diff row list once, then keep the row list stable.
        # Overwrites only update per-row highlight (to show "已处理") and keep the row visible.
        self.snapshot_only_diff = True

        # Toolbar
        bar = ttk.Frame(self.frame)
        bar.pack(fill="x", padx=8, pady=(8, 6))

        ttk.Label(bar, text=f"Sheet: {sheet_name}", font=("Segoe UI", 11, "bold")).pack(side="left")
        self.info = ttk.Label(bar, text="", foreground="#444")
        self.info.pack(side="left", padx=(10, 0))

        # Diff block navigation (fixed position on the right; does not shift with label lengths)
        self.next_diff_btn = tk.Button(bar, text="下一处差异", padx=10, pady=2, command=self._goto_next_diff_block)
        self.prev_diff_btn = tk.Button(bar, text="上一处差异", padx=10, pady=2, command=self._goto_prev_diff_block)
        # Pack on right so it stays at a stable location across sheets
        self.next_diff_btn.pack(side="right", padx=(6, 0))
        self.prev_diff_btn.pack(side="right", padx=(6, 0))
        self._diff_blocks_cache = None  # None = not yet computed; [] = computed, no diffs

        # Some environments fail to toggle BooleanVar reliably; use IntVar with explicit on/off values.
        self.only_diff_var = tk.IntVar(value=int(getattr(self.app, "only_diff_default", 0)))

        self.only_diff_cb = tk.Checkbutton(
            bar,
            text="只看差异内容",
            variable=self.only_diff_var,
            onvalue=1,
            offvalue=0,
            command=self._toggle_only_diff,
            padx=6,
        )
        # Put on the right for a stable position
        self.only_diff_cb.pack(side="right", padx=(6, 0))
        self.force_align_var = tk.IntVar(value=0)
        self.force_align_cb = tk.Checkbutton(
            bar,
            text="强制行对齐(SM)",
            variable=self.force_align_var,
            onvalue=1,
            offvalue=0,
            command=self._toggle_force_align,
            padx=6,
        )
        self.force_align_cb.pack(side="right", padx=(6, 0))
        self.grid_overlay_var = tk.IntVar(value=1)
        self.grid_overlay_cb = tk.Checkbutton(
            bar,
            text="网格显示",
            variable=self.grid_overlay_var,
            onvalue=1,
            offvalue=0,
            command=self._toggle_grid_overlay,
            padx=6,
        )
        self.grid_overlay_cb.pack(side="right", padx=(6, 0))
        if getattr(self.app, "merge_conflict_mode", False):
            try:
                self.only_diff_var.set(1)
                self.only_diff_cb.select()
                self.only_diff_cb.configure(state="disabled")
                self.force_align_var.set(0)
                self.force_align_cb.configure(state="disabled")
            except Exception:
                pass
        self.three_way_var = tk.IntVar(value=1 if getattr(self.app, "merge_mode", False) and getattr(self.app, "has_base", False) else 0)
        if getattr(self.app, "merge_mode", False) and getattr(self.app, "has_base", False):
            tk.Checkbutton(
                bar,
                text="3视图",
                variable=self.three_way_var,
                onvalue=1,
                offvalue=0,
                command=self._toggle_three_way_view,
                padx=6,
            ).pack(side="right", padx=(6, 0))

        # Apply initial visual state from persisted setting
        try:
            if self.only_diff_var.get():
                self.only_diff_cb.select()
            else:
                self.only_diff_cb.deselect()
        except Exception:
            pass
        self._last_only_diff_value = int(self.only_diff_var.get())

        # Debug: provide a force-toggle button to prove the filtering path works even if UI toggling fails.
        if _DEBUG_ENABLED:
            tk.Button(
                bar,
                text="强制切换",
                command=lambda: (self.only_diff_var.set(0 if self.only_diff_var.get() else 1), self._toggle_only_diff()),
                padx=6,
                pady=1,
            ).pack(side="right", padx=(6, 0))

        # Debug: log click + resulting value
        def _log_cb_click(_evt=None):
            _dlog(f"CHECKBOX_CLICK sheet={self.sheet} var={self.only_diff_var.get()}")
            try:
                self.frame.after_idle(lambda: _dlog(f"CHECKBOX_AFTER_IDLE sheet={self.sheet} var={self.only_diff_var.get()}"))
            except Exception:
                pass

        try:
            self.only_diff_cb.bind("<ButtonRelease-1>", _log_cb_click)
        except Exception:
            pass
        # Context merge controls (row-level + region-level)
        # 区域 = 连续的差异行块；以当前鼠标所在行为锚点定位该块。
        # Office-like split button:
        # - Main button executes current mode action.
        # - Arrow menu switches mode only (no overwrite on switch).
        left_row_dir = "BASE2A" if (getattr(self.app, "merge_mode", False) and getattr(self.app, "has_base", False)) else "A2B"
        right_row_dir = "B2A"
        self._left_copy_direction = left_row_dir
        self._right_copy_direction = right_row_dir
        self._copy_scope_mode = "row"
        self._copy_scope_var = tk.StringVar(value="row")

        def _build_split_group(parent, main_text: str, bg: str, command):
            group = tk.Frame(parent, bg=bg, bd=1, relief="solid")
            main_btn = tk.Button(
                group,
                text=main_text,
                bg=bg,
                activebackground=bg,
                padx=10,
                pady=2,
                relief="flat",
                bd=0,
                command=command,
            )
            sep = tk.Frame(group, bg="#9aa7b0", width=1)
            menu_btn = tk.Menubutton(
                group,
                text="▾",
                bg=bg,
                activebackground=bg,
                padx=6,
                pady=2,
                relief="flat",
                bd=0,
                indicatoron=False,
                direction="below",
            )
            main_btn.pack(side="left")
            sep.pack(side="left", fill="y", pady=2)
            menu_btn.pack(side="left")
            return group, main_btn, menu_btn

        self.use_left_group, self.use_left_btn, self.use_left_menu_btn = _build_split_group(
            bar,
            "使用左侧行",
            "#eaf2ff",
            command=lambda: self._run_copy_action_by_mode(self._left_copy_direction),
        )
        self._use_left_menu = tk.Menu(self.use_left_menu_btn, tearoff=0)
        self._use_left_menu.add_radiobutton(
            label="行模式",
            variable=self._copy_scope_var,
            value="row",
            command=lambda: self._set_copy_scope_mode("row"),
        )
        self._use_left_menu.add_radiobutton(
            label="区域模式",
            variable=self._copy_scope_var,
            value="region",
            command=lambda: self._set_copy_scope_mode("region"),
        )
        self.use_left_menu_btn.configure(menu=self._use_left_menu)

        self.use_right_group, self.use_right_btn, self.use_right_menu_btn = _build_split_group(
            bar,
            "使用右侧行",
            "#ffecec",
            command=lambda: self._run_copy_action_by_mode(self._right_copy_direction),
        )
        self._use_right_menu = tk.Menu(self.use_right_menu_btn, tearoff=0)
        self._use_right_menu.add_radiobutton(
            label="行模式",
            variable=self._copy_scope_var,
            value="row",
            command=lambda: self._set_copy_scope_mode("row"),
        )
        self._use_right_menu.add_radiobutton(
            label="区域模式",
            variable=self._copy_scope_var,
            value="region",
            command=lambda: self._set_copy_scope_mode("region"),
        )
        self.use_right_menu_btn.configure(menu=self._use_right_menu)
        self._set_copy_scope_mode("row")

        self.use_base_btn = None
        if getattr(self.app, "merge_mode", False) and getattr(self.app, "has_base", False):
            self.use_base_btn = tk.Button(
                bar,
                text="保留Mine",
                bg="#f3f3ff",
                padx=10,
                pady=2,
                command=lambda: self._copy_selected_row("MINE2A"),
            )
        self.undo_btn = tk.Button(
            bar,
            text="回退",
            bg="#f2f2f2",
            padx=8,
            pady=2,
            command=self._undo_last_action,
        )
        # Keep at top-right (avoid misclick)
        self.use_right_group.pack(side="right", padx=(6, 0))
        if self.use_base_btn is not None:
            self.use_base_btn.pack(side="right", padx=(6, 0))
        self.use_left_group.pack(side="right")
        self.undo_btn.pack(side="right", padx=(6, 0))

        ttk.Button(bar, text="刷新本Sheet", command=self._manual_rescan).pack(side="right", padx=(6, 0))
        self._full_render = False
        self._load_all_btn = ttk.Button(bar, text="加载全部", command=self._load_all_rows)
        if _FAST_OPEN_ENABLED:
            self._load_all_btn.pack(side="right", padx=(6, 0))

        # Path bar (requested red-box area): show full paths above the diff panes
        path_bar = ttk.Frame(self.frame)
        path_bar.pack(fill="x", padx=8, pady=(0, 4))

        self._path_font = ("Segoe UI", 9)
        path_bar.grid_columnconfigure(0, weight=1)
        path_bar.grid_columnconfigure(1, weight=1)
        path_bar.grid_columnconfigure(2, weight=1)

        self._path_font = ("Segoe UI", 9, "bold")

        def _one_line_text(s: str, max_len: int = 120) -> str:
            s = (s or "").replace("\r", " ").replace("\n", " ")
            if len(s) <= max_len:
                return s
            # keep file tail visible when path is long
            return "..." + s[-(max_len - 3):]

        if getattr(self.app, "merge_mode", False) and getattr(self.app, "has_base", False):
            mine_src = getattr(self.app, "raw_mine", None) or self.app.file_a
            base_src = getattr(self.app, "raw_base", None) or getattr(self.app, "base_path", "")
            theirs_src = getattr(self.app, "raw_theirs", None) or self.app.file_b
            label_a = f"mine={self._source_display_name(mine_src)}"
            label_base = f"base={self._source_display_name(base_src)}" if base_src else "base=-"
            label_b = f"theirs={self._source_display_name(theirs_src)}"
        else:
            # Diff mode: keep wording consistent with SVN semantics (left=base, right=mine).
            base_src = getattr(self.app, "raw_base", None) or self.app.file_a
            mine_src = getattr(self.app, "raw_mine", None) or self.app.file_b
            base_disp = _one_line_text(str(base_src or "")) or "-"
            mine_disp = _one_line_text(str(mine_src or "")) or "-"
            label_a = f"base={base_disp}"
            label_b = f"mine={mine_disp}"
            label_base = _one_line_text(getattr(self.app, "base_path", "") or "")
        is_merge_labels = bool(getattr(self.app, "merge_mode", False) and getattr(self.app, "has_base", False))
        path_bg_a = _MINE_BG if is_merge_labels else _BASE_BG
        path_bg_b = _THEIRS_BG if is_merge_labels else _MINE_BG

        self.path_label_a = tk.Label(
            path_bar,
            text=label_a,
            font=self._path_font,
            bg=path_bg_a,
            anchor="w",
            padx=6,
            pady=2,
        )
        self.path_label_a.grid(row=0, column=0, sticky="ew")
        self.path_label_base = tk.Label(
            path_bar,
            text=label_base if label_base else "基础(base): -",
            font=self._path_font,
            bg=_BASE_BG,
            anchor="w",
            padx=6,
            pady=2,
        )
        self.path_label_base.grid(row=0, column=1, sticky="ew")
        self.path_label_b = tk.Label(
            path_bar,
            text=label_b,
            font=self._path_font,
            bg=path_bg_b,
            anchor="w",
            padx=6,
            pady=2,
        )
        self.path_label_b.grid(row=0, column=2, sticky="ew")

        # Extra vertical scrollbar (left side) for convenience; controls both panes.
        # NOTE: must be packed BEFORE the paned window so it remains visible.
        self.vsb_left = ttk.Scrollbar(self.frame, orient="vertical", command=self._yview_both)
        self.vsb_left.pack(side="left", fill="y")
        # Diff minimap next to the left vertical scrollbar (more discoverable).
        self.vdiff_map = tk.Canvas(self.frame, width=10, highlightthickness=0, bg="#ebebeb")
        self.vdiff_map.pack(side="left", fill="y", padx=(0, 2))
        self.vdiff_map.bind("<Button-1>", self._on_vdiff_map_click)

        # Panes
        paned = ttk.PanedWindow(self.frame, orient="horizontal")
        paned.pack(fill="both", expand=True, padx=8, pady=(0, 8))
        self._main_paned = paned

        left_wrap = ttk.Frame(paned)
        mid_wrap = ttk.Frame(paned)
        right_wrap = ttk.Frame(paned)
        self._left_wrap = left_wrap
        self._mid_wrap = mid_wrap
        self._right_wrap = right_wrap
        paned.add(mid_wrap, weight=1)
        paned.add(left_wrap, weight=1)
        paned.add(right_wrap, weight=1)

        def _keep_panes_equal(_evt=None):
            # Keep A/B content panes at 50:50 to avoid visual width mismatch.
            try:
                total = self._main_paned.winfo_width()
                if total and total > 2:
                    if self._is_three_way_enabled():
                        self._main_paned.sashpos(0, total // 3)
                        self._main_paned.sashpos(1, (total * 2) // 3)
                    else:
                        self._main_paned.sashpos(0, total // 2)
            except Exception:
                pass

        self._keep_panes_equal = _keep_panes_equal
        self._main_paned.bind("<Configure>", self._keep_panes_equal)
        self._main_paned.bind("<ButtonRelease-1>", self._keep_panes_equal)
        self.frame.after(0, self._keep_panes_equal)

        self.left_title = ttk.Label(left_wrap, text="Mine", background=_MINE_BG)
        self.left_title.pack(fill="x")
        self.mid_title = ttk.Label(mid_wrap, text="Base", background=_BASE_BG)
        self.mid_title.pack(fill="x")
        self.right_title = ttk.Label(right_wrap, text="Base", background=_THEIRS_BG)
        self.right_title.pack(fill="x")

        # Font size tuned closer to TortoiseMerge (+~20%)
        self.editor_font = ("Consolas", 11)
        left_body = ttk.Frame(left_wrap)
        left_body.pack(fill="both", expand=True)
        mid_body = ttk.Frame(mid_wrap)
        mid_body.pack(fill="both", expand=True)
        right_body = ttk.Frame(right_wrap)
        right_body.pack(fill="both", expand=True)

        self.left_ln = tk.Text(left_body, width=self._row_header_width, wrap="none", undo=False, font=self.editor_font, bg="#efefef", fg="#666666", relief="flat", takefocus=0, cursor="arrow")
        self.base_ln = tk.Text(mid_body, width=self._row_header_width, wrap="none", undo=False, font=self.editor_font, bg="#efefef", fg="#666666", relief="flat", takefocus=0, cursor="arrow")
        self.right_ln = tk.Text(right_body, width=self._row_header_width, wrap="none", undo=False, font=self.editor_font, bg="#efefef", fg="#666666", relief="flat", takefocus=0, cursor="arrow")
        self.left = tk.Text(left_body, wrap="none", undo=False, font=self.editor_font, bg="white")
        self.base = tk.Text(mid_body, wrap="none", undo=False, font=self.editor_font, bg="white")
        self.right = tk.Text(right_body, wrap="none", undo=False, font=self.editor_font, bg="white")

        # Scrollbars
        # Per-pane vertical scrollbars (user requested visible scrollbars on both A and B)
        self.vsb_a = ttk.Scrollbar(left_body, orient="vertical", command=self._yview_both)
        self.vsb_m = ttk.Scrollbar(mid_body, orient="vertical", command=self._yview_both)
        self.vsb_b = ttk.Scrollbar(right_body, orient="vertical", command=self._yview_both)
        self.vsb_a.pack(side="right", fill="y")
        self.vsb_m.pack(side="right", fill="y")
        self.vsb_b.pack(side="right", fill="y")

        # Shared vertical scrollbar on far right removed (redundant / low usability).
        self.vsb = None

        # Horizontal scroll sync: keep A/B panes aligned when scrolling horizontally.
        self._xsyncing = False
        # Suppress C-pane -> main-pane x-sync during programmatic C text refresh.
        self._suppress_c_xsync = False

        def _sync_main_headers(first):
            try:
                frac = float(first)
            except Exception:
                return
            try:
                if getattr(self, "left_colhdr", None) is not None:
                    self.left_colhdr.xview_moveto(frac)
                if getattr(self, "right_colhdr", None) is not None:
                    self.right_colhdr.xview_moveto(frac)
                if self._is_three_way_enabled() and getattr(self, "base_colhdr", None) is not None:
                    self.base_colhdr.xview_moveto(frac)
            except Exception:
                pass

        def _xscroll_left(first, last):
            # called when left xview changes
            if self._is_click_trace_active():
                _dlog(f"xscroll_left first={first} last={last} xsyncing={self._xsyncing}")
            if self._xsyncing:
                self.hsb_left.set(first, last)
                return
            self._xsyncing = True
            try:
                self.hsb_left.set(first, last)
                self._sync_main_x_from_widget(self.left, first)
                self._sync_c_x_from_widget(self.left, first)
            finally:
                self._xsyncing = False
            try:
                self._update_diff_maps()
            except Exception:
                pass

        def _xscroll_right(first, last):
            if self._is_click_trace_active():
                _dlog(f"xscroll_right first={first} last={last} xsyncing={self._xsyncing}")
            if self._xsyncing:
                self.hsb_right.set(first, last)
                return
            self._xsyncing = True
            try:
                self.hsb_right.set(first, last)
                self._sync_main_x_from_widget(self.right, first)
                self._sync_c_x_from_widget(self.right, first)
            finally:
                self._xsyncing = False
            try:
                self._update_diff_maps()
            except Exception:
                pass

        def _xscroll_mid(first, last):
            if self._is_click_trace_active():
                _dlog(f"xscroll_mid first={first} last={last} xsyncing={self._xsyncing}")
            if self._xsyncing:
                self.hsb_mid.set(first, last)
                return
            self._xsyncing = True
            try:
                self.hsb_mid.set(first, last)
                self._sync_main_x_from_widget(self.base, first)
                self._sync_c_x_from_widget(self.base, first)
            finally:
                self._xsyncing = False
            try:
                self._update_diff_maps()
            except Exception:
                pass

        def _xview_left(*args):
            # scrollbar drag/click on left
            self._xsyncing = True
            try:
                self.left.xview(*args)
                first, last = self.left.xview()
                self.hsb_left.set(first, last)
                self._sync_main_x_from_widget(self.left, first)
                self._sync_c_x_from_widget(self.left, first)
            finally:
                self._xsyncing = False
            try:
                self._update_diff_maps()
            except Exception:
                pass

        def _xview_right(*args):
            self._xsyncing = True
            try:
                self.right.xview(*args)
                first, last = self.right.xview()
                self.hsb_right.set(first, last)
                self._sync_main_x_from_widget(self.right, first)
                self._sync_c_x_from_widget(self.right, first)
            finally:
                self._xsyncing = False
            try:
                self._update_diff_maps()
            except Exception:
                pass

        def _xview_mid(*args):
            self._xsyncing = True
            try:
                self.base.xview(*args)
                first, last = self.base.xview()
                self.hsb_mid.set(first, last)
                self._sync_main_x_from_widget(self.base, first)
                self._sync_c_x_from_widget(self.base, first)
            finally:
                self._xsyncing = False
            try:
                self._update_diff_maps()
            except Exception:
                pass

        self.hsb_left = ttk.Scrollbar(left_wrap, orient="horizontal", command=_xview_left)
        self.hsb_mid = ttk.Scrollbar(mid_wrap, orient="horizontal", command=_xview_mid)
        self.hsb_right = ttk.Scrollbar(right_wrap, orient="horizontal", command=_xview_right)
        self.hdiff_left = tk.Canvas(left_wrap, height=10, highlightthickness=0, bg="#ebebeb")
        self.hdiff_mid = tk.Canvas(mid_wrap, height=10, highlightthickness=0, bg="#ebebeb")
        self.hdiff_right = tk.Canvas(right_wrap, height=10, highlightthickness=0, bg="#ebebeb")
        self.hdiff_left.bind("<Button-1>", lambda e: self._on_hdiff_map_click(e, "left"))
        self.hdiff_mid.bind("<Button-1>", lambda e: self._on_hdiff_map_click(e, "mid"))
        self.hdiff_right.bind("<Button-1>", lambda e: self._on_hdiff_map_click(e, "right"))
        self.left.configure(xscrollcommand=_xscroll_left)
        self.base.configure(xscrollcommand=_xscroll_mid)
        self.right.configure(xscrollcommand=_xscroll_right)

        self.left.configure(yscrollcommand=self._yscroll_left)
        self.base.configure(yscrollcommand=self._yscroll_mid)
        self.right.configure(yscrollcommand=self._yscroll_right)
        self.vsb_left.configure(command=self._yview_both)
        self.vsb_a.configure(command=self._yview_both)
        self.vsb_m.configure(command=self._yview_both)
        self.vsb_b.configure(command=self._yview_both)

        # Excel-like column header row for A/Base/B panes.
        left_head = ttk.Frame(left_wrap)
        left_head.pack(fill="x", before=left_body)
        mid_head = ttk.Frame(mid_wrap)
        mid_head.pack(fill="x", before=mid_body)
        right_head = ttk.Frame(right_wrap)
        right_head.pack(fill="x", before=right_body)

        self.left_corner_hdr = tk.Text(left_head, width=self._row_header_width, height=1, wrap="none", undo=False,
                                       font=self.editor_font, bg="#efefef", fg="#666666", relief="flat", takefocus=0, cursor="arrow")
        self.base_corner_hdr = tk.Text(mid_head, width=self._row_header_width, height=1, wrap="none", undo=False,
                                       font=self.editor_font, bg="#efefef", fg="#666666", relief="flat", takefocus=0, cursor="arrow")
        self.right_corner_hdr = tk.Text(right_head, width=self._row_header_width, height=1, wrap="none", undo=False,
                                        font=self.editor_font, bg="#efefef", fg="#666666", relief="flat", takefocus=0, cursor="arrow")
        self.left_colhdr = tk.Text(left_head, height=1, wrap="none", undo=False,
                                   font=self.editor_font, bg="#efefef", fg="#555555", relief="flat", takefocus=0, cursor="arrow")
        self.base_colhdr = tk.Text(mid_head, height=1, wrap="none", undo=False,
                                   font=self.editor_font, bg="#efefef", fg="#555555", relief="flat", takefocus=0, cursor="arrow")
        self.right_colhdr = tk.Text(right_head, height=1, wrap="none", undo=False,
                                    font=self.editor_font, bg="#efefef", fg="#555555", relief="flat", takefocus=0, cursor="arrow")

        self.left_corner_hdr.pack(side="left", fill="y")
        ttk.Separator(left_head, orient="vertical").pack(side="left", fill="y")
        self._vsb_hdr_spc_a = tk.Frame(left_head, width=1, bg="#efefef")
        self._vsb_hdr_spc_a.pack(side="right", fill="y")
        self.left_colhdr.pack(side="left", fill="x", expand=True)
        self.base_corner_hdr.pack(side="left", fill="y")
        ttk.Separator(mid_head, orient="vertical").pack(side="left", fill="y")
        self._vsb_hdr_spc_m = tk.Frame(mid_head, width=1, bg="#efefef")
        self._vsb_hdr_spc_m.pack(side="right", fill="y")
        self.base_colhdr.pack(side="left", fill="x", expand=True)
        self.right_corner_hdr.pack(side="left", fill="y")
        ttk.Separator(right_head, orient="vertical").pack(side="left", fill="y")
        self._vsb_hdr_spc_b = tk.Frame(right_head, width=1, bg="#efefef")
        self._vsb_hdr_spc_b.pack(side="right", fill="y")
        self.right_colhdr.pack(side="left", fill="x", expand=True)
        for w in (self.left_corner_hdr, self.base_corner_hdr, self.right_corner_hdr,
                  self.left_colhdr, self.base_colhdr, self.right_colhdr):
            try:
                w.configure(state="normal")
                w.delete("1.0", "end")
                w.insert("1.0", "")
                w.configure(state="disabled")
            except Exception:
                pass
        # Keep colhdr spacers in sync with the actual scrollbar width so that column
        # headers align precisely with the data widgets at all times.
        def _sync_vsb_hdr_spacers(event=None):
            try:
                w = self.vsb_a.winfo_width()
                if w > 1:
                    self._vsb_hdr_spc_a.configure(width=w)
                    self._vsb_hdr_spc_b.configure(width=w)
                    self._vsb_hdr_spc_m.configure(width=w)
            except Exception:
                pass
        self.vsb_a.bind("<Configure>", _sync_vsb_hdr_spacers, "+")

        self.left_ln.pack(side="left", fill="y")
        ttk.Separator(left_body, orient="vertical").pack(side="left", fill="y")
        self.left.pack(fill="both", expand=True)
        self.hsb_left.pack(fill="x")
        self.hdiff_left.pack(fill="x")
        self.base_ln.pack(side="left", fill="y")
        ttk.Separator(mid_body, orient="vertical").pack(side="left", fill="y")
        self.base.pack(fill="both", expand=True)
        self.hsb_mid.pack(fill="x")
        self.hdiff_mid.pack(fill="x")

        # Save action row: keep a fixed height on both sides so horizontal
        # scrollbars stay aligned even when only one side has a button.
        # Also keep middle pane height identical to left/right to avoid row misalignment.
        save_row_height = 34

        # Save A button (bottom-right of A pane)
        save_a_row = ttk.Frame(left_wrap, height=save_row_height)
        save_a_row.pack(fill="x", pady=(2, 0))
        save_a_row.pack_propagate(False)
        if getattr(self.app, "merge_mode", False):
            tk.Button(save_a_row, text="保存Merged并退出", bg="#eaf2ff", padx=14, pady=4,
                      command=self.app.save_merged_and_exit).pack(side="right")
        else:
            if not getattr(self.app, "diff_base_mine_mode", False):
                if not _is_temp_base_path(getattr(self.app, "file_a", "")):
                    tk.Button(save_a_row, text="保存A", bg="#eaf2ff", padx=14, pady=4,
                              command=self.app.save_a_inplace).pack(side="right")

        # Base pane spacer: maintain same bottom reserved height as A/B panes.
        save_mid_row = ttk.Frame(mid_wrap, height=save_row_height)
        save_mid_row.pack(fill="x", pady=(2, 0))
        save_mid_row.pack_propagate(False)

        self.right_ln.pack(side="left", fill="y")
        ttk.Separator(right_body, orient="vertical").pack(side="left", fill="y")
        self.right.pack(fill="both", expand=True)
        self.hsb_right.pack(fill="x")
        self.hdiff_right.pack(fill="x")

        # Save B button (bottom-right of B pane)
        save_b_row = ttk.Frame(right_wrap, height=save_row_height)
        save_b_row.pack(fill="x", pady=(2, 0))
        save_b_row.pack_propagate(False)
        if not getattr(self.app, "merge_mode", False):
            save_b_text = "Save Mine" if getattr(self.app, "diff_base_mine_mode", False) else "Save B"
            tk.Button(save_b_row, text=save_b_text, bg="#ffecec", padx=14, pady=4,
                      command=self.app.save_b_inplace).pack(side="right")

        # Tags (order matters: diffcell should be applied after diffrow)
        # Closer to TortoiseMerge vibe: left diff block = orange, right diff block = yellow
        self.left.tag_configure("diffrow", background=_MINE_BG)
        self.base.tag_configure("diffrow", background=_BASE_BG)
        self.right.tag_configure("diffrow", background=_THEIRS_BG)

        # Cell-level highlight (red) for exact diffs
        self.left.tag_configure("diffcell", background=_DIFF_CELL_BG)
        self.base.tag_configure("diffcell", background=_DIFF_CELL_BG)
        self.right.tag_configure("diffcell", background=_DIFF_CELL_BG)
        # Alignment padding: grey slot for rows that exist only on the other side.
        # tag_raise ensures paddingrow background overrides diffrow on the empty slot.
        self.left.tag_configure("paddingrow", background="#A0A0A0")
        self.base.tag_configure("paddingrow", background="#A0A0A0")
        self.right.tag_configure("paddingrow", background="#A0A0A0")
        self.left.tag_raise("paddingrow")
        self.base.tag_raise("paddingrow")
        self.right.tag_raise("paddingrow")

        # selection should not overwrite diff colors
        self.left.tag_configure("selrow", underline=1, font=("Consolas", 11, "bold"))
        self.base.tag_configure("selrow", underline=1, font=("Consolas", 11, "bold"))
        self.right.tag_configure("selrow", underline=1, font=("Consolas", 11, "bold"))
        # Selected-cell highlight (same blue as C区 cselcell), applied on A/B/(Base).
        self.left.tag_configure("selcell", background="#8EB9FF")
        self.base.tag_configure("selcell", background="#8EB9FF")
        self.right.tag_configure("selcell", background="#8EB9FF")
        self.left.tag_raise("selcell")
        self.base.tag_raise("selcell")
        self.right.tag_raise("selcell")
        self.left_ln.tag_configure("diffrow", background="#ffd9d9")
        self.base_ln.tag_configure("diffrow", background="#ffd9d9")
        self.right_ln.tag_configure("diffrow", background="#ffd9d9")

        # Bindings
        self._syncing = False
        for w in (self.left, self.base, self.right):
            w.bind("<MouseWheel>", self._on_mousewheel)
            w.bind("<Button-4>", self._on_mousewheel)
            w.bind("<Button-5>", self._on_mousewheel)
            w.bind("<KeyRelease>", lambda e: self._update_cursor_lines())
            w.bind("<ButtonRelease-1>", lambda e: self._update_cursor_lines())
            if getattr(self.app, "merge_conflict_mode", False):
                #快捷键：下一处/上一处冲突
                w.bind("<F4>", lambda e: (self._goto_next_diff_block(), "break"))
                w.bind("<Shift-F4>", lambda e: (self._goto_prev_diff_block(), "break"))
                w.bind("<Control-n>", lambda e: (self._goto_next_diff_block(), "break"))
                w.bind("<Control-p>", lambda e: (self._goto_prev_diff_block(), "break"))

        # Click handling (selection + arrow action)
        left_click_dir = "MINE2A" if (getattr(self.app, "merge_mode", False) and getattr(self.app, "has_base", False)) else "A2B"
        self.left.bind("<Button-1>", lambda e, d=left_click_dir: self._on_click_with_arrow(self.left, e, d))
        self.base.bind("<Button-1>", lambda e: self._on_click_with_arrow(self.base, e, "BASE2A"))
        self.right.bind("<Button-1>", lambda e: self._on_click_with_arrow(self.right, e, "B2A"))
        self.left.bind("<Motion>", lambda e: self._on_cell_hover_tooltip(self.left, e, "A"))
        self.base.bind("<Motion>", lambda e: self._on_cell_hover_tooltip(self.base, e, "BASE"))
        self.right.bind("<Motion>", lambda e: self._on_cell_hover_tooltip(self.right, e, "B"))
        self.left.bind("<Leave>", lambda e: self._hide_cell_tooltip())
        self.base.bind("<Leave>", lambda e: self._hide_cell_tooltip())
        self.right.bind("<Leave>", lambda e: self._hide_cell_tooltip())
        # Double-click merge (single cell)
        self.left.bind("<Double-Button-1>", lambda e, d=left_click_dir: self._copy_cell(d, e))
        self.base.bind("<Double-Button-1>", lambda e: self._copy_cell("BASE2A", e))
        self.right.bind("<Double-Button-1>", lambda e: self._copy_cell("B2A", e))
        self.left_ln.bind("<Button-1>", lambda e, d=left_click_dir: self._on_row_header_click(self.left_ln, e, d))
        self.base_ln.bind("<Button-1>", lambda e: self._on_row_header_click(self.base_ln, e, "BASE2A"))
        self.right_ln.bind("<Button-1>", lambda e: self._on_row_header_click(self.right_ln, e, "B2A"))
        # Keep old habit compatibility: right-click on row header also triggers row apply.
        self.left_ln.bind("<Button-3>", lambda e, d=left_click_dir: self._on_row_header_click(self.left_ln, e, d))
        self.base_ln.bind("<Button-3>", lambda e: self._on_row_header_click(self.base_ln, e, "BASE2A"))
        self.right_ln.bind("<Button-3>", lambda e: self._on_row_header_click(self.right_ln, e, "B2A"))
        self.left_ln.bind("<Motion>", lambda e, d=left_click_dir: self._on_row_header_hover(self.left_ln, e, d))
        self.base_ln.bind("<Motion>", lambda e: self._on_row_header_hover(self.base_ln, e, "BASE2A"))
        self.right_ln.bind("<Motion>", lambda e: self._on_row_header_hover(self.right_ln, e, "B2A"))
        self.left_ln.bind("<Leave>", lambda e: self._clear_row_header_hover(self.left_ln))
        self.base_ln.bind("<Leave>", lambda e: self._clear_row_header_hover(self.base_ln))
        self.right_ln.bind("<Leave>", lambda e: self._clear_row_header_hover(self.right_ln))

        # C区: compact cursor compare block + cell-aligned view
        self.c_area = ttk.Notebook(self.frame)
        self.c_area.pack(fill="x", padx=8, pady=(0, 4))

        # ---- C1: compact row compare (2 lines in 2-way, 3 lines in 3-way) ----
        c_text_frame = ttk.Frame(self.c_area)
        self.c_area.add(c_text_frame, text="C区-行对比")

        # C区 column header + row-header layout (Excel-like)
        c_head = ttk.Frame(c_text_frame)
        c_head.pack(side="top", fill="x")
        self.cursor_cmp_corner = tk.Text(c_head, width=self._row_header_width, height=1, wrap="none", undo=False,
                                         font=self.editor_font, bg="#efefef", fg="#666666", relief="flat", takefocus=0, cursor="arrow")
        self.cursor_cmp_colhdr = tk.Text(c_head, height=1, wrap="none", undo=False,
                                         font=self.editor_font, bg="#efefef", fg="#555555", relief="flat", takefocus=0, cursor="arrow")
        self.cursor_cmp_corner.pack(side="left", fill="y")
        ttk.Separator(c_head, orient="vertical").pack(side="left", fill="y")
        self.cursor_cmp_colhdr.pack(side="left", fill="x", expand=True)

        c_body = ttk.Frame(c_text_frame)
        c_body.pack(side="top", fill="x")
        self.cursor_cmp_ln = tk.Text(c_body, width=self._row_header_width, height=3 if self._is_three_way_enabled() else 2, wrap="none", undo=False,
                                     font=self.editor_font, bg="#efefef", fg="#666666", relief="flat", takefocus=0, cursor="arrow")
        self.cursor_cmp_ln.pack(side="left", fill="y")
        ttk.Separator(c_body, orient="vertical").pack(side="left", fill="y")

        self.cursor_cmp = tk.Text(
            c_body,
            height=3 if self._is_three_way_enabled() else 2,
            wrap="none",
            font=self.editor_font,
            bd=1,
            relief="solid",
        )
        # Make base colors stronger (user feedback: previous too light)
        self.cursor_cmp.tag_configure("a", background=_MINE_BG)
        self.cursor_cmp.tag_configure("base", background=_BASE_BG)
        self.cursor_cmp.tag_configure("b", background=_THEIRS_BG)
        self.cursor_cmp.tag_configure("missing", background="#a6a6a6")
        # Diff cell highlight (match main panes)
        self.cursor_cmp.tag_configure("diffcell", background=_DIFF_CELL_BG)
        # Explicit C-area click target highlight
        self.cursor_cmp.tag_configure("cselcell", background="#8EB9FF")
        self.cursor_cmp.pack(side="left", fill="x", expand=True)

        for w in (self.cursor_cmp_corner, self.cursor_cmp_colhdr, self.cursor_cmp_ln):
            try:
                w.configure(state="normal")
                w.delete("1.0", "end")
                w.insert("1.0", "")
                w.configure(state="disabled")
            except Exception:
                pass

        # Horizontal scrollbar for C区行对比（sync with main panes）
        self.cursor_hsb = ttk.Scrollbar(c_text_frame, orient="horizontal", command=self._xview_cursor_cmp)
        self.cursor_cmp.configure(xscrollcommand=self._xscroll_cursor_cmp)
        self.cursor_hsb.pack(side="top", fill="x")
        self.cursor_cmp.bind("<Button-1>", self._on_cursor_cmp_click)
        self.cursor_cmp.bind("<Double-Button-1>", self._on_cursor_cmp_double_click)

        # ---- C2: cell-aligned view (optional; can be hidden if not useful/performance) ----
        self._enable_c_cell = False  # user feedback: not useful; keep hidden by default
        c_cell_frame = ttk.Frame(self.c_area)
        self.c_area.add(c_cell_frame, text="C区-单元格对齐")
        if not self._enable_c_cell:
            try:
                self.c_area.tab(c_cell_frame, state="hidden")
            except Exception:
                pass

        top_row = ttk.Frame(c_cell_frame)
        top_row.pack(fill="x", pady=(2, 2))
        self.c_only_diff_cells = tk.IntVar(value=1)
        tk.Checkbutton(
            top_row,
            text="只显示差异单元格",
            variable=self.c_only_diff_cells,
            onvalue=1,
            offvalue=0,
            command=lambda: self._update_cursor_lines(),
        ).pack(side="left")

        self.cell_cmp_text = tk.Text(c_cell_frame, height=6, wrap="none", font=self.editor_font, bd=1, relief="solid")
        self.cell_cmp_text.tag_configure("a", background=_MINE_BG)
        self.cell_cmp_text.tag_configure("b", background=_THEIRS_BG)
        self.cell_cmp_text.tag_configure("diffcell", background=_DIFF_CELL_BG)

        self.cell_cmp_hsb = ttk.Scrollbar(c_cell_frame, orient="horizontal", command=self._xview_cell_cmp)
        self.cell_cmp_text.configure(xscrollcommand=self._xscroll_cell_cmp)

        self.cell_cmp_text.pack(side="top", fill="x", expand=True)
        self.cell_cmp_hsb.pack(side="top", fill="x")

        # initial render should respect the persisted only-diff setting
        # Defer heavy initial refresh; SowMergeApp will lazy-load the active sheet.
        # (Still create the UI widgets now.)
        # self.refresh(row_only=None, rescan=True)
        # self._update_cursor_lines()
        # Initial panel state (must run after C区 widgets are created)
        self._toggle_three_way_view(init_only=True)

    # ---------- Scrolling sync ----------
    def _is_three_way_enabled(self) -> bool:
        try:
            return bool(getattr(self, "three_way_var", None) and self.three_way_var.get() and getattr(self.app, "merge_mode", False) and getattr(self.app, "has_base", False))
        except Exception:
            return False

    @staticmethod
    def _source_display_name(path_like: str) -> str:
        """Prefer stable workbook names and keep SVN revision hints when available."""
        s = (path_like or "").strip().strip('"')
        if not s:
            return "-"

        # Synthetic marker produced by parser when BASE is exported from .svn.
        m = re.match(r"(.+?)@BASE\(\.svn\)$", s, re.IGNORECASE)
        if m:
            return f"{os.path.basename(m.group(1))}@BASE"

        bn = os.path.basename(s)
        ext_pat = r"(?:xlsx|xlsm|csv)"

        # Exported revision snapshots created by this tool.
        m = re.match(rf"{re.escape(APP_NAME)}_svncat_r(\d+)_\d{{8}}_\d{{6}}_(.+?\.{ext_pat})$", bn, re.IGNORECASE)
        if m:
            return f"{m.group(2)}@r{m.group(1)}"
        m = re.match(rf"{re.escape(APP_NAME)}_svncat_BASE_\d{{8}}_\d{{6}}_(.+?\.{ext_pat})$", bn, re.IGNORECASE)
        if m:
            return f"{m.group(1)}@BASE"

        # Temp wrapper pattern: xxx.xlsx-rev123.svn456.tmp.xlsx -> xxx.xlsx@r123
        m = re.match(rf"(.+?\.{ext_pat})-rev(\d+)\.svn\d+\.tmp\.{ext_pat}$", bn, re.IGNORECASE)
        if m:
            return f"{m.group(1)}@r{m.group(2)}"

        # SVN conflict side files: file.merge-left.r123 -> file@r123
        m = re.match(rf"(.+?)\.merge-(?:left|right)\.r(\d+)$", bn, re.IGNORECASE)
        if m:
            return f"{m.group(1)}@r{m.group(2)}"

        # Generic prefixed temp: keep original workbook tail if present.
        m = re.match(rf".*?_(.+?\.{ext_pat})$", bn, re.IGNORECASE)
        if m and ".merge-" not in m.group(1).lower():
            return m.group(1)
        return bn

    def _toggle_three_way_view(self, init_only: bool = False):
        enabled = self._is_three_way_enabled()
        def _one_line_text(s: str, max_len: int = 120) -> str:
            s = (s or "").replace("\r", " ").replace("\n", " ")
            if len(s) <= max_len:
                return s
            return "..." + s[-(max_len - 3):]
        try:
            panes = list(self._main_paned.panes())
            mid_id = str(self._mid_wrap)
            has_mid = mid_id in panes
            if enabled and (not has_mid):
                self._main_paned.insert(0, self._mid_wrap, weight=1)
            elif (not enabled) and has_mid:
                self._main_paned.forget(self._mid_wrap)
        except Exception:
            pass
        try:
            if enabled:
                # Layout: left=Base, center=Mine, right=Theirs — reorder header labels to match
                self.path_label_base.grid(row=0, column=0, sticky="ew")
                self.path_label_a.grid(row=0, column=1, sticky="ew")
                self.path_label_b.grid(row=0, column=2, sticky="ew")
                self.left_title.configure(text="Mine", background=_MINE_BG)
                self.mid_title.configure(text="Base", background=_BASE_BG)
                self.right_title.configure(text="Theirs", background=_THEIRS_BG)
                mine_src = getattr(self.app, "raw_mine", None) or self.app.file_a
                base_src = getattr(self.app, "raw_base", None) or getattr(self.app, "base_path", "")
                theirs_src = getattr(self.app, "raw_theirs", None) or self.app.file_b
                self.path_label_a.configure(text=f"mine={self._source_display_name(mine_src)}", bg=_MINE_BG)
                self.path_label_base.configure(text=f"base={self._source_display_name(base_src)}" if base_src else "base=-", bg=_BASE_BG)
                self.path_label_b.configure(text=f"theirs={self._source_display_name(theirs_src)}", bg=_THEIRS_BG)
            else:
                self.path_label_base.grid_remove()
                # Restore 2-way labels to left/right columns
                self.path_label_a.grid(row=0, column=0, sticky="ew")
                self.path_label_b.grid(row=0, column=2, sticky="ew")
                self.left_title.configure(text="Base", background=_BASE_BG)
                self.right_title.configure(text="Mine", background=_MINE_BG)
                self.mid_title.configure(text="Base", background=_BASE_BG)
                base_src = getattr(self.app, "raw_base", None) or self.app.file_a
                mine_src = getattr(self.app, "raw_mine", None) or self.app.file_b
                base_disp = _one_line_text(str(base_src or "")) or "-"
                mine_disp = _one_line_text(str(mine_src or "")) or "-"
                self.path_label_a.configure(text=f"base={base_disp}", bg=_BASE_BG)
                self.path_label_b.configure(text=f"mine={mine_disp}", bg=_MINE_BG)
        except Exception:
            pass
        try:
            self.cursor_cmp.configure(height=3 if enabled else 2)
        except Exception:
            pass
        try:
            self.cursor_cmp_ln.configure(height=3 if enabled else 2)
        except Exception:
            pass
        if not init_only:
            try:
                self.refresh(row_only=None, rescan=False)
                self._update_cursor_lines()
            except Exception:
                pass
        try:
            self.frame.after(0, self._keep_panes_equal)
        except Exception:
            pass

    def _sync_main_x_to_frac(self, first):
        try:
            frac = float(first)
        except Exception:
            return

        prev_xsync = bool(getattr(self, "_xsyncing", False))
        self._xsyncing = True
        try:
            # Use left pane as anchor, then map by pixel position to peers.
            # This avoids drift when panes render different glyph widths.
            self.left.xview_moveto(frac)
            self._sync_main_x_from_widget(self.left, frac)
        except Exception:
            pass
        finally:
            self._xsyncing = prev_xsync

    def _sync_c_x_to_frac(self, first):
        try:
            frac = float(first)
        except Exception:
            return

        # Programmatic C-pane sync must not feed back into main panes.
        prev_suppress = bool(getattr(self, "_suppress_c_xsync", False))
        self._suppress_c_xsync = True
        try:
            try:
                self.cursor_cmp.xview_moveto(frac)
                if getattr(self, "cursor_cmp_colhdr", None) is not None:
                    self.cursor_cmp_colhdr.xview_moveto(frac)
                cf, cl = self.cursor_cmp.xview()
                self.cursor_hsb.set(cf, cl)
            except Exception:
                pass
            try:
                self.cell_cmp_text.xview_moveto(frac)
                sf, sl = self.cell_cmp_text.xview()
                self.cell_cmp_hsb.set(sf, sl)
            except Exception:
                pass
        finally:
            self._suppress_c_xsync = prev_suppress

    @staticmethod
    def _clamp(value: float, lo: float, hi: float) -> float:
        try:
            v = float(value)
        except Exception:
            return lo
        if v < lo:
            return lo
        if v > hi:
            return hi
        return v

    def _map_xfirst_between_widgets(self, src_widget: tk.Text, dst_widget: tk.Text, src_first) -> float:
        """Map horizontal scroll position by pixel offset, with end-point locking."""
        try:
            sf, sl = src_widget.xview()
            sf = float(sf)
            sl = float(sl)
            df, dl = dst_widget.xview()
            df = float(df)
            dl = float(dl)
            src_first_f = float(src_first)
            src_vis = max(1e-6, sl - sf)
            dst_vis = max(1e-6, dl - df)
            src_max = max(0.0, 1.0 - src_vis)
            dst_max = max(0.0, 1.0 - dst_vis)
            src_first_f = self._clamp(src_first_f, 0.0, src_max)

            # Lock endpoints: when source is at extreme left/right, target should also
            # reach its own extreme to avoid "not fully right" under different viewports.
            eps = 1e-4
            if src_first_f <= eps:
                return 0.0
            if src_max <= eps:
                return 0.0
            if src_first_f >= src_max - eps:
                return dst_max

            src_vp = max(1.0, float(src_widget.winfo_width()))
            dst_vp = max(1.0, float(dst_widget.winfo_width()))
            src_total = src_vp / src_vis
            dst_total = dst_vp / dst_vis
            if src_total <= 1e-6 or dst_total <= 1e-6:
                return self._clamp(src_first_f, 0.0, dst_max)

            # Convert to absolute left-pixel position in source, then map to destination ratio.
            left_px = src_first_f * src_total
            dst_first = left_px / dst_total
            return self._clamp(dst_first, 0.0, dst_max)
        except Exception:
            try:
                return float(src_first)
            except Exception:
                return 0.0

    def _sync_main_x_from_widget(self, src_widget: tk.Text, src_first):
        """Sync A/B(/BASE) using source-widget absolute x position."""
        try:
            left_first = self._map_xfirst_between_widgets(src_widget, self.left, src_first)
            right_first = self._map_xfirst_between_widgets(src_widget, self.right, src_first)
            self.left.xview_moveto(left_first)
            self.right.xview_moveto(right_first)
            if self._is_three_way_enabled():
                base_first = self._map_xfirst_between_widgets(src_widget, self.base, src_first)
                self.base.xview_moveto(base_first)
            if getattr(self, "left_colhdr", None) is not None:
                self.left_colhdr.xview_moveto(left_first)
            if getattr(self, "right_colhdr", None) is not None:
                self.right_colhdr.xview_moveto(right_first)
            if self._is_three_way_enabled() and getattr(self, "base_colhdr", None) is not None:
                base_first = float((self.base.xview() or (0.0, 1.0))[0])
                self.base_colhdr.xview_moveto(base_first)
            lf, ll = self.left.xview()
            rf, rl = self.right.xview()
            self.hsb_left.set(lf, ll)
            self.hsb_right.set(rf, rl)
            if self._is_three_way_enabled():
                mf, ml = self.base.xview()
                self.hsb_mid.set(mf, ml)
        except Exception:
            pass

    def _sync_c_x_from_widget(self, src_widget: tk.Text, src_first):
        """Sync C panes using source-widget absolute x position."""
        prev_suppress = bool(getattr(self, "_suppress_c_xsync", False))
        self._suppress_c_xsync = True
        try:
            try:
                c_first = self._map_xfirst_between_widgets(src_widget, self.cursor_cmp, src_first)
                self.cursor_cmp.xview_moveto(c_first)
                if getattr(self, "cursor_cmp_colhdr", None) is not None:
                    self.cursor_cmp_colhdr.xview_moveto(c_first)
                cf, cl = self.cursor_cmp.xview()
                self.cursor_hsb.set(cf, cl)
            except Exception:
                pass
            try:
                if hasattr(self, "cell_cmp_text"):
                    cell_first = self._map_xfirst_between_widgets(src_widget, self.cell_cmp_text, src_first)
                    self.cell_cmp_text.xview_moveto(cell_first)
                    sf, sl = self.cell_cmp_text.xview()
                    self.cell_cmp_hsb.set(sf, sl)
            except Exception:
                pass
        finally:
            self._suppress_c_xsync = prev_suppress

    def _is_click_trace_active(self) -> bool:
        try:
            return float(getattr(self, "_trace_click_until", 0.0) or 0.0) > time.time()
        except Exception:
            return False

    def _log_click_trace_state(self, stage: str):
        if not self._is_click_trace_active():
            return
        try:
            lx = float((self.left.xview() or (0.0, 1.0))[0])
        except Exception:
            lx = -1.0
        try:
            rx = float((self.right.xview() or (0.0, 1.0))[0])
        except Exception:
            rx = -1.0
        try:
            cx = float((self.cursor_cmp.xview() or (0.0, 1.0))[0])
        except Exception:
            cx = -1.0
        try:
            li = str(self.left.index("insert"))
        except Exception:
            li = "?"
        try:
            ri = str(self.right.index("insert"))
        except Exception:
            ri = "?"
        _dlog(f"click_trace {stage} x(left={lx:.6f},right={rx:.6f},c={cx:.6f}) insert(left={li},right={ri})")

    def _post_click_x_guard(self, saved_x: float, stage: str):
        try:
            now_x = float((self.left.xview() or (0.0, 1.0))[0])
        except Exception:
            now_x = saved_x
        drift = abs(now_x - saved_x)
        if drift > 1e-4:
            try:
                self._sync_main_x_to_frac(saved_x)
                self._sync_c_x_to_frac(saved_x)
            except Exception:
                pass
            _dlog(f"click_guard_restore stage={stage} saved={saved_x:.6f} now={now_x:.6f} drift={drift:.6f}")
        else:
            _dlog(f"click_guard_ok stage={stage} saved={saved_x:.6f} now={now_x:.6f}")
        self._log_click_trace_state(f"post_guard:{stage}")

    def _xview_cursor_cmp(self, *args):
        if self._xsyncing:
            return
        self._xsyncing = True
        try:
            # Drive the main pane directly so the full A/B scroll range is accessible.
            # C区's viewport is wider, so using cursor_cmp as driver would clamp
            # the scroll position and prevent reaching the rightmost content in A/B.
            self.left.xview(*args)
            first, last = self.left.xview()
            self._sync_main_x_to_frac(first)
            self._sync_c_x_from_widget(self.left, first)
            # Reflect main-pane scroll state in cursor_hsb for full-range thumb.
            self.cursor_hsb.set(first, last)
        finally:
            self._xsyncing = False
        try:
            self._update_diff_maps()
        except Exception:
            pass

    def _xscroll_cursor_cmp(self, first, last):
        if self._is_click_trace_active():
            _dlog(f"xscroll_cursor_cmp first={first} last={last} xsyncing={self._xsyncing} suppress={getattr(self, '_suppress_c_xsync', False)}")
        # Passive C-pane xscroll callback should never drive main panes.
        # Only explicit C scrollbar command handlers (_xview_cursor_cmp/_xview_cell_cmp)
        # are allowed to sync main xview.
        self.cursor_hsb.set(first, last)
        return

    def _xview_cell_cmp(self, *args):
        if self._xsyncing:
            return
        self._xsyncing = True
        try:
            self.cell_cmp_text.xview(*args)
            first, last = self.cell_cmp_text.xview()
            self.cell_cmp_hsb.set(first, last)
            self._sync_main_x_from_widget(self.cell_cmp_text, first)
            self._sync_c_x_from_widget(self.cell_cmp_text, first)
        finally:
            self._xsyncing = False
        try:
            self._update_diff_maps()
        except Exception:
            pass

    def _xscroll_cell_cmp(self, first, last):
        if self._is_click_trace_active():
            _dlog(f"xscroll_cell_cmp first={first} last={last} xsyncing={self._xsyncing} suppress={getattr(self, '_suppress_c_xsync', False)}")
        # Same rule as cursor_cmp: passive callback updates only its own scrollbar.
        self.cell_cmp_hsb.set(first, last)
        return

    def _is_grid_overlay_enabled(self) -> bool:
        try:
            return bool(getattr(self, "grid_overlay_var", None) and self.grid_overlay_var.get())
        except Exception:
            return False

    def _gridify_parts(self, parts: list[str]) -> list[str]:
        if not self._is_grid_overlay_enabled():
            return parts
        # Keep tab layout unchanged; prepend a visual splitter in each cell.
        return [f"|{p}" for p in parts]

    def _toggle_grid_overlay(self):
        try:
            self._invalidate_render_cache()
            # pair_text_a/b cache formatted line strings; must be cleared so
            # _build_row_and_diff_pair re-runs with the new grid-on/off separator.
            self.pair_text_a = {}
            self.pair_text_b = {}
            # Rescan widths to avoid stale narrow columns when toggling grid mode.
            self.refresh(row_only=None, rescan=True)
            self._update_cursor_lines()
        except Exception:
            pass

    def _yscroll_all(self, first, last):
        for sb in (self.vsb_left, self.vsb_a, self.vsb_m, self.vsb_b):
            try:
                sb.set(first, last)
            except Exception:
                pass
        try:
            self._update_diff_maps()
        except Exception:
            pass

    def _yscroll_left(self, first, last):
        if self._syncing:
            return
        self._syncing = True
        try:
            self.left_ln.yview_moveto(first)
            if self._is_three_way_enabled():
                self.base.yview_moveto(first)
                self.base_ln.yview_moveto(first)
            self.right.yview_moveto(first)
            self.right_ln.yview_moveto(first)
            self._yscroll_all(first, last)
        finally:
            self._syncing = False
        self._maybe_load_more_rows(last)

    def _yscroll_mid(self, first, last):
        if self._syncing:
            return
        self._syncing = True
        try:
            self.left.yview_moveto(first)
            self.left_ln.yview_moveto(first)
            self.right.yview_moveto(first)
            self.right_ln.yview_moveto(first)
            self.base_ln.yview_moveto(first)
            self._yscroll_all(first, last)
        finally:
            self._syncing = False
        self._maybe_load_more_rows(last)

    def _yscroll_right(self, first, last):
        if self._syncing:
            return
        self._syncing = True
        try:
            if self._is_three_way_enabled():
                self.base.yview_moveto(first)
                self.base_ln.yview_moveto(first)
            self.left.yview_moveto(first)
            self.left_ln.yview_moveto(first)
            self.right_ln.yview_moveto(first)
            self._yscroll_all(first, last)
        finally:
            self._syncing = False
        self._maybe_load_more_rows(last)

    def _yview_both(self, *args):
        self._syncing = True
        try:
            self.left.yview(*args)
            self.left_ln.yview(*args)
            if self._is_three_way_enabled():
                self.base.yview(*args)
                self.base_ln.yview(*args)
            self.right.yview(*args)
            self.right_ln.yview(*args)
            try:
                first, last = self.left.yview()
                self._yscroll_all(first, last)
            except Exception:
                pass
        finally:
            self._syncing = False
        try:
            _first, last = self.left.yview()
            self._maybe_load_more_rows(last)
        except Exception:
            pass

    def _on_mousewheel(self, event):
        if getattr(event, "num", None) == 4:
            delta = 120
        elif getattr(event, "num", None) == 5:
            delta = -120
        else:
            delta = event.delta
        steps = int(-1 * (delta / 120))
        self._yview_both("scroll", steps, "units")
        return "break"

    def _on_vdiff_map_click(self, event):
        try:
            h = max(1, self.vdiff_map.winfo_height())
            frac = min(1.0, max(0.0, float(event.y) / float(h)))
            self._yview_both("moveto", frac)
        except Exception:
            pass

    def _on_hdiff_map_click(self, event, pane: str):
        try:
            canvas = self.hdiff_left if pane == "left" else (self.hdiff_mid if pane == "mid" else self.hdiff_right)
            w = max(1, canvas.winfo_width())
            frac = min(1.0, max(0.0, float(event.x) / float(w)))
            self._sync_main_x_to_frac(frac)
            self._sync_c_x_to_frac(frac)
        except Exception:
            pass

    def _update_diff_maps(self):
        # Vertical diff map (by display row index)
        try:
            self.vdiff_map.delete("all")
            h = max(1, self.vdiff_map.winfo_height())
            w = max(1, self.vdiff_map.winfo_width())
            rows = self._full_display_rows if self._full_display_rows else self.display_rows
            n = max(1, len(rows))
            diff_mask = [bool(self.pair_diff_cols.get(pidx)) for pidx in rows]
            diff_count = sum(1 for v in diff_mask if v)
            # Dynamic marker height:
            # - fewer diffs -> thicker marker for visibility
            # - many diffs -> thinner marker to reduce overlap
            marker_min_h = max(2, min(12, int(h / max(1, diff_count)))) if diff_count > 0 else 2

            # Draw contiguous diff blocks as filled segments for better discoverability.
            i = 0
            while i < n:
                if not diff_mask[i]:
                    i += 1
                    continue
                j = i
                while j + 1 < n and diff_mask[j + 1]:
                    j += 1
                y1 = int((i / n) * h)
                y2 = int(((j + 1) / n) * h)
                if y2 - y1 < marker_min_h:
                    y2 = min(h, y1 + marker_min_h)
                self.vdiff_map.create_rectangle(0, y1, w, y2, outline="", fill="#ff2d2d")
                i = j + 1
            try:
                first, last = self.left.yview()
                y1 = int(first * h)
                y2 = max(y1 + 2, int(last * h))
                self.vdiff_map.create_rectangle(0, y1, w, y2, outline="#1e78ff")
            except Exception:
                pass
        except Exception:
            pass

        # Horizontal diff maps (under each pane scrollbar; by diff columns)
        try:
            diff_cols = set()
            for cols in (self.pair_diff_cols or {}).values():
                if cols:
                    diff_cols.update(cols)
            max_col = max(1, int(self.max_col or 1))
            canvases = [self.hdiff_left, self.hdiff_right]
            if self._is_three_way_enabled():
                canvases.insert(1, self.hdiff_mid)
            for canvas in canvases:
                canvas.delete("all")
                cw = max(1, canvas.winfo_width())
                ch = max(1, canvas.winfo_height())
                if diff_cols:
                    marker_min_w = max(2, min(14, int(cw / max(1, len(diff_cols)))))
                    sorted_cols = sorted(diff_cols)
                    seg_start = sorted_cols[0]
                    seg_end = sorted_cols[0]
                    for c in sorted_cols[1:]:
                        if c == seg_end + 1:
                            seg_end = c
                        else:
                            x1 = int(((seg_start - 1) / max_col) * cw)
                            x2 = int((seg_end / max_col) * cw)
                            if x2 - x1 < marker_min_w:
                                x2 = min(cw, x1 + marker_min_w)
                            canvas.create_rectangle(x1, 0, x2, ch, outline="", fill="#c46a00")
                            seg_start = c
                            seg_end = c
                    x1 = int(((seg_start - 1) / max_col) * cw)
                    x2 = int((seg_end / max_col) * cw)
                    if x2 - x1 < marker_min_w:
                        x2 = min(cw, x1 + marker_min_w)
                    canvas.create_rectangle(x1, 0, x2, ch, outline="", fill="#c46a00")
                try:
                    lf, ll = self.left.xview()
                    canvas.create_rectangle(int(lf * cw), 0, max(int(ll * cw), int(lf * cw) + 2), ch, outline="#1e78ff")
                except Exception:
                    pass
        except Exception:
            pass

    # ---------- Selection + toolbar buttons ----------
    def _widget_line(self, w: tk.Text):
        try:
            idx = w.index("@%d,%d" % (w.winfo_pointerx() - w.winfo_rootx(), w.winfo_pointery() - w.winfo_rooty()))
        except Exception:
            idx = w.index("insert")
        return int(str(idx).split(".")[0])

    def _pair_idx_for_line(self, line: int) -> int | None:
        if not (1 <= line <= len(self.display_rows)):
            return None
        return self.display_rows[line - 1]

    def _pair_for_line(self, line: int):
        idx = self._pair_idx_for_line(line)
        if idx is None or idx >= len(self.row_pairs):
            return None
        return self.row_pairs[idx]

    def _side_for_widget(self, w: tk.Text) -> str:
        if w is self.left:
            return "A"
        if w is self.base:
            return "BASE"
        return "B"

    def _col_from_char(self, char_pos: int) -> int | None:
        """Resolve display column index by character offset in a rendered row line."""
        try:
            ch = int(char_pos)
        except Exception:
            return None
        spans = self._spans_for_line()
        for c, (s, e) in spans.items():
            if s <= ch < e:
                return c
        return None

    def _set_main_selected_cell(self, line: int | None, col: int | None):
        try:
            ln = int(line) if line is not None else None
            cc = int(col) if col is not None else None
        except Exception:
            ln = None
            cc = None
        if ln is None or cc is None or ln < 1 or cc < 1:
            self._main_sel_line = None
            self._main_sel_col = None
            return
        self._main_sel_line = ln
        self._main_sel_col = cc

    def _apply_main_selected_cell_highlight(self):
        # Remove previously applied selcell highlight in O(1).
        prev_line = self._applied_main_sel_line
        prev_col = self._applied_main_sel_col
        if prev_line and prev_col:
            try:
                spans_prev = self._spans_for_line()
                if prev_col in spans_prev:
                    s0, e0 = spans_prev[prev_col]
                    for t in (self.left, self.base, self.right):
                        t.tag_remove("selcell", f"{prev_line}.{s0}", f"{prev_line}.{e0}")
            except Exception:
                pass
        self._applied_main_sel_line = None
        self._applied_main_sel_col = None

        line = self._main_sel_line
        col = self._main_sel_col
        if not line or not col:
            return
        if not (1 <= int(line) <= max(1, len(self.display_rows))):
            return

        try:
            spans = self._spans_for_line()
            if col not in spans:
                return
            s, e = spans[col]
            pair = self._pair_for_line(int(line))
            ra = self._row_for_side(pair, "A")
            rb = self._row_for_side(pair, "B")
            base_r = self._row_for_side(pair, "BASE") if self._is_three_way_enabled() else None
            if ra is not None:
                self.left.tag_add("selcell", f"{line}.{s}", f"{line}.{e}")
            if self._is_three_way_enabled() and base_r is not None:
                self.base.tag_add("selcell", f"{line}.{s}", f"{line}.{e}")
            if rb is not None:
                self.right.tag_add("selcell", f"{line}.{s}", f"{line}.{e}")
            self._applied_main_sel_line = int(line)
            self._applied_main_sel_col = int(col)
        except Exception:
            pass

    @staticmethod
    def _row_for_side(pair, side: str) -> int | None:
        if not pair:
            return None
        if side == "A":
            return pair[0]
        if side == "BASE":
            # Base must map to its own aligned row identity (A-side index only).
            # Falling back to B causes duplicated base row numbers in gap regions.
            return pair[0]
        return pair[1]

    def _select_from_widget(self, w: tk.Text, event):
        # Set insert mark to the clicked position so follow-up actions can use it.
        x_before = 0.0
        try:
            x_before = float((self.left.xview() or (0.0, 1.0))[0])
        except Exception:
            x_before = 0.0
        try:
            idx = w.index(f"@{event.x},{event.y}")
            w.mark_set("insert", idx)
        except Exception:
            idx = None

        line = self._widget_line(w)
        col = 0
        try:
            if idx is not None:
                col = int(str(idx).split(".")[1])
        except Exception:
            col = 0

        # Resolve clicked character position to display column index.
        hit_col = self._col_from_char(col)
        self._set_main_selected_cell(line, hit_col)

        # Mirror main-pane cell selection into C区 selected-cell hint.
        # 2-way: A->line1, B->line2
        # 3-way: BASE->line1, A->line2, B->line3
        c_line = None
        if hit_col is not None and int(hit_col) > 0:
            side = self._side_for_widget(w)
            if self._is_three_way_enabled():
                if side == "BASE":
                    c_line = 1
                elif side == "A":
                    c_line = 2
                else:
                    c_line = 3
            else:
                c_line = 1 if side in ("A", "BASE") else 2
        self._cursor_cmp_sel_col = int(hit_col) if c_line is not None else None
        self._cursor_cmp_sel_line = int(c_line) if c_line is not None else None

        # Keep both panes aligned for the cursor compare block:
        # when user clicks either side, keep row+column insert marks consistent.
        for other in (self.left, self.base, self.right):
            if other is w:
                continue
            try:
                other.mark_set("insert", f"{line}.{max(0, col)}")
            except Exception:
                pass

        self._highlight_selected_line(line)
        pair = self._pair_for_line(line)
        self.selected_pair_idx = self._pair_idx_for_line(line)
        self.selected_excel_row_a = self._row_for_side(pair, "A")
        self.selected_excel_row_b = self._row_for_side(pair, "B")
        self.selected_excel_row = self.selected_excel_row_a or self.selected_excel_row_b
        # No button state updates (performance): buttons are always visible and logic no-ops when no diff.

        self._update_cursor_lines()
        self._update_diff_nav_state()

        try:
            x_after = float((self.left.xview() or (0.0, 1.0))[0])
            _dlog(f"select_from_widget xview before={x_before:.6f} after={x_after:.6f} line={line} col={col}")
        except Exception:
            pass

    def _select_line(self, line: int):
        if line < 1:
            return
        for w in (self.left, self.base, self.right):
            try:
                w.mark_set("insert", f"{line}.0")
            except Exception:
                pass
        self._highlight_selected_line(line)
        pair = self._pair_for_line(line)
        self.selected_pair_idx = self._pair_idx_for_line(line)
        self.selected_excel_row_a = self._row_for_side(pair, "A")
        self.selected_excel_row_b = self._row_for_side(pair, "B")
        self.selected_excel_row = self.selected_excel_row_a or self.selected_excel_row_b
        self._set_main_selected_cell(line, None)
        self._update_cursor_lines()
        self._update_diff_nav_state()

    def _on_row_header_click(self, w: tk.Text, event, direction: str):
        try:
            idx = w.index(f"@{event.x},{event.y}")
            line = int(str(idx).split(".")[0])
        except Exception:
            line = 1
        self._clear_row_header_hover(w)
        self._select_line(line)

        # 3-way UX: mine row-header is selection-only (no overwrite action).
        if self._is_three_way_enabled() and w is self.left_ln:
            return "break"

        self._copy_selected_row(direction, row_header=True)
        return "break"

    def _row_header_side(self, w: tk.Text) -> str:
        if w is self.left_ln:
            return "A"
        if w is self.base_ln:
            return "BASE"
        return "B"

    def _set_row_header_text(self, w: tk.Text, line: int, txt: str):
        try:
            # Preserve existing row-header visual tags (e.g. diffrow background)
            # when replacing text for hover arrow rendering.
            keep_tags = []
            try:
                keep_tags = list(w.tag_names(f"{line}.0"))
            except Exception:
                keep_tags = []

            w.configure(state="normal")
            w.delete(f"{line}.0", f"{line}.end")
            w.insert(f"{line}.0", txt)
            for tag in keep_tags:
                try:
                    w.tag_add(tag, f"{line}.0", f"{line}.end")
                except Exception:
                    pass
            w.configure(state="disabled")
        except Exception:
            pass

    def _clear_row_header_hover(self, w: tk.Text):
        line = getattr(self, "_hover_ln_line_left", None) if w is self.left_ln else (
            getattr(self, "_hover_ln_line_mid", None) if w is self.base_ln else getattr(self, "_hover_ln_line_right", None)
        )
        if line is None:
            return
        pair_idx = self._pair_idx_for_line(line)
        if pair_idx is None:
            return
        rn_w = max(3, len(str(self.max_row)))
        side = self._row_header_side(w)
        txt = self._row_label_for_pair_idx(pair_idx, side).rjust(rn_w)
        self._set_row_header_text(w, line, txt)
        try:
            w.configure(cursor="arrow")
        except Exception:
            pass
        if w is self.left_ln:
            self._hover_ln_line_left = None
        elif w is self.base_ln:
            self._hover_ln_line_mid = None
        else:
            self._hover_ln_line_right = None

    def _on_row_header_hover(self, w: tk.Text, event, direction: str):
        try:
            idx = w.index(f"@{event.x},{event.y}")
            line = int(str(idx).split(".")[0])
        except Exception:
            self._clear_row_header_hover(w)
            return
        if not (1 <= line <= len(self.display_rows)):
            self._clear_row_header_hover(w)
            return

        # 3-way UX: mine row-header should not display action arrow.
        if self._is_three_way_enabled() and w is self.left_ln:
            self._clear_row_header_hover(w)
            try:
                w.configure(cursor="arrow")
            except Exception:
                pass
            return

        pair = self._pair_for_line(line)
        pair_idx = self._pair_idx_for_line(line)
        side = self._row_header_side(w)
        r = self._row_for_side(pair, side)
        cols = self.pair_diff_cols.get(pair_idx, set()) if pair_idx is not None else set()
        allow_base_row_action = (self._is_three_way_enabled() and direction == "BASE2A" and w is self.base_ln)
        if r is None or ((not cols) and (not allow_base_row_action)):
            self._clear_row_header_hover(w)
            return
        hover_line = getattr(self, "_hover_ln_line_left", None) if w is self.left_ln else (
            getattr(self, "_hover_ln_line_mid", None) if w is self.base_ln else getattr(self, "_hover_ln_line_right", None)
        )
        if hover_line == line:
            return
        self._clear_row_header_hover(w)
        rn_w = max(3, len(str(self.max_row)))
        arrow = _ROW_ARROW_RIGHT if direction in ("A2B", "MINE2A", "BASE2A") else _ROW_ARROW_LEFT
        self._set_row_header_text(w, line, arrow.rjust(rn_w))
        try:
            w.configure(cursor="hand2")
        except Exception:
            pass
        if w is self.left_ln:
            self._hover_ln_line_left = line
        elif w is self.base_ln:
            self._hover_ln_line_mid = line
        else:
            self._hover_ln_line_right = line

    def _hide_cell_tooltip(self):
        tip = getattr(self, "_cell_tip_win", None)
        if tip is not None:
            try:
                tip.destroy()
            except Exception:
                pass
        self._cell_tip_win = None
        self._cell_tip_key = None

    def _show_cell_tooltip(self, text: str, x_root: int, y_root: int, key):
        if not text:
            self._hide_cell_tooltip()
            return
        if getattr(self, "_cell_tip_key", None) == key and getattr(self, "_cell_tip_win", None) is not None:
            try:
                self._cell_tip_win.geometry(f"+{x_root + 14}+{y_root + 18}")
            except Exception:
                pass
            return
        self._hide_cell_tooltip()
        try:
            tip = tk.Toplevel(self.root)
            tip.wm_overrideredirect(True)
            tip.wm_geometry(f"+{x_root + 14}+{y_root + 18}")
            lbl = tk.Label(tip, text=text, justify="left", relief="solid", borderwidth=1, bg="#fffbe6", fg="#222", font=("Consolas", 10))
            self._cell_tip_win = tip
            self._cell_tip_key = key
            lbl.pack(ipadx=4, ipady=2)
        except Exception:
            self._hide_cell_tooltip()

    def _on_cell_hover_tooltip(self, w: tk.Text, event, side: str):
        try:
            idx = w.index(f"@{event.x},{event.y}")
            line = int(idx.split(".")[0])
            col_char = int(idx.split(".")[1])
        except Exception:
            self._hide_cell_tooltip()
            return
        if not (1 <= line <= len(self.display_rows)):
            self._hide_cell_tooltip()
            return
        pair_idx = self._pair_idx_for_line(line)
        if pair_idx is None or pair_idx >= len(self.row_pairs):
            self._hide_cell_tooltip()
            return
        spans = self._spans_for_line()
        target_col = None
        for c, (s, e) in spans.items():
            if s <= col_char < e:
                target_col = c
                break
        if target_col is None:
            self._hide_cell_tooltip()
            return
        pair = self.row_pairs[pair_idx]
        r = self._row_for_side(pair, side)
        if r is None:
            self._hide_cell_tooltip()
            return
        try:
            if side == "A":
                ws = self.app.ws_a_val(self.sheet)
            elif side == "BASE":
                ws = self.app.ws_base_val(self.sheet)
            else:
                ws = self.app.ws_b_val(self.sheet)
            full_text = _val_to_str(ws.cell(row=r, column=target_col).value)
        except Exception:
            self._hide_cell_tooltip()
            return
        width = max(1, int(self.col_char_widths.get(target_col, 1)))
        if len(full_text) <= width:
            self._hide_cell_tooltip()
            return
        key = (self.sheet, side, r, target_col, full_text)
        self._show_cell_tooltip(full_text, event.x_root, event.y_root, key)

    def _on_click_with_arrow(self, w: tk.Text, event, direction: str):
        # Keep horizontal position stable on click; Tk default Text binding may call see(insert).
        saved_x = 0.0
        try:
            saved_x = float((self.left.xview() or (0.0, 1.0))[0])
        except Exception:
            saved_x = 0.0

        try:
            self._click_trace_seq = int(getattr(self, "_click_trace_seq", 0)) + 1
        except Exception:
            self._click_trace_seq = 1
        seq = self._click_trace_seq
        self._trace_click_until = time.time() + 1.2
        self._log_click_trace_state(f"click_start#{seq}")

        # Select row/column first.
        self._select_from_widget(w, event)
        self._log_click_trace_state(f"after_select#{seq}")

        # Multi-stage guard: catch delayed xview writes from Tk/idle callbacks.
        try:
            self.left.after_idle(lambda sx=saved_x, st=f"idle#{seq}": self._post_click_x_guard(sx, st))
            self.left.after(30, lambda sx=saved_x, st=f"t30#{seq}": self._post_click_x_guard(sx, st))
            self.left.after(120, lambda sx=saved_x, st=f"t120#{seq}": self._post_click_x_guard(sx, st))
        except Exception:
            self._post_click_x_guard(saved_x, f"fallback#{seq}")

        try:
            idx = w.index(f"@{event.x},{event.y}")
            line = int(idx.split(".")[0])
        except Exception:
            return "break"

        if not (1 <= line <= len(self.display_rows)):
            return "break"

        pair = self._pair_for_line(line)
        r = self._row_for_side(pair, self._side_for_widget(w))
        pair_idx = self._pair_idx_for_line(line)
        cols = self.pair_diff_cols.get(pair_idx, set()) if pair_idx is not None else set()
        allow_base_row_action = (
            self._is_three_way_enabled()
            and direction == "BASE2A"
            and w is self.base
        )

        # Row overwrite entry is row-header/button driven; main data-area click should only select.
        # Return break to suppress Tk Text default click handler that can reset xview.
        if (not cols) and (not allow_base_row_action):
            return "break"
        if r is None:
            return "break"
        return "break"

    def _on_hover(self, w: tk.Text, event, direction: str):
        try:
            idx = w.index(f"@{event.x},{event.y}")
            line = int(idx.split(".")[0])
            col = int(idx.split(".")[1])
        except Exception:
            self._clear_hover(w)
            return
        if not (1 <= line <= len(self.display_rows)):
            self._clear_hover(w)
            return
        pair = self._pair_for_line(line)
        r = self._row_for_side(pair, self._side_for_widget(w))
        pair_idx = self._pair_idx_for_line(line)
        cols = self.pair_diff_cols.get(pair_idx, set()) if pair_idx is not None else set()
        allow_base_row_action = (
            self._is_three_way_enabled()
            and direction == "BASE2A"
            and w is self.base
        )
        if (not cols) and (not allow_base_row_action):
            self._clear_hover(w)
            return
        if r is None:
            self._clear_hover(w)
            return
        self._show_hover_arrow(w, line, r, direction)

    def _clear_hover(self, w: tk.Text):
        if w is self.left:
            line = self._hover_line_left
        elif w is self.base:
            line = self._hover_line_mid
        else:
            line = self._hover_line_right
        if line is None:
            return
        self._restore_rownum(w, line)
        try:
            if w is self.left:
                w.configure(cursor=self._left_cursor_default)
            elif w is self.base:
                w.configure(cursor=self._mid_cursor_default)
            else:
                w.configure(cursor=self._right_cursor_default)
        except Exception:
            pass
        if w is self.left:
            self._hover_line_left = None
        elif w is self.base:
            self._hover_line_mid = None
        else:
            self._hover_line_right = None

    def _show_hover_arrow(self, w: tk.Text, line: int, r: int, direction: str):
        if w is self.left:
            if self._hover_line_left == line:
                return
        elif w is self.base:
            if self._hover_line_mid == line:
                return
        else:
            if self._hover_line_right == line:
                return
        # restore previous
        self._clear_hover(w)
        self._replace_rownum_with_arrow(w, line, r, direction)
        try:
            w.configure(cursor="hand2")
        except Exception:
            pass
        if w is self.left:
            self._hover_line_left = line
        elif w is self.base:
            self._hover_line_mid = line
        else:
            self._hover_line_right = line

    def _replace_rownum_with_arrow(self, w: tk.Text, line: int, r: int, direction: str):
        if direction == "B2A":
            arrow = _ROW_ARROW_LEFT
        else:
            arrow = _ROW_ARROW_RIGHT
        rownum = str(r)
        new_label = arrow + (" " * max(0, len(rownum) - 1))
        start = f"{line}.0"
        end = f"{line}.{len(rownum)}"
        try:
            w.delete(start, end)
            w.insert(start, new_label)
        except Exception:
            pass

    def _restore_rownum(self, w: tk.Text, line: int):
        if not (1 <= line <= len(self.display_rows)):
            return
        pair = self._pair_for_line(line)
        r = self._row_for_side(pair, self._side_for_widget(w))
        if r is None:
            return
        rownum = str(r)
        start = f"{line}.0"
        end = f"{line}.{len(rownum)}"
        try:
            w.delete(start, end)
            w.insert(start, rownum)
        except Exception:
            pass

    def _highlight_selected_line(self, line: int):
        # Remove highlight only from the previously selected line (O(1))
        if self._last_selected_line is not None:
            prev = self._last_selected_line
            for t in (self.left, self.base, self.right):
                t.tag_remove("selrow", f"{prev}.0", f"{prev}.end")
        for t in (self.left, self.base, self.right):
            t.tag_add("selrow", f"{line}.0", f"{line}.end")
        self._last_selected_line = line

    def _capture_view_anchor(self):
        """Capture viewport and selection to restore after heavy refresh."""
        first = 0.0
        x_main = 0.0
        x_c = 0.0
        line = 1
        col = 0
        pair_idx = self.selected_pair_idx
        row_a = self.selected_excel_row_a
        row_b = self.selected_excel_row_b
        try:
            first = float((self.left.yview() or (0.0, 1.0))[0])
        except Exception:
            first = 0.0
        try:
            x_main = float((self.left.xview() or (0.0, 1.0))[0])
        except Exception:
            x_main = 0.0
        try:
            x_c = float((self.cursor_cmp.xview() or (0.0, 1.0))[0])
        except Exception:
            x_c = x_main
        try:
            parts = str(self.left.index("insert")).split(".")
            line = int(parts[0])
            col = int(parts[1])
        except Exception:
            line = 1
            col = 0
        return (first, x_main, x_c, line, col, pair_idx, row_a, row_b)

    def _restore_view_anchor(self, anchor):
        if not anchor:
            return
        # backward compatibility with older anchor tuple shape
        if len(anchor) >= 8:
            first, x_main, x_c, line, col, pair_idx, row_a, row_b = anchor
        else:
            first, line, pair_idx, row_a, row_b = anchor
            x_main = 0.0
            x_c = 0.0
            col = 0
        try:
            self.left.yview_moveto(first)
            if self._is_three_way_enabled():
                self.base.yview_moveto(first)
            self.right.yview_moveto(first)
        except Exception:
            pass
        try:
            self._sync_main_x_to_frac(x_main)
            self._sync_c_x_to_frac(x_main if x_c is None else x_c)
        except Exception:
            pass

        target_line = None
        # Prefer relocating by real excel row id; pair indices may shift after rescan.
        try:
            p = None
            if row_a is not None:
                p = self.row_a_to_pair_idx.get(row_a)
            if p is None and row_b is not None:
                p = self.row_b_to_pair_idx.get(row_b)
            if p is not None and p in self.row_to_line:
                target_line = self.row_to_line.get(p)
        except Exception:
            target_line = None
        try:
            if target_line is None and pair_idx is not None and pair_idx in self.row_to_line:
                target_line = self.row_to_line.get(pair_idx)
        except Exception:
            target_line = None

        if target_line is None:
            try:
                max_line = max(1, len(self.display_rows))
            except Exception:
                max_line = 1
            target_line = max(1, min(int(line or 1), max_line))

        idx = f"{target_line}.{max(0, int(col or 0))}"
        for w in (self.left, self.base, self.right):
            try:
                w.mark_set("insert", idx)
            except Exception:
                pass
        try:
            self._highlight_selected_line(target_line)
            self.selected_pair_idx = self._pair_idx_for_line(target_line)
            pair = self._pair_for_line(target_line)
            self.selected_excel_row_a = self._row_for_side(pair, "A")
            self.selected_excel_row_b = self._row_for_side(pair, "B")
            self.selected_excel_row = self.selected_excel_row_a or self.selected_excel_row_b
        except Exception:
            pass

    def _base_to_mine_diff_cols(self, row_a: int | None, row_b: int | None, max_col: int) -> set[int]:
        """Columns that differ between base and mine for the target row in 3-way mode."""
        cols: set[int] = set()
        if not self._is_three_way_enabled():
            return cols
        if not getattr(self.app, "has_base", False):
            return cols
        r = row_a if row_a is not None else row_b
        if r is None:
            return cols
        try:
            ws_a = self.app.ws_a_val(self.sheet)
            ws_base = self.app.ws_base_val(self.sheet)
        except Exception:
            return cols
        for c in range(1, max_col + 1):
            try:
                va = ws_a.cell(row=r, column=c).value
                vb = ws_base.cell(row=r, column=c).value
            except Exception:
                va = None
                vb = None
            if _val_to_str(va) != _val_to_str(vb):
                cols.add(c)
        return cols

    def _update_cursor_lines(self):
        """Update compact row compare block.

        2-way: line1=mine(A), line2=theirs(B)
        3-way: line1=base, line2=mine, line3=theirs
        """
        prev_suppress_c_xsync = bool(getattr(self, "_suppress_c_xsync", False))
        self._suppress_c_xsync = True
        cursor_first = 0.0
        cell_first = 0.0
        try:
            try:
                # Keep C area aligned with main panes by default.
                cursor_first = float((self.left.xview() or (0.0, 1.0))[0])
            except Exception:
                cursor_first = 0.0
            if hasattr(self, "cell_cmp_text"):
                try:
                    cell_first = cursor_first
                except Exception:
                    cell_first = 0.0

            la = int(self.left.index("insert").split(".")[0])
            lb = int(self.right.index("insert").split(".")[0])
            lm = int(self.base.index("insert").split(".")[0])

            a_text = self.left.get(f"{la}.0", f"{la}.end") if la >= 1 else ""
            b_text = self.right.get(f"{lb}.0", f"{lb}.end") if lb >= 1 else ""

            # Determine selected pair (based on line in the view)
            pair_idx = self._pair_idx_for_line(la)
            pair = self.row_pairs[pair_idx] if pair_idx is not None and pair_idx < len(self.row_pairs) else None
            ra = self._row_for_side(pair, "A")
            rb = self._row_for_side(pair, "B")
            diff_cols = self.pair_diff_cols.get(pair_idx, set()) if pair_idx is not None else set()
            base_text = self.base.get(f"{lm}.0", f"{lm}.end") if lm >= 1 else ""
            if self._is_three_way_enabled() and pair_idx is not None:
                base_text = self._build_base_line(pair_idx)

            is_three = self._is_three_way_enabled()
            self._render_cursor_row_headers(pair, is_three)
            # Force strict rendering order:
            # 2-way: mine/theirs
            # 3-way: base/mine/theirs
            self.cursor_cmp.configure(state="normal")
            self.cursor_cmp.delete("1.0", "end")
            if is_three:
                self.cursor_cmp.insert("1.0", f"{base_text}\n{a_text}\n{b_text}")
            else:
                self.cursor_cmp.insert("1.0", f"{a_text}\n{b_text}")

            # Clear & apply base tags
            self.cursor_cmp.tag_remove("a", "1.0", "end")
            self.cursor_cmp.tag_remove("base", "1.0", "end")
            self.cursor_cmp.tag_remove("b", "1.0", "end")
            self.cursor_cmp.tag_remove("missing", "1.0", "end")
            self.cursor_cmp.tag_remove("diffcell", "1.0", "end")
            self.cursor_cmp.tag_remove("cselcell", "1.0", "end")
            if is_three:
                base_r = self._row_for_side(pair, "BASE")
                if base_r is None:
                    self.cursor_cmp.tag_add("missing", "1.0", "1.end")
                else:
                    self.cursor_cmp.tag_add("base", "1.0", "1.end")
                if ra is None:
                    self.cursor_cmp.tag_add("missing", "2.0", "2.end")
                else:
                    self.cursor_cmp.tag_add("a", "2.0", "2.end")
                if rb is None:
                    self.cursor_cmp.tag_add("missing", "3.0", "3.end")
                else:
                    self.cursor_cmp.tag_add("b", "3.0", "3.end")
            else:
                if ra is None:
                    self.cursor_cmp.tag_add("missing", "1.0", "1.end")
                else:
                    self.cursor_cmp.tag_add("a", "1.0", "1.end")
                if rb is None:
                    self.cursor_cmp.tag_add("missing", "2.0", "2.end")
                else:
                    self.cursor_cmp.tag_add("b", "2.0", "2.end")

            spans_a = self._spans_for_line(a_text)
            spans_b = self._spans_for_line(b_text)
            spans_base = self._spans_for_line(base_text) if is_three else {}

            # Cell-level diff highlight
            if diff_cols:
                for c in diff_cols:
                    if is_three and c in spans_base:
                        s, e = spans_base[c]
                        self.cursor_cmp.tag_add("diffcell", f"1.{s}", f"1.{e}")
                    if c in spans_a:
                        s, e = spans_a[c]
                        self.cursor_cmp.tag_add("diffcell", f"{2 if is_three else 1}.{s}", f"{2 if is_three else 1}.{e}")
                    if c in spans_b:
                        s, e = spans_b[c]
                        self.cursor_cmp.tag_add("diffcell", f"{3 if is_three else 2}.{s}", f"{3 if is_three else 2}.{e}")

            # Keep a visible clicked-cell hint in C区 after re-render.
            try:
                sel_col = int(getattr(self, "_cursor_cmp_sel_col", 0) or 0)
                sel_line = int(getattr(self, "_cursor_cmp_sel_line", 0) or 0)
            except Exception:
                sel_col = 0
                sel_line = 0
            if sel_col > 0:
                if is_three:
                    if sel_line == 1 and sel_col in spans_base:
                        s0, e0 = spans_base[sel_col]
                        self.cursor_cmp.tag_add("cselcell", f"1.{s0}", f"1.{e0}")
                    elif sel_line == 2 and sel_col in spans_a:
                        s0, e0 = spans_a[sel_col]
                        self.cursor_cmp.tag_add("cselcell", f"2.{s0}", f"2.{e0}")
                    elif sel_line == 3 and sel_col in spans_b:
                        s0, e0 = spans_b[sel_col]
                        self.cursor_cmp.tag_add("cselcell", f"3.{s0}", f"3.{e0}")
                else:
                    if sel_line == 1 and sel_col in spans_a:
                        s0, e0 = spans_a[sel_col]
                        self.cursor_cmp.tag_add("cselcell", f"1.{s0}", f"1.{e0}")
                    elif sel_line == 2 and sel_col in spans_b:
                        s0, e0 = spans_b[sel_col]
                        self.cursor_cmp.tag_add("cselcell", f"2.{s0}", f"2.{e0}")

            # Apply selected-cell highlight on main panes (A/B and Base in 3-way).
            self._apply_main_selected_cell_highlight()

            self.cursor_cmp.configure(state="disabled")

            # ---- Update C区单元格对齐（可选） ----
            if getattr(self, "_enable_c_cell", False) and hasattr(self, "cell_cmp_text"):
                try:
                    self.cell_cmp_text.configure(state="normal")
                    self.cell_cmp_text.delete("1.0", "end")
                    self.cell_cmp_text.tag_remove("a", "1.0", "end")
                    self.cell_cmp_text.tag_remove("b", "1.0", "end")
                    self.cell_cmp_text.tag_remove("diffcell", "1.0", "end")

                    if pair is not None:
                        ws_a_val = self.app.ws_a_val(self.sheet)
                        ws_b_val = self.app.ws_b_val(self.sheet)
                        show_only_diff = bool(self.c_only_diff_cells.get())
                        cols_to_show = sorted(diff_cols) if show_only_diff else list(range(1, self.max_col + 1))

                        if show_only_diff:
                            parts_a = []
                            parts_b = []
                            widths = []
                            for c in cols_to_show:
                                va = ws_a_val.cell(row=ra, column=c).value if ra is not None else None
                                vb = ws_b_val.cell(row=rb, column=c).value if rb is not None else None
                                a_s = _val_to_str(va)
                                b_s = _val_to_str(vb)
                                parts_a.append(a_s)
                                parts_b.append(b_s)
                                widths.append(max(4, min(max(len(a_s), len(b_s)), _COL_MAX_DISPLAY_WIDTH)))

                            sep = _COL_SEP
                            trail = " \u2502"
                            a_line = sep.join(_format_cell(parts_a[i], widths[i]) for i in range(len(parts_a))) + (trail if parts_a else "")
                            b_line = sep.join(_format_cell(parts_b[i], widths[i]) for i in range(len(parts_b))) + (trail if parts_b else "")
                            self.cell_cmp_text.insert("end", a_line + "\n" + b_line + "\n")

                            self.cell_cmp_text.tag_add("a", "1.0", "1.end")
                            self.cell_cmp_text.tag_add("b", "2.0", "2.end")

                            # All shown columns are diffs (show_only_diff path) — highlight whole lines.
                            if a_line:
                                self.cell_cmp_text.tag_add("diffcell", "1.0", "1.end")
                            if b_line:
                                self.cell_cmp_text.tag_add("diffcell", "2.0", "2.end")
                        else:
                            line_no = 1
                            for c in cols_to_show:
                                va = ws_a_val.cell(row=ra, column=c).value if ra is not None else None
                                vb = ws_b_val.cell(row=rb, column=c).value if rb is not None else None
                                a_s = _val_to_str(va)
                                b_s = _val_to_str(vb)

                                self.cell_cmp_text.insert("end", a_s + "\n")
                                self.cell_cmp_text.insert("end", b_s + "\n")

                                self.cell_cmp_text.tag_add("a", f"{line_no}.0", f"{line_no}.end")
                                self.cell_cmp_text.tag_add("b", f"{line_no+1}.0", f"{line_no+1}.end")

                                if va != vb:
                                    self.cell_cmp_text.tag_add("diffcell", f"{line_no}.0", f"{line_no}.end")
                                    self.cell_cmp_text.tag_add("diffcell", f"{line_no+1}.0", f"{line_no+1}.end")

                                line_no += 2

                    self.cell_cmp_text.configure(state="disabled")
                except Exception:
                    try:
                        self.cell_cmp_text.configure(state="disabled")
                    except Exception:
                        pass

            # Restore C pane horizontal viewport without driving main panes.
            try:
                self.cursor_cmp.xview_moveto(cursor_first)
                cf, cl = self.cursor_cmp.xview()
                self.cursor_hsb.set(cf, cl)
                # Also restore column header: delete+insert resets its xview to 0.
                if getattr(self, "cursor_cmp_colhdr", None) is not None:
                    self.cursor_cmp_colhdr.xview_moveto(cursor_first)
            except Exception:
                pass
            if hasattr(self, "cell_cmp_text"):
                try:
                    self.cell_cmp_text.xview_moveto(cell_first)
                    sf, sl = self.cell_cmp_text.xview()
                    self.cell_cmp_hsb.set(sf, sl)
                except Exception:
                    pass
            self._suppress_c_xsync = prev_suppress_c_xsync
        except Exception:
            try:
                self._suppress_c_xsync = prev_suppress_c_xsync
            except Exception:
                pass
            pass

    def _copy_single_cell_by_pair(self, pair_idx: int, direction: str, c: int):
        try:
            if pair_idx is None or pair_idx >= len(self.row_pairs):
                return
            pair = self.row_pairs[pair_idx]
            ra = self._row_for_side(pair, "A")
            rb = self._row_for_side(pair, "B")
            if direction == "A2B":
                if ra is None or rb is None:
                    return
                src_r, dst_r = ra, rb
            elif direction == "BASE2A":
                if ra is None:
                    return
                src_r = ra
                dst_r = ra
            else:
                if rb is None:
                    return
                src_r = rb
                dst_r = ra if ra is not None else rb

            anchor = self._capture_view_anchor()
            if direction == "A2B":
                old_edit = self.app.ws_b_edit(self.sheet).cell(row=dst_r, column=c).value
                old_val = self.app.ws_b_val(self.sheet).cell(row=dst_r, column=c).value
                v_edit = self.app.ws_a_edit(self.sheet).cell(row=src_r, column=c).value
                v_val = self.app.ws_a_val(self.sheet).cell(row=src_r, column=c).value
                self.app.ws_b_edit(self.sheet).cell(row=dst_r, column=c).value = _choose_edit_value(v_val, v_edit)
                self.app.ws_b_val(self.sheet).cell(row=dst_r, column=c).value = v_val
                self.app.modified_b = True
                self.app.modified_sheets_b.add(self.sheet)
                self.app.push_undo({"sheet": self.sheet, "target": "B", "cells": [(dst_r, c, old_edit, old_val)]})
            elif direction == "BASE2A":
                old_edit = self.app.ws_a_edit(self.sheet).cell(row=dst_r, column=c).value
                old_val = self.app.ws_a_val(self.sheet).cell(row=dst_r, column=c).value
                v_edit = self.app.ws_base_edit(self.sheet).cell(row=src_r, column=c).value
                v_val = self.app.ws_base_val(self.sheet).cell(row=src_r, column=c).value
                new_edit = _choose_edit_value(v_val, v_edit)
                self.app.ws_a_edit(self.sheet).cell(row=dst_r, column=c).value = new_edit
                self.app.ws_a_val(self.sheet).cell(row=dst_r, column=c).value = v_val
                self.app.record_manual_a_cell(self.sheet, dst_r, c, new_edit)
                self.app.modified_a = True
                self.app.modified_sheets_a.add(self.sheet)
                self.app.push_undo({"sheet": self.sheet, "target": "A", "cells": [(dst_r, c, old_edit, old_val)]})
            else:
                old_edit = self.app.ws_a_edit(self.sheet).cell(row=dst_r, column=c).value
                old_val = self.app.ws_a_val(self.sheet).cell(row=dst_r, column=c).value
                v_edit = self.app.ws_b_edit(self.sheet).cell(row=src_r, column=c).value
                v_val = self.app.ws_b_val(self.sheet).cell(row=src_r, column=c).value
                new_edit = _choose_edit_value(v_val, v_edit)
                self.app.ws_a_edit(self.sheet).cell(row=dst_r, column=c).value = new_edit
                self.app.ws_a_val(self.sheet).cell(row=dst_r, column=c).value = v_val
                self.app.record_manual_a_cell(self.sheet, dst_r, c, new_edit)
                self.app.modified_a = True
                self.app.modified_sheets_a.add(self.sheet)
                self.app.push_undo({"sheet": self.sheet, "target": "A", "cells": [(dst_r, c, old_edit, old_val)]})

            touched_r = ra or rb
            if touched_r is not None:
                self.touched_rows.add(touched_r)
            self._invalidate_render_cache()
            if bool(self.only_diff_var.get()) and self.snapshot_only_diff:
                self._recalc_row_diff_and_update(dst_r)
            self.refresh(row_only=dst_r, rescan=False)
            self._restore_view_anchor(anchor)
            self._update_cursor_lines()
        except Exception as e:
            messagebox.showerror("Error", f"C区覆盖失败：\n{e}")

    def _on_cursor_cmp_click(self, event):
        """Single-click in C区: map clicked cell back to current pair/column selection."""
        try:
            idx = self.cursor_cmp.index(f"@{event.x},{event.y}")
            line_no = int(str(idx).split(".")[0])
            char_no = int(str(idx).split(".")[1])
        except Exception:
            return "break"

        try:
            line_text = self.cursor_cmp.get(f"{line_no}.0", f"{line_no}.end")
        except Exception:
            line_text = ""
        spans = self._spans_for_line(line_text)
        hit_col = None
        hit_char = 0
        for c, (s0, e0) in spans.items():
            if s0 <= char_no < e0:
                hit_col = c
                hit_char = s0
                break
        if hit_col is None:
            return "break"

        # Resolve current pair/line in the main panes.
        pair_idx = self.selected_pair_idx
        if pair_idx is None:
            try:
                line_guess = int(self.left.index("insert").split(".")[0])
                pair_idx = self._pair_idx_for_line(line_guess)
            except Exception:
                pair_idx = None

        target_line = None
        try:
            if pair_idx is not None and pair_idx in self.row_to_line:
                target_line = int(self.row_to_line.get(pair_idx))
        except Exception:
            target_line = None
        if target_line is None:
            try:
                target_line = int(self.left.index("insert").split(".")[0])
            except Exception:
                target_line = 1
        try:
            target_line = max(1, min(int(target_line), max(1, len(self.display_rows))))
        except Exception:
            target_line = 1

        saved_x = 0.0
        try:
            saved_x = float((self.left.xview() or (0.0, 1.0))[0])
        except Exception:
            saved_x = 0.0

        target_idx = f"{target_line}.{max(0, int(hit_char))}"
        for w in (self.left, self.base, self.right):
            try:
                w.mark_set("insert", target_idx)
            except Exception:
                pass

        try:
            self._highlight_selected_line(target_line)
            self.selected_pair_idx = self._pair_idx_for_line(target_line)
            pair = self._pair_for_line(target_line)
            self.selected_excel_row_a = self._row_for_side(pair, "A")
            self.selected_excel_row_b = self._row_for_side(pair, "B")
            self.selected_excel_row = self.selected_excel_row_a or self.selected_excel_row_b
        except Exception:
            pass

        # Store C selection so _update_cursor_lines can render visible clicked-cell highlight.
        self._cursor_cmp_sel_col = int(hit_col)
        self._cursor_cmp_sel_line = int(line_no)
        self._set_main_selected_cell(target_line, hit_col)

        # Re-render C block and keep x aligned.
        self._update_cursor_lines()
        try:
            self._sync_main_x_to_frac(saved_x)
            self._sync_c_x_to_frac(saved_x)
        except Exception:
            pass
        self._update_diff_nav_state()
        return "break"

    def _on_cursor_cmp_double_click(self, event):
        try:
            idx = self.cursor_cmp.index(f"@{event.x},{event.y}")
            line_no = int(str(idx).split(".")[0])
            char_no = int(str(idx).split(".")[1])
        except Exception:
            return

        pair_idx = self.selected_pair_idx
        if pair_idx is None:
            try:
                line = int(self.left.index("insert").split(".")[0])
                pair_idx = self._pair_idx_for_line(line)
            except Exception:
                pair_idx = None
        if pair_idx is None or pair_idx >= len(self.row_pairs):
            return

        diff_cols = set(self.pair_diff_cols.get(pair_idx, set()))
        if not diff_cols:
            return

        is_three = self._is_three_way_enabled()
        if is_three:
            if line_no == 1:
                direction = "BASE2A"
            elif line_no == 3:
                direction = "B2A"
            else:
                return
        else:
            if line_no == 2:
                direction = "B2A"
            else:
                return

        line_text = self.cursor_cmp.get(f"{line_no}.0", f"{line_no}.end")
        spans = self._spans_for_line(line_text)
        hit_col = None
        for c, (s, e) in spans.items():
            if s <= char_no < e:
                hit_col = c
                break
        if hit_col is None or hit_col not in diff_cols:
            return
        self._copy_single_cell_by_pair(pair_idx, direction, hit_col)

    def _set_copy_scope_mode(self, mode: str):
        mode_norm = str(mode or "").strip().lower()
        mode_norm = "region" if mode_norm == "region" else "row"
        self._copy_scope_mode = mode_norm
        try:
            self._copy_scope_var.set(mode_norm)
        except Exception:
            pass
        self._refresh_copy_scope_buttons()

    def _refresh_copy_scope_buttons(self):
        mode = getattr(self, "_copy_scope_mode", "row")
        if mode == "region":
            left_text = "使用左侧区域"
            right_text = "使用右侧区域"
        else:
            left_text = "使用左侧行"
            right_text = "使用右侧行"
        try:
            self.use_left_btn.configure(text=left_text)
        except Exception:
            pass
        try:
            self.use_right_btn.configure(text=right_text)
        except Exception:
            pass

    def _run_copy_action_by_mode(self, direction: str):
        mode = getattr(self, "_copy_scope_mode", "row")
        if mode == "region":
            self._copy_selected_region(direction)
        else:
            self._copy_selected_row(direction)

    def _update_merge_buttons_for_row(self, excel_row: int):
        # Buttons are always visible; no UI updates needed.
        return

    # ---------- Diff block navigation ----------
    def _compute_diff_blocks(self):
        """Return list of (start_line, end_line) diff blocks in current view."""
        blocks = []
        start = None
        for line_idx, pair_idx in enumerate(self.display_rows, start=1):
            has = bool(self.pair_diff_cols.get(pair_idx, set()))
            if has and start is None:
                start = line_idx
            elif (not has) and start is not None:
                blocks.append((start, line_idx - 1))
                start = None
        if start is not None:
            blocks.append((start, len(self.display_rows)))
        self._diff_blocks_cache = blocks
        return blocks

    def _current_line(self) -> int:
        try:
            return int(self.left.index("insert").split(".")[0])
        except Exception:
            return 1

    def _current_diff_block_for_line(self, line: int):
        blocks = self._diff_blocks_cache if self._diff_blocks_cache is not None else self._compute_diff_blocks()
        for start, end in blocks:
            if start <= line <= end:
                return (start, end)
        return None

    def _copy_selected_region(self, direction: str):
        """Copy contiguous diff block around current line using diff-cell columns only."""
        try:
            line = self._current_line()
            block = self._current_diff_block_for_line(line)
            if not block:
                return
            start, end = block
            anchor = self._capture_view_anchor()
            # Collect all undo cells into one list so the entire region is a
            # single undo entry regardless of how many rows it spans.
            undo_cells_region: list = []
            undo_target = "A" if direction in ("B2A", "BASE2A") else "B"
            changed_any = False
            for ln in range(start, end + 1):
                if not (1 <= ln <= len(self.display_rows)):
                    continue
                pair_idx = self.display_rows[ln - 1]
                cols = set(self.pair_diff_cols.get(pair_idx, set()))
                if not cols:
                    continue
                # In 3-way mode, "采用Base" is kept as its own button.
                # Left/Right region actions follow left/right semantics only.
                row_changed = self._copy_selected_row(
                    direction,
                    row_header=False,
                    override_pair_idx=pair_idx,
                    override_cols=cols,
                    suppress_refresh=True,
                    _undo_out=undo_cells_region,
                )
                if row_changed:
                    changed_any = True
            if changed_any:
                if undo_cells_region:
                    self.app.push_undo({"sheet": self.sheet, "target": undo_target, "cells": undo_cells_region})
                self._invalidate_render_cache()
                # pair_text_a/b were not updated during suppress_refresh loop;
                # clear them so refresh rebuilds each row from the new cell values.
                self.pair_text_a = {}
                self.pair_text_b = {}
                self.refresh(row_only=None, rescan=False)
                self._restore_view_anchor(anchor)
                self._update_cursor_lines()
        except Exception as e:
            messagebox.showerror("Error", f"覆盖区域失败：\n{e}")

    def _update_diff_nav_state(self):
        blocks = self._compute_diff_blocks()
        if not blocks:
            self.prev_diff_btn.configure(state="disabled")
            self.next_diff_btn.configure(state="disabled")
            return

        cur = self._current_line()
        has_prev = any(b[1] < cur for b in blocks)  # only enable when a block ends before cursor
        has_next = any(b[0] > cur for b in blocks)
        self.prev_diff_btn.configure(state=("normal" if has_prev else "disabled"))
        self.next_diff_btn.configure(state=("normal" if has_next else "disabled"))

    def _goto_block_start(self, start_line: int):
        # Scroll so the line is visible
        try:
            # Preserve horizontal position: see() would reset xview to column 0.
            saved_x = 0.0
            try:
                saved_x = float((self.left.xview() or (0.0, 1.0))[0])
            except Exception:
                pass
            for w in (self.left, self.right):
                w.mark_set("insert", f"{start_line}.0")
                w.see(f"{start_line}.0")
            # Restore horizontal position after see() reset it.
            try:
                self._sync_main_x_to_frac(saved_x)
                self._sync_c_x_to_frac(saved_x)
            except Exception:
                pass
            self._highlight_selected_line(start_line)
            pair = self._pair_for_line(start_line)
            self.selected_pair_idx = self._pair_idx_for_line(start_line)
            self.selected_excel_row_a = self._row_for_side(pair, "A")
            self.selected_excel_row_b = self._row_for_side(pair, "B")
            self.selected_excel_row = self.selected_excel_row_a or self.selected_excel_row_b
            self._set_main_selected_cell(start_line, None)
            self._update_cursor_lines()
        except Exception:
            pass
        self._update_diff_nav_state()

    def _goto_next_diff_block(self):
        blocks = self._compute_diff_blocks()
        cur = self._current_line()
        for start, _end in blocks:
            if start > cur:
                self._goto_block_start(start)
                return
        self._update_diff_nav_state()

    def _goto_prev_diff_block(self):
        blocks = self._compute_diff_blocks()
        cur = self._current_line()
        prev = None
        for start, end in blocks:
            if end < cur:
                prev = start
            elif start >= cur:
                break
        if prev is not None:
            self._goto_block_start(prev)
        self._update_diff_nav_state()

    # ---------- Diff calculation helpers ----------
    def _get_row_values(self, ws, r: int):
        # Fast row read using iter_rows(values_only=True)
        try:
            row = next(ws.iter_rows(min_row=r, max_row=r, min_col=1, max_col=self.max_col, values_only=True))
        except StopIteration:
            row = ()
        if row is None:
            row = ()
        # Ensure length == max_col
        if len(row) < self.max_col:
            row = tuple(row) + (None,) * (self.max_col - len(row))
        return row

    def _show_loading(self):
        """Show a loading placeholder while background diff computation is in progress."""
        try:
            for w in (self.left, self.right):
                w.configure(state="normal")
                w.delete("1.0", "end")
                w.insert("1.0", "计算中...\n")
            for w in (self.left_ln, self.base_ln, self.right_ln):
                w.configure(state="normal")
                w.delete("1.0", "end")
                w.insert("1.0", "\n")
                w.configure(state="disabled")
            self.info.configure(text="正在后台计算差异，请稍候...")
        except Exception:
            pass

    @staticmethod
    def _row_label(r: int | None) -> str:
        return str(r) if r is not None else ""

    def _build_line_from_row_label(self, label: str, row_vals) -> str:
        parts = [_val_to_str(v) for v in row_vals]
        parts = self._gridify_parts(parts)
        return "\t".join(parts)

    def _build_row_and_diff_pair(self, ws_a_val, ws_b_val, ws_a_edit, ws_b_edit, ra: int | None, rb: int | None):
        """Build display lines and diff columns for a row pair.
        col_char_widths must be pre-populated before calling (see _prescan_col_widths)."""
        raw_a = []
        raw_b = []
        cols = set()
        for c in range(1, self.max_col + 1):
            da, db, eq = _cell_display_and_equal_by_row(ws_a_val, ws_b_val, ws_a_edit, ws_b_edit, ra, rb, c)
            raw_a.append(_val_to_str(da))
            raw_b.append(_val_to_str(db))
            if not eq:
                cols.add(c)
        grid_on = self._is_grid_overlay_enabled()
        sep = _COL_SEP if grid_on else "   "
        trail = " \u2502" if grid_on else ""
        cells_a = sep.join(_format_cell(raw_a[i], self.col_char_widths.get(i + 1, 1)) for i in range(len(raw_a))) + trail
        cells_b = sep.join(_format_cell(raw_b[i], self.col_char_widths.get(i + 1, 1)) for i in range(len(raw_b))) + trail
        line_a = cells_a
        line_b = cells_b
        return line_a, line_b, cols

    def _prescan_col_widths(self, ws_a_val, ws_b_val, ws_base_val=None, max_pairs: int = 0):
        """Quick first-pass scan to populate col_char_widths before building formatted lines.
        max_pairs>0 limits scanning to first N pairs (for large sheets)."""
        self.col_char_widths = {}
        self._rownum_display_width = 0
        pairs = self.row_pairs[:max_pairs] if max_pairs > 0 else self.row_pairs
        for ra, rb in pairs:
            r_base = ra if ra is not None else rb
            for c in range(1, self.max_col + 1):
                sa = _val_to_str(ws_a_val.cell(row=ra, column=c).value if ra is not None else None)
                sb = _val_to_str(ws_b_val.cell(row=rb, column=c).value if rb is not None else None)
                w = min(max(len(sa), len(sb)), _COL_MAX_DISPLAY_WIDTH)
                if ws_base_val is not None and r_base is not None:
                    try:
                        sv = _val_to_str(ws_base_val.cell(row=r_base, column=c).value)
                    except Exception:
                        sv = ""
                    w = min(max(w, len(sv)), _COL_MAX_DISPLAY_WIDTH)
                if w > self.col_char_widths.get(c, 0):
                    self.col_char_widths[c] = w
        # Avoid ultra-narrow columns (which render as repeated "...").
        # Keep a readable lower bound after grid on/off toggles.
        for c in range(1, self.max_col + 1):
            self.col_char_widths[c] = max(4, int(self.col_char_widths.get(c, 1)))

    def _build_row_pairs(self, ws_a_val, ws_b_val, force: bool = False):
        # Align rows between A and B to avoid cascading diffs on insert/delete.
        max_row_a = ws_a_val.max_row or 1
        max_row_b = ws_b_val.max_row or 1
        max_row = max(max_row_a, max_row_b)
        if max_row <= 0:
            return []
        if (not force) and max_row >= _ROW_ALIGN_MAX_ROWS:
            # Large-sheet fast path: skip SequenceMatcher and pair rows directly.
            return self._build_row_pairs_direct(max_row_a, max_row_b)

        def _row_sig_list(ws, max_row_local: int):
            # Read all rows in one pass (much faster than per-row iter_rows calls)
            try:
                all_rows = list(ws.iter_rows(
                    min_row=1, max_row=max_row_local,
                    min_col=1, max_col=self.max_col,
                    values_only=True,
                ))
            except Exception:
                all_rows = []
            sigs = []
            for row in all_rows:
                if row is None:
                    row = ()
                sigs.append("\x1f".join(_merge_cmp_value(v) for v in row))
            return sigs

        sig_a = _row_sig_list(ws_a_val, max_row_a)
        sig_b = _row_sig_list(ws_b_val, max_row_b)

        def _sim_score(sa: str, sb: str) -> float:
            if sa == sb:
                return 2.0
            if (not sa) or (not sb):
                return 0.0
            # Non-exact similarity fallback for replace blocks:
            # avoids mapping A[x] -> B[y] when B[z] is clearly closer.
            try:
                return difflib.SequenceMatcher(a=sa, b=sb, autojunk=False).ratio()
            except Exception:
                return 0.0

        sm = difflib.SequenceMatcher(a=sig_a, b=sig_b, autojunk=False)
        pairs: list[tuple[int | None, int | None]] = []
        for tag, i1, i2, j1, j2 in sm.get_opcodes():
            if tag == "equal":
                for i, j in zip(range(i1, i2), range(j1, j2)):
                    pairs.append((i + 1, j + 1))
            elif tag == "replace":
                len_a = i2 - i1
                len_b = j2 - j1
                common = min(len_a, len_b)
                # Choose head/tail mapping by similarity score (not only exact hits).
                head_score = 0.0
                tail_score = 0.0
                for k in range(common):
                    head_score += _sim_score(sig_a[i1 + k], sig_b[j1 + k])
                    tail_score += _sim_score(sig_a[i2 - common + k], sig_b[j2 - common + k])
                # Prefer tail on tie to better match append-heavy editing patterns.
                use_tail = tail_score >= head_score
                if use_tail:
                    extra_a = len_a - common
                    extra_b = len_b - common
                    for k in range(extra_a):
                        pairs.append((i1 + k + 1, None))
                    for k in range(extra_b):
                        pairs.append((None, j1 + k + 1))
                    a_start = i2 - common
                    b_start = j2 - common
                    for k in range(common):
                        pairs.append((a_start + k + 1, b_start + k + 1))
                else:
                    for k in range(common):
                        pairs.append((i1 + k + 1, j1 + k + 1))
                    for k in range(common, len_a):
                        pairs.append((i1 + k + 1, None))
                    for k in range(common, len_b):
                        pairs.append((None, j1 + k + 1))
            elif tag == "delete":
                for i in range(i1, i2):
                    pairs.append((i + 1, None))
            elif tag == "insert":
                for j in range(j1, j2):
                    pairs.append((None, j + 1))
        return pairs

    @staticmethod
    def _build_row_pairs_direct(max_row_a: int, max_row_b: int):
        """Direct row pairing (1:1 by row number), used for very large sheets."""
        max_row = max(max_row_a, max_row_b)
        pairs: list[tuple[int | None, int | None]] = []
        for r in range(1, max_row + 1):
            ra = r if r <= max_row_a else None
            rb = r if r <= max_row_b else None
            pairs.append((ra, rb))
        return pairs

    def _precompute_large_diff_by_blocks(self, ws_a_val, ws_b_val, ws_a_edit, ws_b_edit, max_row_a: int, max_row_b: int):
        """Large-sheet only-diff precompute using tail-first block scan."""
        max_row = max(max_row_a, max_row_b)
        block = _LARGE_SHEET_BLOCK_ROWS
        for block_end in range(max_row, 0, -block):
            block_start = max(1, block_end - block + 1)
            block_len = block_end - block_start + 1

            rows_a = {}
            rows_b = {}
            if block_start <= max_row_a:
                for idx, row in enumerate(
                    ws_a_val.iter_rows(
                        min_row=block_start,
                        max_row=min(block_end, max_row_a),
                        min_col=1,
                        max_col=self.max_col,
                        values_only=True,
                    ),
                    start=block_start,
                ):
                    rows_a[idx] = row or ()
            if block_start <= max_row_b:
                for idx, row in enumerate(
                    ws_b_val.iter_rows(
                        min_row=block_start,
                        max_row=min(block_end, max_row_b),
                        min_col=1,
                        max_col=self.max_col,
                        values_only=True,
                    ),
                    start=block_start,
                ):
                    rows_b[idx] = row or ()

            sig_a = []
            sig_b = []
            for r in range(block_start, block_end + 1):
                row_a = rows_a.get(r, ())
                row_b = rows_b.get(r, ())
                if len(row_a) < self.max_col:
                    row_a = tuple(row_a) + (None,) * (self.max_col - len(row_a))
                if len(row_b) < self.max_col:
                    row_b = tuple(row_b) + (None,) * (self.max_col - len(row_b))
                sig_a.append(tuple(_merge_cmp_value(v) for v in row_a))
                sig_b.append(tuple(_merge_cmp_value(v) for v in row_b))

            if sig_a == sig_b:
                continue

            # Tail-first within changed block (newer rows first).
            for off in range(block_len - 1, -1, -1):
                if sig_a[off] == sig_b[off]:
                    continue
                r = block_start + off
                pair_idx = self.row_a_to_pair_idx.get(r)
                if pair_idx is None:
                    pair_idx = self.row_b_to_pair_idx.get(r)
                if pair_idx is None:
                    continue
                if pair_idx >= len(self.row_pairs):
                    continue
                ra, rb = self.row_pairs[pair_idx]
                line_a, line_b, cols = self._build_row_and_diff_pair(ws_a_val, ws_b_val, ws_a_edit, ws_b_edit, ra, rb)
                self.pair_diff_cols[pair_idx] = cols
                self.pair_text_a[pair_idx] = line_a
                self.pair_text_b[pair_idx] = line_b

    def _build_row_and_diff(self, ws_a_val, ws_b_val, ws_a_edit, ws_b_edit, r: int):
        parts_a = []
        parts_b = []
        cols = set()
        for c in range(1, self.max_col + 1):
            da, db, eq = _cell_display_and_equal(ws_a_val, ws_b_val, ws_a_edit, ws_b_edit, r, c)
            parts_a.append(_val_to_str(da))
            parts_b.append(_val_to_str(db))
            if not eq:
                cols.add(c)
        line_a = str(r) + "\t" + "\t".join(parts_a)
        line_b = str(r) + "\t" + "\t".join(parts_b)
        return line_a, line_b, cols

    def _compute_diff_cols_from_rows(self, row_a, row_b):
        cols = set()
        # row tuples are 0-indexed; cols are 1-indexed
        for i, (va, vb) in enumerate(zip(row_a, row_b), start=1):
            if va != vb:
                cols.add(i)
        return cols

    def _build_line_from_row(self, r: int, row_vals) -> str:
        return str(r) + "\t" + "\t".join(_val_to_str(v) for v in row_vals)

    def _spans_for_line(self, line: str = "") -> dict:
        """Return {colIndex: (start, end)} character positions in the formatted line.
        Computed directly from col_char_widths — no string parsing needed."""
        pos = 0
        spans = {}
        for c in range(1, self.max_col + 1):
            w = max(1, self.col_char_widths.get(c, 1))
            spans[c] = (pos, pos + w)
            pos += w + _COL_SEP_LEN
        return spans

    def _apply_rownum_diff_tag_line(self, line_idx: int, pair_idx: int):
        pass  # Row headers are rendered in dedicated widgets (left_ln/base_ln/right_ln).

    # ---------- Only-diff toggle ----------
    def _toggle_only_diff(self):
        # Snapshot mode confirmed by user: diff rows list is generated once when opening (or manual refresh).
        # Toggling "只看差异" only switches display, without recomputing the diff map.
        try:
            _dlog(f"TOGGLE only_diff={bool(self.only_diff_var.get())} raw={self.only_diff_var.get()} sheet={self.sheet}")
        except Exception:
            pass

        # Always trust current UI value; do not auto-flip state.
        # Auto-flip can invert explicit programmatic toggles and break only-diff filtering.
        cur = int(self.only_diff_var.get())
        self._last_only_diff_value = cur

        # Performance optimization:
        # - For normal sheets with ready diff data, toggling only-diff only changes
        #   display mode and does not need a full rescan.
        # - For large sheets when enabling only-diff, keep the existing rescan path
        #   to build the diff snapshot deterministically.
        need_rescan = False
        if not getattr(self, "_data_ready", False):
            need_rescan = True
        elif bool(cur) and bool(getattr(self, "_is_large_sheet", False)):
            need_rescan = True
        try:
            _dlog(
                f"TOGGLE only_diff refresh: rescan={need_rescan} "
                f"data_ready={getattr(self, '_data_ready', False)} "
                f"large={getattr(self, '_is_large_sheet', False)} "
                f"sheet={self.sheet}"
            )
        except Exception:
            pass

        self._suppress_bg_apply = True
        try:
            self.refresh(row_only=None, rescan=need_rescan)
        finally:
            self._suppress_bg_apply = False  # Fresh data rendered; allow bg applies again.
        self._update_cursor_lines()
        self._update_diff_nav_state()

        # Persist setting (debounced: write 1 s after last toggle to avoid per-keypress I/O)
        try:
            self.app.only_diff_default = int(self.only_diff_var.get())
            if hasattr(self, "_settings_save_id"):
                try:
                    self.frame.after_cancel(self._settings_save_id)
                except Exception:
                    pass
            self._settings_save_id = self.frame.after(1000, self._flush_settings)
        except Exception as e:
            _dlog(f"settings debounce failed: {e}")

    def _toggle_force_align(self):
        """Manual override for large-sheet row pairing accuracy."""
        try:
            self._force_sequence_align = bool(self.force_align_var.get())
            _dlog(f"TOGGLE force_align={self._force_sequence_align} sheet={self.sheet}")
        except Exception:
            self._force_sequence_align = bool(self.force_align_var.get())
        self._suppress_bg_apply = True
        try:
            self.refresh(row_only=None, rescan=True)
        finally:
            self._suppress_bg_apply = False  # Fresh data rendered; allow bg applies again.
        self._update_cursor_lines()
        self._update_diff_nav_state()

    def _flush_settings(self):
        """Debounced settings write: called 1 s after the last only-diff toggle."""
        try:
            os.makedirs(os.path.dirname(_SETTINGS_PATH), exist_ok=True)
            with open(_SETTINGS_PATH, "w", encoding="utf-8") as f:
                json.dump({"only_diff": int(self.only_diff_var.get())}, f, ensure_ascii=False)
        except Exception as e:
            _dlog(f"settings save failed: {e}")

    def _manual_rescan(self):
        self._suppress_bg_apply = True
        try:
            self.refresh(row_only=None, rescan=True)
        finally:
            self._suppress_bg_apply = False  # Fresh data rendered; allow bg applies again.
        self._update_cursor_lines()
        self._update_diff_nav_state()

    # ---------- Merge operations ----------
    def _copy_cell(self, direction: str, event):
        try:
            anchor = self._capture_view_anchor()
            if direction == "A2B":
                src = self.left
            elif direction == "MINE2A":
                src = self.left
            elif direction == "BASE2A":
                src = self.base
            else:
                src = self.right
            idx = src.index(f"@{event.x},{event.y}")
            src.mark_set("insert", idx)
            line = int(idx.split(".")[0])
            col_char = int(idx.split(".")[1])

            if not (1 <= line <= len(self.display_rows)):
                return
            pair = self._pair_for_line(line)
            ra = self._row_for_side(pair, "A")
            rb = self._row_for_side(pair, "B")
            if direction == "A2B":
                if ra is None:
                    return
                if rb is None:
                    # A-only row: insert new row in B then return (full-row insert).
                    pair_idx = self.display_rows[line - 1]
                    self._insert_row_copy(pair_idx, "A2B", ra, False, None, anchor)
                    return
                src_r = ra
                dst_r = rb
            elif direction == "MINE2A":
                if ra is None and rb is None:
                    return
                src_r = ra if ra is not None else rb
                dst_r = ra if ra is not None else rb
            elif direction == "BASE2A":
                if ra is None:
                    return
                src_r = ra
                dst_r = ra
            else:
                if rb is None:
                    return
                if ra is None:
                    # B-only row: insert new row in A then return (full-row insert).
                    pair_idx = self.display_rows[line - 1]
                    self._insert_row_copy(pair_idx, "B2A", rb, False, None, anchor)
                    return
                src_r = rb
                dst_r = ra

            # Use strict character spans to resolve clicked column.
            # Separator clicks must not map to adjacent cells.
            spans = self._spans_for_line()
            c = None
            for col_num, (s0, e0) in spans.items():
                if s0 <= col_char < e0:
                    c = col_num
                    break
            if c is None:
                return
            if c > self.max_col:
                c = self.max_col

            # Merge conflict mode:
            # - "A2B" means keep mine, just mark resolved.
            # - "B2A" means apply theirs to mine, then mark resolved.
            if getattr(self.app, "merge_conflict_mode", False):
                if direction == "A2B":
                    self.app.user_touched_conflicts = True
                    self._resolve_conflict_cell(dst_r, c)
                    return

            if direction == "A2B":
                old_edit = self.app.ws_b_edit(self.sheet).cell(row=dst_r, column=c).value
                old_val = self.app.ws_b_val(self.sheet).cell(row=dst_r, column=c).value
                v_edit = self.app.ws_a_edit(self.sheet).cell(row=src_r, column=c).value
                v_val = self.app.ws_a_val(self.sheet).cell(row=src_r, column=c).value
                # Cached-value mode: always write the cached value
                self.app.ws_b_edit(self.sheet).cell(row=dst_r, column=c).value = _choose_edit_value(v_val, v_edit)
                self.app.ws_b_val(self.sheet).cell(row=dst_r, column=c).value = v_val
                self.app.modified_b = True
                self.app.modified_sheets_b.add(self.sheet)
                self.app.push_undo({"sheet": self.sheet, "target": "B", "cells": [(dst_r, c, old_edit, old_val)]})
            elif direction == "MINE2A":
                # Keep mine value; in conflict mode this means "accept mine".
                if getattr(self.app, "merge_conflict_mode", False):
                    self.app.user_touched_conflicts = True
                    self._resolve_conflict_cell(dst_r, c)
                return
            elif direction == "B2A":
                old_edit = self.app.ws_a_edit(self.sheet).cell(row=dst_r, column=c).value
                old_val = self.app.ws_a_val(self.sheet).cell(row=dst_r, column=c).value
                v_edit = self.app.ws_b_edit(self.sheet).cell(row=src_r, column=c).value
                v_val = self.app.ws_b_val(self.sheet).cell(row=src_r, column=c).value
                new_edit = _choose_edit_value(v_val, v_edit)
                self.app.ws_a_edit(self.sheet).cell(row=dst_r, column=c).value = new_edit
                self.app.ws_a_val(self.sheet).cell(row=dst_r, column=c).value = v_val
                self.app.record_manual_a_cell(self.sheet, dst_r, c, new_edit)
                self.app.modified_a = True
                self.app.modified_sheets_a.add(self.sheet)
                self.app.push_undo({"sheet": self.sheet, "target": "A", "cells": [(dst_r, c, old_edit, old_val)]})
                # In conflict mode, B2A applies theirs; mark conflict resolved.
                if getattr(self.app, "merge_conflict_mode", False):
                    self.app.user_touched_conflicts = True
                    self._resolve_conflict_cell(dst_r, c)
                    return
            else:
                old_edit = self.app.ws_a_edit(self.sheet).cell(row=dst_r, column=c).value
                old_val = self.app.ws_a_val(self.sheet).cell(row=dst_r, column=c).value
                v_edit = self.app.ws_base_edit(self.sheet).cell(row=src_r, column=c).value
                v_val = self.app.ws_base_val(self.sheet).cell(row=src_r, column=c).value
                new_edit = _choose_edit_value(v_val, v_edit)
                self.app.ws_a_edit(self.sheet).cell(row=dst_r, column=c).value = new_edit
                self.app.ws_a_val(self.sheet).cell(row=dst_r, column=c).value = v_val
                self.app.record_manual_a_cell(self.sheet, dst_r, c, new_edit)
                self.app.modified_a = True
                self.app.modified_sheets_a.add(self.sheet)
                self.app.push_undo({"sheet": self.sheet, "target": "A", "cells": [(dst_r, c, old_edit, old_val)]})

            # Mark as touched: keep row visible in "只看差异" even if diffs are resolved.
            pair = self._pair_for_line(line)
            touched_r = self._row_for_side(pair, "A") or self._row_for_side(pair, "B")
            if touched_r is not None:
                self.touched_rows.add(touched_r)
            self._invalidate_render_cache()

            # Minimize flicker: use row-only incremental refresh after overwrite.
            # Full-sheet rescan can be done manually by user when needed.
            if bool(self.only_diff_var.get()) and self.snapshot_only_diff:
                self._recalc_row_diff_and_update(dst_r)
            self.refresh(row_only=dst_r, rescan=False)
            self._restore_view_anchor(anchor)
            self._update_cursor_lines()
        except Exception as e:
            messagebox.showerror("Error", f"覆盖单元格失败：\n{e}")

    # ---------- Row-insert helpers ----------

    def _find_row_insert_position(self, pair_idx: int, side: str) -> int:
        """Return the worksheet row index at which to insert a new row.

        Scans row_pairs[0..pair_idx-1] to find the last non-None row on the
        given side ('A' or 'B').  Returns last_row + 1, or 1 if no prior rows
        exist on that side.
        """
        last_row = 0
        side_idx = 0 if side == "A" else 1
        for i in range(pair_idx):
            r = self.row_pairs[i][side_idx]
            if r is not None:
                last_row = r
        return last_row + 1

    def _insert_row_copy(
        self,
        pair_idx: int,
        direction: str,
        src_row: int,
        suppress_refresh: bool,
        _undo_out,
        anchor,
    ) -> bool:
        """Insert a new row in the destination worksheet and copy src_row data into it.

        Called by _copy_selected_row when the destination side has no paired row:
        - direction="B2A": rb exists, ra is None → insert new row in A.
        - direction="A2B": ra exists, rb is None → insert new row in B.

        In 3-way mode, B2A also inserts an empty row in Base at the same position
        to keep A-Base row-number alignment intact.
        """
        try:
            ws_a_val = self.app.ws_a_val(self.sheet)
            ws_b_val = self.app.ws_b_val(self.sheet)
            ws_a_edit = self.app.ws_a_edit(self.sheet)
            ws_b_edit = self.app.ws_b_edit(self.sheet)

            if direction == "B2A":
                ws_dst_val = ws_a_val
                ws_dst_edit = ws_a_edit
                ws_src_val = ws_b_val
                ws_src_edit = ws_b_edit
                insert_pos = self._find_row_insert_position(pair_idx, "A")
            else:  # A2B
                ws_dst_val = ws_b_val
                ws_dst_edit = ws_b_edit
                ws_src_val = ws_a_val
                ws_src_edit = ws_a_edit
                insert_pos = self._find_row_insert_position(pair_idx, "B")

            # Insert blank row in destination worksheet.
            ws_dst_val.insert_rows(idx=insert_pos)
            ws_dst_edit.insert_rows(idx=insert_pos)

            # When inserting into A, shift any existing manual-edit records whose
            # row numbers are >= insert_pos (they moved up by one in the worksheet).
            if direction == "B2A":
                try:
                    to_shift = {k: v for k, v in self.app.manual_a_cell_ops.items()
                                if k[0] == self.sheet and k[1] >= insert_pos}
                    for k in to_shift:
                        del self.app.manual_a_cell_ops[k]
                    for (s, r, c), v in to_shift.items():
                        self.app.manual_a_cell_ops[(s, r + 1, c)] = v
                except Exception:
                    pass

            # In 3-way mode, inserting into A also inserts an empty row in Base
            # so that A-Base row-number alignment is preserved for BASE2A operations.
            base_inserted = False
            if direction == "B2A" and self._is_three_way_enabled():
                try:
                    ws_bv = self.app.ws_base_val(self.sheet)
                    ws_be = self.app.ws_base_edit(self.sheet)
                    if ws_bv is not None and ws_be is not None:
                        ws_bv.insert_rows(idx=insert_pos)
                        ws_be.insert_rows(idx=insert_pos)
                        base_inserted = True
                except Exception:
                    pass

            # Copy cell values from src_row into the newly inserted row.
            max_col = max(1, ws_src_val.max_column or 1, ws_src_edit.max_column or 1)
            for c in range(1, max_col + 1):
                v_val = ws_src_val.cell(row=src_row, column=c).value
                v_edit = ws_src_edit.cell(row=src_row, column=c).value
                new_edit = _choose_edit_value(v_val, v_edit)
                ws_dst_val.cell(row=insert_pos, column=c).value = v_val
                ws_dst_edit.cell(row=insert_pos, column=c).value = new_edit
                if direction == "B2A":
                    self.app.record_manual_a_cell(self.sheet, insert_pos, c, new_edit)

            if direction == "B2A":
                self.app.modified_a = True
                self.app.modified_sheets_a.add(self.sheet)
            else:
                self.app.modified_b = True
                self.app.modified_sheets_b.add(self.sheet)

            target_tag = "A_INSERT_ROW" if direction == "B2A" else "B_INSERT_ROW"
            self.app.push_undo({
                "sheet": self.sheet,
                "target": target_tag,
                "row": insert_pos,
                "count": 1,
                "base_inserted": base_inserted,
            })

            if not suppress_refresh:
                self._invalidate_render_cache()
                self.refresh(row_only=None, rescan=True)
                if anchor is not None:
                    self._restore_view_anchor(anchor)
                self._update_cursor_lines()
            return True
        except Exception as e:
            messagebox.showerror("Error", f"插入行失败：\n{e}")
            return False

    def _copy_selected_row(
        self,
        direction: str,
        row_header: bool = False,
        override_pair_idx: int | None = None,
        override_cols: set[int] | None = None,
        suppress_refresh: bool = False,
        _undo_out: list | None = None,
    ) -> bool:
        t0 = datetime.now()
        try:
            anchor = None if suppress_refresh else self._capture_view_anchor()
            resolved_only = False
            changed = False
            # use last selected excel row (set on click); fallback to cursor line
            pair_idx = override_pair_idx if override_pair_idx is not None else self.selected_pair_idx
            if pair_idx is None:
                widget = self.left
                try:
                    focus = self.root.focus_get()
                    if focus == self.right:
                        widget = self.right
                except Exception:
                    pass
                try:
                    line = int((widget.index("insert").split(".")[0]))
                except Exception:
                    line = 1
                if not (1 <= line <= len(self.display_rows)):
                    return False
                pair_idx = self.display_rows[line - 1]
            pair = self.row_pairs[pair_idx] if pair_idx is not None and pair_idx < len(self.row_pairs) else None
            ra = self._row_for_side(pair, "A")
            rb = self._row_for_side(pair, "B")
            if direction == "A2B":
                if ra is None:
                    return False
                src_r = ra
                dst_r = rb  # may be None if A-only row; insert path handled in second block
            elif direction == "MINE2A":
                if ra is None and rb is None:
                    return False
                src_r = ra if ra is not None else rb
                dst_r = ra if ra is not None else rb
            elif direction == "BASE2A":
                if ra is None:
                    return False
                src_r = ra
                dst_r = ra
            else:
                if rb is None:
                    return False
                src_r = rb
                dst_r = ra  # may be None if B-only row; insert path handled in second block
            ws_a_val = self.app.ws_a_val(self.sheet)
            ws_b_val = self.app.ws_b_val(self.sheet)
            ws_base_val = self.app.ws_base_val(self.sheet) if getattr(self.app, "has_base", False) else None
            ws_a_edit = self.app.ws_a_edit(self.sheet)
            ws_b_edit = self.app.ws_b_edit(self.sheet)
            ws_base_edit = self.app.ws_base_edit(self.sheet) if getattr(self.app, "has_base", False) else None

            # Default row action overwrites full row range.
            full_max_col = max(
                self.max_col,
                ws_a_val.max_column or 1,
                ws_b_val.max_column or 1,
                (ws_base_val.max_column or 1) if ws_base_val is not None else 1,
                ws_a_edit.max_column or 1,
                ws_b_edit.max_column or 1,
                (ws_base_edit.max_column or 1) if ws_base_edit is not None else 1,
            )
            action_direction = direction
            cols = set(range(1, full_max_col + 1)) if override_cols is None else set(override_cols)

            # 3-way row-header behavior:
            # - Base row number: apply only diff cells to mine
            # - Theirs row number: apply full row to mine
            if row_header and self._is_three_way_enabled() and direction == "BASE2A":
                # Use base-vs-mine diffs for base-row action (not mine-vs-theirs).
                cols = self._base_to_mine_diff_cols(ra, rb, full_max_col)

            # Recompute src/dst based on final action direction.
            if action_direction == "A2B":
                if ra is None:
                    return False
                if rb is None:
                    # A-only row: insert a new row in B at the corresponding position.
                    return self._insert_row_copy(pair_idx, "A2B", ra, suppress_refresh, _undo_out, anchor)
                src_r = ra
                dst_r = rb
            elif action_direction == "MINE2A":
                if ra is None and rb is None:
                    return False
                src_r = ra if ra is not None else rb
                dst_r = ra if ra is not None else rb
            elif action_direction == "BASE2A":
                if ra is None:
                    return False
                src_r = ra
                dst_r = ra
            else:
                if rb is None:
                    return False
                if ra is None:
                    # B-only row: insert a new row in A at the corresponding position.
                    return self._insert_row_copy(pair_idx, "B2A", rb, suppress_refresh, _undo_out, anchor)
                src_r = rb
                dst_r = ra

            # Merge conflict mode:
            # - "A2B" means keep mine, just mark resolved.
            # - "B2A" means apply theirs to mine, then mark resolved.
            if getattr(self.app, "merge_conflict_mode", False):
                rows = self.app.merge_conflict_cells_by_sheet.get(self.sheet) if getattr(self.app, "merge_conflict_cells_by_sheet", None) else None
                conflict_row = ra or rb
                if rows and conflict_row in rows:
                    cols = set(rows.get(conflict_row, set())) if action_direction == "A2B" else cols
                if action_direction == "A2B":
                    self.app.user_touched_conflicts = True
                    self._resolve_conflict_row(conflict_row, cols)
                    resolved_only = True
                    changed = True
                elif action_direction == "MINE2A":
                    self.app.user_touched_conflicts = True
                    self._resolve_conflict_row(conflict_row, cols)
                    resolved_only = True
                    changed = True

            if not cols:
                return False

            if action_direction == "A2B":
                if not resolved_only:
                    undo_cells = []
                    for c in cols:
                        old_edit = ws_b_edit.cell(row=dst_r, column=c).value
                        old_val = ws_b_val.cell(row=dst_r, column=c).value
                        v_edit = ws_a_edit.cell(row=src_r, column=c).value
                        v_val = ws_a_val.cell(row=src_r, column=c).value
                        ws_b_edit.cell(row=dst_r, column=c).value = _choose_edit_value(v_val, v_edit)
                        ws_b_val.cell(row=dst_r, column=c).value = v_val
                        undo_cells.append((dst_r, c, old_edit, old_val))
                    self.app.modified_b = True
                    self.app.modified_sheets_b.add(self.sheet)
                    if undo_cells:
                        changed = True
                        if _undo_out is not None:
                            _undo_out.extend(undo_cells)
                        else:
                            self.app.push_undo({"sheet": self.sheet, "target": "B", "cells": undo_cells})
            elif action_direction == "MINE2A":
                # Keep mine row as-is. In conflict mode, row can still be "changed"
                # by conflict resolution metadata updates above.
                return bool(changed)
            elif action_direction == "B2A":
                undo_cells = []
                for c in cols:
                    old_edit = ws_a_edit.cell(row=dst_r, column=c).value
                    old_val = ws_a_val.cell(row=dst_r, column=c).value
                    v_edit = ws_b_edit.cell(row=src_r, column=c).value
                    v_val = ws_b_val.cell(row=src_r, column=c).value
                    new_edit = _choose_edit_value(v_val, v_edit)
                    ws_a_edit.cell(row=dst_r, column=c).value = new_edit
                    ws_a_val.cell(row=dst_r, column=c).value = v_val
                    self.app.record_manual_a_cell(self.sheet, dst_r, c, new_edit)
                    undo_cells.append((dst_r, c, old_edit, old_val))
                self.app.modified_a = True
                self.app.modified_sheets_a.add(self.sheet)
                if undo_cells:
                    changed = True
                    if _undo_out is not None:
                        _undo_out.extend(undo_cells)
                    else:
                        self.app.push_undo({"sheet": self.sheet, "target": "A", "cells": undo_cells})
                # In conflict mode, B2A applies theirs; mark conflict resolved.
                if getattr(self.app, "merge_conflict_mode", False):
                    self.app.user_touched_conflicts = True
                    self._resolve_conflict_row(conflict_row, cols)
                    resolved_only = True
                    changed = True
            else:
                undo_cells = []
                if ws_base_edit is None or ws_base_val is None:
                    return False
                for c in cols:
                    old_edit = ws_a_edit.cell(row=dst_r, column=c).value
                    old_val = ws_a_val.cell(row=dst_r, column=c).value
                    v_edit = ws_base_edit.cell(row=src_r, column=c).value
                    v_val = ws_base_val.cell(row=src_r, column=c).value
                    new_edit = _choose_edit_value(v_val, v_edit)
                    ws_a_edit.cell(row=dst_r, column=c).value = new_edit
                    ws_a_val.cell(row=dst_r, column=c).value = v_val
                    self.app.record_manual_a_cell(self.sheet, dst_r, c, new_edit)
                    undo_cells.append((dst_r, c, old_edit, old_val))
                self.app.modified_a = True
                self.app.modified_sheets_a.add(self.sheet)
                if undo_cells:
                    changed = True
                    if _undo_out is not None:
                        _undo_out.extend(undo_cells)
                    else:
                        self.app.push_undo({"sheet": self.sheet, "target": "A", "cells": undo_cells})
                if getattr(self.app, "merge_conflict_mode", False):
                    self.app.user_touched_conflicts = True
                    self._resolve_conflict_row(dst_r, cols)
                    resolved_only = True
                    changed = True

            if not changed:
                return False

            # Mark as touched: keep row visible in "只看差异" even if diffs are resolved.
            touched_r = ra or rb
            if touched_r is not None:
                self.touched_rows.add(touched_r)
            if not suppress_refresh:
                self._invalidate_render_cache()
                # Minimize flicker: use row-only incremental refresh after overwrite.
                if bool(self.only_diff_var.get()) and self.snapshot_only_diff:
                    self._recalc_row_diff_and_update(dst_r)
                self.refresh(row_only=dst_r, rescan=False)
                self._restore_view_anchor(anchor)
                self._update_cursor_lines()
            return True
        except Exception as e:
            messagebox.showerror("Error", f"覆盖整行失败：\n{e}")
            return False
        finally:
            try:
                dt = (datetime.now() - t0).total_seconds() * 1000.0
                _dlog(f"OVERWRITE_ROW {self.sheet} dir={direction} ms={dt:.1f}")
            except Exception:
                pass

    def _undo_last_action(self):
        try:
            action = self.app.pop_undo()
            if not action:
                return
            sheet = action.get("sheet")
            target = action.get("target")
            if target == "A_APPEND":
                start_row = action.get("start_row")
                count = action.get("count")
                if not start_row or not count:
                    return
                ws_edit = self.app.ws_a_edit(sheet)
                ws_val = self.app.ws_a_val(sheet)
                ws_edit.delete_rows(start_row, count)
                ws_val.delete_rows(start_row, count)
                try:
                    keys_to_drop = [k for k in self.app.manual_a_cell_ops.keys() if k[0] == sheet and k[1] >= int(start_row)]
                    for k in keys_to_drop:
                        self.app.manual_a_cell_ops.pop(k, None)
                except Exception:
                    pass
                self.app.modified_a = True
                self.app.modified_sheets_a.add(sheet)
                if sheet == self.sheet:
                    self._invalidate_render_cache()
                    self.refresh(row_only=None, rescan=True)
                    self._update_cursor_lines()
                return
            if target in ("A_INSERT_ROW", "B_INSERT_ROW"):
                row = action.get("row", 1)
                count = action.get("count", 1)
                base_inserted = action.get("base_inserted", False)
                if target == "A_INSERT_ROW":
                    ws_edit = self.app.ws_a_edit(sheet)
                    ws_val = self.app.ws_a_val(sheet)
                    ws_edit.delete_rows(row, count)
                    ws_val.delete_rows(row, count)
                    try:
                        row_int = int(row)
                        # Drop manual-edit records for the deleted row(s).
                        to_drop = [k for k in self.app.manual_a_cell_ops
                                   if k[0] == sheet and row_int <= k[1] < row_int + count]
                        for k in to_drop:
                            del self.app.manual_a_cell_ops[k]
                        # Shift records for rows that moved down after the delete.
                        to_shift = {k: v for k, v in self.app.manual_a_cell_ops.items()
                                    if k[0] == sheet and k[1] >= row_int + count}
                        for k in to_shift:
                            del self.app.manual_a_cell_ops[k]
                        for (s, r, c), v in to_shift.items():
                            self.app.manual_a_cell_ops[(s, r - count, c)] = v
                    except Exception:
                        pass
                    self.app.modified_a = True
                    self.app.modified_sheets_a.add(sheet)
                else:
                    ws_edit = self.app.ws_b_edit(sheet)
                    ws_val = self.app.ws_b_val(sheet)
                    ws_edit.delete_rows(row, count)
                    ws_val.delete_rows(row, count)
                    self.app.modified_b = True
                    self.app.modified_sheets_b.add(sheet)
                if base_inserted:
                    try:
                        ws_bv = self.app.ws_base_val(sheet)
                        ws_be = self.app.ws_base_edit(sheet)
                        if ws_bv is not None and ws_be is not None:
                            ws_bv.delete_rows(row, count)
                            ws_be.delete_rows(row, count)
                    except Exception:
                        pass
                if sheet == self.sheet:
                    self._invalidate_render_cache()
                    self.refresh(row_only=None, rescan=True)
                    self._update_cursor_lines()
                return
            cells = action.get("cells", [])
            if not cells:
                return
            if target == "A":
                ws_edit = self.app.ws_a_edit(sheet)
                ws_val = self.app.ws_a_val(sheet)
                self.app.modified_a = True
                self.app.modified_sheets_a.add(sheet)
            else:
                ws_edit = self.app.ws_b_edit(sheet)
                ws_val = self.app.ws_b_val(sheet)
                self.app.modified_b = True
                self.app.modified_sheets_b.add(sheet)
            rows = set()
            for r, c, old_edit, old_val in cells:
                ws_edit.cell(row=r, column=c).value = old_edit
                ws_val.cell(row=r, column=c).value = old_val
                if target == "A":
                    self.app.record_manual_a_cell(sheet, r, c, old_edit)
                rows.add(r)
            # refresh current sheet if applicable
            if sheet == self.sheet:
                for r in rows:
                    self.touched_rows.add(r)
                if self._align_rows_enabled:
                    # Full refresh supersedes per-row work; avoid N redundant partial renders.
                    self.refresh(row_only=None, rescan=True)
                else:
                    for r in rows:
                        if bool(self.only_diff_var.get()) and self.snapshot_only_diff:
                            self._recalc_row_diff_and_update(r)
                        self.refresh(row_only=r, rescan=False)
                self._update_cursor_lines()
        except Exception as e:
            messagebox.showerror("撤销失败", f"撤销操作失败，数据可能未恢复：\n{e}")

    def _resolve_conflict_cell(self, r: int, c: int):
        try:
            if self.app.resolve_conflict_cell(self.sheet, r, c):
                # update view based on updated conflict map
                self.refresh(row_only=None, rescan=False)
                self._update_cursor_lines()
        except Exception:
            pass

    def _resolve_conflict_row(self, r: int, cols):
        try:
            if self.app.resolve_conflict_row(self.sheet, r, cols):
                self.refresh(row_only=None, rescan=False)
                self._update_cursor_lines()
        except Exception:
            pass

    def _refresh_row_text_only(self, r: int):
        """Update the rendered row text for row r without recomputing diff_cols_by_row."""
        try:
            ws_a = self.app.ws_a_val(self.sheet)
            ws_b = self.app.ws_b_val(self.sheet)
            ws_a_edit = self.app.ws_a_edit(self.sheet)
            ws_b_edit = self.app.ws_b_edit(self.sheet)

            self.max_col = max(ws_a.max_column or 1, ws_b.max_column or 1)

            pair_idx = self.row_a_to_pair_idx.get(r)
            if pair_idx is None:
                pair_idx = self.row_b_to_pair_idx.get(r)
            if pair_idx is None:
                return
            ra, rb = self.row_pairs[pair_idx]
            line_a, line_b, _cols = self._build_row_and_diff_pair(ws_a, ws_b, ws_a_edit, ws_b_edit, ra, rb)
            self.pair_text_a[pair_idx] = line_a
            self.pair_text_b[pair_idx] = line_b

            line = self.row_to_line.get(pair_idx)
            if line is None:
                return

            self.left.delete(f"{line}.0", f"{line}.end")
            self._render_base_line(line, pair_idx)
            self.right.delete(f"{line}.0", f"{line}.end")
            self.left.insert(f"{line}.0", self.pair_text_a[pair_idx])
            self.right.insert(f"{line}.0", self.pair_text_b[pair_idx])
            self._render_row_header_line(line, pair_idx)
        except Exception:
            pass

    def _recalc_row_diff_and_update(self, r: int):
        """Recompute diff for row r and update its highlight, without changing the row list (snapshot mode)."""
        try:
            ws_a = self.app.ws_a_val(self.sheet)
            ws_b = self.app.ws_b_val(self.sheet)
            ws_a_edit = self.app.ws_a_edit(self.sheet)
            ws_b_edit = self.app.ws_b_edit(self.sheet)

            self.max_col = max(ws_a.max_column or 1, ws_b.max_column or 1)
            pair_idx = self.row_a_to_pair_idx.get(r)
            if pair_idx is None:
                pair_idx = self.row_b_to_pair_idx.get(r)
            if pair_idx is None:
                return
            if pair_idx >= len(self.row_pairs):
                return
            ra, rb = self.row_pairs[pair_idx]
            line_a, line_b, cols = self._build_row_and_diff_pair(ws_a, ws_b, ws_a_edit, ws_b_edit, ra, rb)
            self.pair_diff_cols[pair_idx] = cols
            self.pair_text_a[pair_idx] = line_a
            self.pair_text_b[pair_idx] = line_b

            line = self.row_to_line.get(pair_idx)
            if line is None:
                # if not visible and touched, rebuild snapshot to include it
                if bool(self.only_diff_var.get()) and (r in self.touched_rows):
                    self.refresh(row_only=None, rescan=False)
                return

            # update text
            self.left.delete(f"{line}.0", f"{line}.end")
            self._render_base_line(line, pair_idx)
            self.right.delete(f"{line}.0", f"{line}.end")
            self.left.insert(f"{line}.0", self.pair_text_a[pair_idx])
            self.right.insert(f"{line}.0", self.pair_text_b[pair_idx])
            self._render_row_header_line(line, pair_idx)

            # update tags for this line
            for w in (self.left, self.base, self.right):
                w.tag_remove("diffrow", f"{line}.0", f"{line}.end")
                w.tag_remove("diffcell", f"{line}.0", f"{line}.end")

            if cols:
                self.left.tag_add("diffrow", f"{line}.0", f"{line}.end")
                self.right.tag_add("diffrow", f"{line}.0", f"{line}.end")

                spans = self._spans_for_line()
                for c in cols:
                    if c in spans:
                        s, e = spans[c]
                        self.left.tag_add("diffcell", f"{line}.{s}", f"{line}.{e}")
                        self.right.tag_add("diffcell", f"{line}.{s}", f"{line}.{e}")
        except Exception:
            pass

    def _invalidate_render_cache(self):
        self._data_version += 1
        self._render_cache.clear()
        self._diff_blocks_cache = None

    def _build_base_line(self, pair_idx: int) -> str:
        if not self._is_three_way_enabled():
            return ""
        if not getattr(self.app, "has_base", False):
            return ""
        if pair_idx >= len(self.row_pairs):
            return ""
        pair = self.row_pairs[pair_idx]
        if not pair:
            return ""
        ra, rb = pair
        r = ra
        if r is None:
            return ""
        try:
            ws_base = self.app.ws_base_val(self.sheet)
        except Exception:
            return ""
        raw = []
        for c in range(1, self.max_col + 1):
            try:
                v = ws_base.cell(row=r, column=c).value
            except Exception:
                v = None
            raw.append(_val_to_str(v))
        grid_on = self._is_grid_overlay_enabled()
        sep = _COL_SEP if grid_on else "   "
        trail = " \u2502" if grid_on else ""
        cells = sep.join(_format_cell(raw[i], self.col_char_widths.get(i + 1, 1)) for i in range(len(raw))) + trail
        return cells

    def _render_base_full(self):
        if not self._is_three_way_enabled():
            try:
                self.base.delete("1.0", "end")
                self.base.tag_remove("selrow", "1.0", "end")
            except Exception:
                pass
            return
        lines = [self._build_base_line(pair_idx) for pair_idx in self.display_rows]
        try:
            self.base.delete("1.0", "end")
            self.base.insert("1.0", "\n".join(lines) + ("\n" if lines else ""))
        except Exception:
            pass

    def _render_base_line(self, line: int, pair_idx: int):
        if not self._is_three_way_enabled():
            return
        txt = self._build_base_line(pair_idx)
        try:
            self.base.delete(f"{line}.0", f"{line}.end")
            self.base.insert(f"{line}.0", txt)
        except Exception:
            pass

    def _row_label_for_pair_idx(self, pair_idx: int, side: str) -> str:
        try:
            pair = self.row_pairs[pair_idx]
            r = self._row_for_side(pair, side)
            return str(r) if r is not None else ""
        except Exception:
            return ""

    def _render_row_headers_full(self):
        rn_w = max(3, len(str(self.max_row)))
        left_lines = []
        base_lines = []
        right_lines = []
        for pidx in self.display_rows:
            left_lines.append(self._row_label_for_pair_idx(pidx, "A").rjust(rn_w))
            base_lines.append(self._row_label_for_pair_idx(pidx, "BASE").rjust(rn_w))
            right_lines.append(self._row_label_for_pair_idx(pidx, "B").rjust(rn_w))
        for w, lines in ((self.left_ln, left_lines), (self.base_ln, base_lines), (self.right_ln, right_lines)):
            try:
                w.configure(state="normal")
                w.delete("1.0", "end")
                w.insert("1.0", "\n".join(lines) + ("\n" if lines else ""))
                w.tag_remove("diffrow", "1.0", "end")
                w.configure(state="disabled")
            except Exception:
                pass

    def _render_row_header_line(self, line: int, pair_idx: int):
        rn_w = max(3, len(str(self.max_row)))
        vals = (
            self._row_label_for_pair_idx(pair_idx, "A").rjust(rn_w),
            self._row_label_for_pair_idx(pair_idx, "BASE").rjust(rn_w),
            self._row_label_for_pair_idx(pair_idx, "B").rjust(rn_w),
        )
        for w, txt in ((self.left_ln, vals[0]), (self.base_ln, vals[1]), (self.right_ln, vals[2])):
            try:
                w.configure(state="normal")
                w.delete(f"{line}.0", f"{line}.end")
                w.insert(f"{line}.0", txt)
                w.tag_remove("diffrow", f"{line}.0", f"{line}.end")
                w.configure(state="disabled")
            except Exception:
                pass

    def _build_col_header_line(self) -> str:
        if self.max_col <= 0:
            return ""
        sep = _COL_SEP if self._is_grid_overlay_enabled() else "   "
        trail = " │" if self._is_grid_overlay_enabled() else ""
        parts = []
        for c in range(1, self.max_col + 1):
            label = get_column_letter(c)
            parts.append(_format_cell(label, self.col_char_widths.get(c, 1)))
        return sep.join(parts) + (trail if parts else "")

    def _render_col_headers(self):
        hdr = self._build_col_header_line()
        rn_w = max(3, len(str(self.max_row)))
        corner = "".rjust(rn_w)
        for w in (self.left_corner_hdr, self.base_corner_hdr, self.right_corner_hdr, self.cursor_cmp_corner):
            try:
                w.configure(state="normal")
                w.delete("1.0", "end")
                w.insert("1.0", corner)
                w.configure(state="disabled")
            except Exception:
                pass
        for w in (self.left_colhdr, self.base_colhdr, self.right_colhdr, self.cursor_cmp_colhdr):
            try:
                w.configure(state="normal")
                w.delete("1.0", "end")
                w.insert("1.0", hdr)
                w.configure(state="disabled")
            except Exception:
                pass

    def _render_cursor_row_headers(self, pair, is_three: bool):
        if not hasattr(self, "cursor_cmp_ln"):
            return
        rn_w = max(3, len(str(self.max_row)))
        ra = self._row_for_side(pair, "A") if pair else None
        rb = self._row_for_side(pair, "B") if pair else None
        rows = []
        if is_three:
            base_r = self._row_for_side(pair, "BASE") if pair else None
            rows.append(str(base_r) if base_r is not None else "")
            rows.append(str(ra) if ra is not None else "")
        else:
            rows.append(str(ra) if ra is not None else "")
        rows.append(str(rb) if rb is not None else "")
        rows_txt = [r.rjust(rn_w) for r in rows]
        try:
            self.cursor_cmp_ln.configure(state="normal")
            self.cursor_cmp_ln.delete("1.0", "end")
            self.cursor_cmp_ln.insert("1.0", "\n".join(rows_txt) + ("\n" if rows_txt else ""))
            self.cursor_cmp_ln.configure(state="disabled")
        except Exception:
            pass

    # ---------- Rendering ----------
    def _load_all_rows(self):
        self._full_render = True
        self.refresh(row_only=None, rescan=False)

    def _append_rows(self, new_rows: list[int]):
        if not new_rows:
            return
        ws_a = self.app.ws_a_val(self.sheet)
        ws_b = self.app.ws_b_val(self.sheet)
        try:
            wb_a_edit = getattr(self.app, "_wb_a_edit", None)
            ws_a_edit = wb_a_edit[self.sheet] if wb_a_edit is not None else None
        except Exception:
            ws_a_edit = None
        try:
            wb_b_edit = getattr(self.app, "_wb_b_edit", None)
            ws_b_edit = wb_b_edit[self.sheet] if wb_b_edit is not None else None
        except Exception:
            ws_b_edit = None

        # Accuracy guard:
        # cached-value mode may read formula cells as None when cache is missing.
        # Load edit workbooks so formula-text fallback can suppress false diffs.
        if _USE_CACHED_VALUES_ONLY and (ws_a_edit is None or ws_b_edit is None):
            try:
                self.app._ensure_edit_loaded()
                ws_a_edit = self.app.ws_a_edit(self.sheet)
                ws_b_edit = self.app.ws_b_edit(self.sheet)
                _dlog(f"formula fallback enabled: loaded edit wbs for sheet={self.sheet}")
            except Exception as e:
                _dlog(f"formula fallback load failed: sheet={self.sheet} err={e}")

        start_line = len(self.display_rows) + 1
        # Preserve current scroll position to avoid jumps
        try:
            first, _last = self.left.yview()
        except Exception:
            first = None

        # _spans_for_line() result depends only on col_char_widths (invariant during this call)
        _col_spans = self._spans_for_line()

        for idx, pair_idx in enumerate(new_rows, start=0):
            if pair_idx not in self.pair_text_a or pair_idx not in self.pair_text_b:
                ra, rb = self.row_pairs[pair_idx]
                line_a, line_b, cols = self._build_row_and_diff_pair(ws_a, ws_b, ws_a_edit, ws_b_edit, ra, rb)
                self.pair_diff_cols[pair_idx] = cols
                self.pair_text_a[pair_idx] = line_a
                self.pair_text_b[pair_idx] = line_b
            else:
                cols = self.pair_diff_cols.get(pair_idx, set())
                line_a = self.pair_text_a.get(pair_idx, "")
                line_b = self.pair_text_b.get(pair_idx, "")

            line_no = start_line + idx
            self.left.insert("end", line_a + "\n")
            if self._is_three_way_enabled():
                self.base.insert("end", self._build_base_line(pair_idx) + "\n")
            self.right.insert("end", line_b + "\n")
            self._render_row_header_line(line_no, pair_idx)

            if cols:
                self._display_diff_row_count += 1
                self.left.tag_add("diffrow", f"{line_no}.0", f"{line_no}.end")
                self.base.tag_add("diffrow", f"{line_no}.0", f"{line_no}.end")
                self.right.tag_add("diffrow", f"{line_no}.0", f"{line_no}.end")
                for c in cols:
                    if c in _col_spans:
                        s, e = _col_spans[c]
                        self.left.tag_add("diffcell", f"{line_no}.{s}", f"{line_no}.{e}")
                        self.right.tag_add("diffcell", f"{line_no}.{s}", f"{line_no}.{e}")

        self.display_rows.extend(new_rows)
        for i, pair_idx in enumerate(new_rows, start=start_line):
            self.row_to_line[pair_idx] = i

        # row numbers are rendered in dedicated row-header widgets

        mode = "只看差异" if self.only_diff_var.get() else "全量"
        total_rows = len(self.row_pairs) if self.row_pairs else self.max_row
        self.info.configure(text=f"{mode} | RowsShown: {len(self.display_rows)} / {total_rows}   Cols: {self.max_col}   DiffRows: {self._display_diff_row_count}")

        if first is not None:
            try:
                self.left.yview_moveto(first)
                if self._is_three_way_enabled():
                    self.base.yview_moveto(first)
                self.right.yview_moveto(first)
            except Exception:
                pass

        self._invalidate_render_cache()

    def _maybe_load_more_rows(self, last_fraction: float):
        if not _FAST_OPEN_ENABLED:
            return
        try:
            last_fraction = float(last_fraction)
        except Exception:
            return
        if self._full_render:
            return
        if bool(self.only_diff_var.get()):
            return
        if getattr(self.app, "merge_conflict_mode", False):
            return
        # Only for full-list mode (not only-diff or conflict-only)
        if not self._full_display_rows:
            return
        if last_fraction < 0.98:
            return
        if len(self.display_rows) >= len(self._full_display_rows):
            return
        old_limit = len(self.display_rows)
        new_limit = min(len(self._full_display_rows), self._render_limit + _FAST_RENDER_BATCH)
        self._render_limit = new_limit
        new_rows = self._full_display_rows[old_limit:new_limit]
        self._append_rows(new_rows)

    def refresh(self, row_only: int | None, rescan: bool):
        _dlog(f"REFRESH sheet={self.sheet} row_only={row_only} rescan={rescan} only_diff={bool(self.only_diff_var.get())} raw={self.only_diff_var.get()}")
        if rescan and (not self._full_render):
            self._render_limit = _FAST_RENDER_ROW_LIMIT
        conflict_cells_by_row = None
        if getattr(self.app, "merge_conflict_mode", False):
            rows_map = getattr(self.app, "merge_conflict_cells_by_sheet", None)
            conflict_cells_by_row = rows_map.get(self.sheet) if rows_map else None
        ws_a = self.app.ws_a_val(self.sheet)
        ws_b = self.app.ws_b_val(self.sheet)
        # Non-blocking edit sheets: use loaded edit workbook if already available.
        # Do not trigger expensive load_workbook() during pure view refresh/toggle.
        try:
            wb_a_edit = getattr(self.app, "_wb_a_edit", None)
            ws_a_edit = wb_a_edit[self.sheet] if wb_a_edit is not None else None
        except Exception:
            ws_a_edit = None
        try:
            wb_b_edit = getattr(self.app, "_wb_b_edit", None)
            ws_b_edit = wb_b_edit[self.sheet] if wb_b_edit is not None else None
        except Exception:
            ws_b_edit = None

        if rescan or (not self._bounds_checked):
            a_r, a_c = _effective_bounds(ws_a)
            b_r, b_c = _effective_bounds(ws_b)
            self.max_row = max(a_r, b_r)
            self.max_col = max(a_c, b_c)
            self._bounds_checked = True
            self._is_large_sheet = self.max_row >= _LARGE_SHEET_ROW_THRESHOLD
            if self._is_large_sheet:
                self._prefer_only_diff_when_ready = True

        # Full rescan diff map + cache row text if requested
        # Use _data_ready flag instead of checking pair_diff_cols emptiness:
        # pair_diff_cols can legitimately be empty (no diffs found) while still being valid data.
        if rescan or not self._data_ready:
            if not rescan:
                # Data not yet ready (background computation still running).
                # Skip this call; _apply_sheet_cache will call refresh() when done.
                return
            self.pair_diff_cols = {}
            self.pair_text_a = {}
            self.pair_text_b = {}
            self.row_a_to_pair_idx = {}
            self.row_b_to_pair_idx = {}
            self._diff_partial = False

            if conflict_cells_by_row is not None:
                # Conflict-only fast path: avoid full-sheet diff scan.
                self._align_rows_enabled = False
                conflict_rows = sorted(conflict_cells_by_row.keys())
                self.row_pairs = [(r, r) for r in conflict_rows]
                for idx, (ra, rb) in enumerate(self.row_pairs):
                    if ra is not None:
                        self.row_a_to_pair_idx[ra] = idx
                    if rb is not None:
                        self.row_b_to_pair_idx[rb] = idx
                # Pre-scan column widths before formatting
                ws_base_val_opt = None
                if getattr(self.app, "has_base", False):
                    try:
                        ws_base_val_opt = self.app.ws_base_val(self.sheet)
                    except Exception:
                        pass
                self._prescan_col_widths(ws_a, ws_b, ws_base_val_opt)
                for idx, (ra, rb) in enumerate(self.row_pairs):
                    line_a, line_b, _cols = self._build_row_and_diff_pair(ws_a, ws_b, ws_a_edit, ws_b_edit, ra, rb)
                    cols = set(conflict_cells_by_row.get(ra, set())) if ra is not None else set()
                    self.pair_diff_cols[idx] = cols
                    self.pair_text_a[idx] = line_a
                    self.pair_text_b[idx] = line_b
            else:
                max_row_a = ws_a.max_row or 1
                max_row_b = ws_b.max_row or 1

                force_align = bool(getattr(self, "_force_sequence_align", False))

                # Large sheets: skip expensive row-alignment on open unless user forces SM.
                if (self.max_row >= _ROW_ALIGN_MAX_ROWS) and (not force_align):
                    self._align_rows_enabled = False
                    self.row_pairs = self._build_row_pairs_direct(max_row_a, max_row_b)
                else:
                    self._align_rows_enabled = (not getattr(self.app, "merge_conflict_mode", False))
                    if self._align_rows_enabled:
                        self.row_pairs = self._build_row_pairs(ws_a, ws_b, force=force_align)
                    else:
                        self.row_pairs = self._build_row_pairs_direct(max_row_a, max_row_b)

                for idx, (ra, rb) in enumerate(self.row_pairs):
                    if ra is not None:
                        self.row_a_to_pair_idx[ra] = idx
                    if rb is not None:
                        self.row_b_to_pair_idx[rb] = idx

                # Pre-scan column widths for aligned display before building any formatted lines.
                ws_base_val_opt = None
                if getattr(self.app, "has_base", False):
                    try:
                        ws_base_val_opt = self.app.ws_base_val(self.sheet)
                    except Exception:
                        pass
                _prescan_limit = _FAST_RENDER_ROW_LIMIT if self._is_large_sheet else 0
                self._prescan_col_widths(ws_a, ws_b, ws_base_val_opt, max_pairs=_prescan_limit)

                # Large-sheet strategy:
                # - full mode: lazy row compute (first 200 visible rows only)
                # - only-diff mode: block scan from tail to head (1000 rows/block)
                if self._is_large_sheet and bool(self.only_diff_var.get()):
                    self._precompute_large_diff_by_blocks(ws_a, ws_b, ws_a_edit, ws_b_edit, max_row_a, max_row_b)
                elif not self._is_large_sheet:
                    for idx, (ra, rb) in enumerate(self.row_pairs):
                        line_a, line_b, cols = self._build_row_and_diff_pair(ws_a, ws_b, ws_a_edit, ws_b_edit, ra, rb)
                        self.pair_diff_cols[idx] = cols
                        self.pair_text_a[idx] = line_a
                        self.pair_text_b[idx] = line_b

            self._data_ready = True

        # Build display rows list (pair indices)
        if conflict_cells_by_row is not None:
            # Always show conflict rows only
            rows = []
            for r in sorted(conflict_cells_by_row.keys()):
                idx = self.row_a_to_pair_idx.get(r)
                if idx is not None:
                    rows.append(idx)
            self._full_display_rows = rows
        elif bool(self.only_diff_var.get()):
            if (not self.snapshot_only_diff) or rescan or (row_only is None) or (not self.display_rows):
                # build snapshot: diff rows + touched rows
                rows = [idx for idx, cols in self.pair_diff_cols.items() if cols]
                rows_set = set(rows)
                for r in self.touched_rows:
                    idx = self.row_a_to_pair_idx.get(r)
                    if idx is not None:
                        rows_set.add(idx)
                self._full_display_rows = sorted(rows_set)
            else:
                # snapshot mode: keep existing row list stable
                pass
        else:
            self._full_display_rows = list(range(0, len(self.row_pairs)))

        # Fast render: limit initial rows unless user opted to load all
        if self._full_render or (not _FAST_OPEN_ENABLED):
            self.display_rows = list(self._full_display_rows)
        else:
            # reset render limit if full list shrank
            self._render_limit = min(self._render_limit, len(self._full_display_rows)) if self._full_display_rows else _FAST_RENDER_ROW_LIMIT
            if len(self._full_display_rows) > _FAST_RENDER_ROW_LIMIT and self._render_limit < _FAST_RENDER_ROW_LIMIT:
                self._render_limit = _FAST_RENDER_ROW_LIMIT
            if self._is_large_sheet and rescan:
                self._render_limit = min(_LARGE_SHEET_INITIAL_ROWS, len(self._full_display_rows)) if self._full_display_rows else _LARGE_SHEET_INITIAL_ROWS
            self.display_rows = self._full_display_rows[:self._render_limit]
        _dlog(f"  build display_rows: {len(self.display_rows)} / {self.max_row} (only_diff={bool(self.only_diff_var.get())} raw={self.only_diff_var.get()})")

        # Ensure pair text/diff exists for currently displayed rows (lazy fill)
        if self.display_rows:
            missing = [idx for idx in self.display_rows if idx not in self.pair_text_a or idx not in self.pair_text_b]
            if missing:
                for idx in missing:
                    if idx >= len(self.row_pairs):
                        continue
                    ra, rb = self.row_pairs[idx]
                    line_a, line_b, cols = self._build_row_and_diff_pair(ws_a, ws_b, ws_a_edit, ws_b_edit, ra, rb)
                    self.pair_diff_cols[idx] = cols
                    self.pair_text_a[idx] = line_a
                    self.pair_text_b[idx] = line_b

        self.row_to_line = {r: i + 1 for i, r in enumerate(self.display_rows)}

        if row_only is None:
            try:
                frac = float((self.left.xview() or (0.0, 1.0))[0])
            except Exception:
                frac = 0.0
            try:
                self._render_col_headers()
                # _render_col_headers() resets header xview to 0; restore all panes uniformly.
                self._sync_main_x_to_frac(frac)
                self._sync_c_x_to_frac(frac)
            except Exception:
                pass

        # Partial refresh: update a single excel row if it is visible
        if row_only is not None:
            r = row_only
            pair_idx = self.row_a_to_pair_idx.get(r)
            if pair_idx is None:
                pair_idx = self.row_b_to_pair_idx.get(r)
            if pair_idx is None:
                return
            ra, rb = self.row_pairs[pair_idx]
            # recompute diff cols + cache text for that pair only
            line_a, line_b, cols = self._build_row_and_diff_pair(ws_a, ws_b, ws_a_edit, ws_b_edit, ra, rb)
            self.pair_text_a[pair_idx] = line_a
            self.pair_text_b[pair_idx] = line_b
            if conflict_cells_by_row is not None and ra is not None:
                self.pair_diff_cols[pair_idx] = set(conflict_cells_by_row.get(ra, set()))
            else:
                self.pair_diff_cols[pair_idx] = cols

            # If only-diff enabled, row might need to be added/removed
            if bool(self.only_diff_var.get()):
                visible = pair_idx in self.row_to_line
                has = bool(self.pair_diff_cols[pair_idx])

                # If diffs are resolved but this row was touched, keep it visible as a record.
                keep = (r in self.touched_rows)

                if self.snapshot_only_diff:
                    # Snapshot mode: never auto-remove rows from the list.
                    # If a touched row is not visible (was not in initial snapshot), allow adding it.
                    if (not visible) and keep:
                        self.refresh(row_only=None, rescan=False)
                        return
                else:
                    if visible and (not has) and (not keep):
                        # remove the line
                        line = self.row_to_line[pair_idx]
                        self.left.delete(f"{line}.0", f"{line + 1}.0")
                        if self._is_three_way_enabled():
                            self.base.delete(f"{line}.0", f"{line + 1}.0")
                        self.right.delete(f"{line}.0", f"{line + 1}.0")
                        # rebuild
                        self.refresh(row_only=None, rescan=False)
                        return

                    if (not visible) and (has or keep):
                        # add row: simplest is full rebuild (diff list is small)
                        self.refresh(row_only=None, rescan=False)
                        return

            line = self.row_to_line.get(pair_idx)
            if line is None:
                # not visible
                return

            line_a = self.pair_text_a.get(pair_idx, "")
            line_b = self.pair_text_b.get(pair_idx, "")

            # update text
            self.left.delete(f"{line}.0", f"{line}.end")
            self._render_base_line(line, pair_idx)
            self.right.delete(f"{line}.0", f"{line}.end")
            self.left.insert(f"{line}.0", line_a)
            self.right.insert(f"{line}.0", line_b)

            # clear tags on this line then apply diff highlight (unless touched row resolved)
            for w in (self.left, self.base, self.right):
                w.tag_remove("diffrow", f"{line}.0", f"{line}.end")
                w.tag_remove("diffcell", f"{line}.0", f"{line}.end")

            cols = self.pair_diff_cols.get(pair_idx, set())
            # If this row was touched and has no diffs anymore, keep it visible but don't show diff highlight.
            show_diff = bool(cols)
            if show_diff:
                self.left.tag_add("diffrow", f"{line}.0", f"{line}.end")
                self.base.tag_add("diffrow", f"{line}.0", f"{line}.end")
                self.right.tag_add("diffrow", f"{line}.0", f"{line}.end")

                spans = self._spans_for_line()
                for c in cols:
                    if c in spans:
                        s, e = spans[c]
                        self.left.tag_add("diffcell", f"{line}.{s}", f"{line}.{e}")
                        self.right.tag_add("diffcell", f"{line}.{s}", f"{line}.{e}")
            # keep fast; do not rebuild sheet nav here
            try:
                self._display_diff_row_count = sum(1 for idx in self.display_rows if self.pair_diff_cols.get(idx))
                mode = "只看差异" if self.only_diff_var.get() else "全量"
                total_rows = len(self.row_pairs) if self.row_pairs else self.max_row
                self.info.configure(text=f"{mode} | RowsShown: {len(self.display_rows)} / {total_rows}   Cols: {self.max_col}   DiffRows: {self._display_diff_row_count}")
            except Exception:
                pass
            try:
                self._update_diff_maps()
            except Exception:
                pass
            return

        # Full render (use cache when possible)
        mode_key = "diff" if (conflict_cells_by_row is not None or bool(self.only_diff_var.get())) else "full"
        head = tuple(self.display_rows[:5])
        tail = tuple(self.display_rows[-5:]) if len(self.display_rows) > 5 else tuple(self.display_rows)
        cache_key = (mode_key, self._render_limit, len(self.display_rows), head, tail, self._data_version)
        if row_only is None and (not rescan):
            cached = self._render_cache.get(cache_key)
            if cached is not None:
                text_a, text_b, tag_rows, tag_cells, diff_row_count = cached
                self.left.delete("1.0", "end")
                self.base.delete("1.0", "end")
                self.right.delete("1.0", "end")
                self.left.insert("1.0", text_a)
                self._render_base_full()
                self.right.insert("1.0", text_b)
                self._render_row_headers_full()
                # clear tags
                self.left.tag_remove("diffrow", "1.0", "end")
                self.base.tag_remove("diffrow", "1.0", "end")
                self.right.tag_remove("diffrow", "1.0", "end")
                self.left.tag_remove("diffcell", "1.0", "end")
                self.base.tag_remove("diffcell", "1.0", "end")
                self.right.tag_remove("diffcell", "1.0", "end")
                self.left.tag_remove("paddingrow", "1.0", "end")
                self.base.tag_remove("paddingrow", "1.0", "end")
                self.right.tag_remove("paddingrow", "1.0", "end")
                # apply cached tags in bulk (one Tcl call per tag per widget)
                if tag_rows:
                    cached_diffrow_args = []
                    for line_idx in tag_rows:
                        cached_diffrow_args.extend([f"{line_idx}.0", f"{line_idx}.end"])
                    self.left.tag_add("diffrow", *cached_diffrow_args)
                    self.base.tag_add("diffrow", *cached_diffrow_args)
                    self.right.tag_add("diffrow", *cached_diffrow_args)
                    self.left_ln.tag_add("diffrow", *cached_diffrow_args)
                    self.base_ln.tag_add("diffrow", *cached_diffrow_args)
                    self.right_ln.tag_add("diffrow", *cached_diffrow_args)
                if tag_cells:
                    cached_cell_left = []
                    cached_cell_right = []
                    for line_idx, spans_a, spans_b in tag_cells:
                        for s, e in spans_a:
                            cached_cell_left.extend([f"{line_idx}.{s}", f"{line_idx}.{e}"])
                        for s, e in spans_b:
                            cached_cell_right.extend([f"{line_idx}.{s}", f"{line_idx}.{e}"])
                    if cached_cell_left:
                        self.left.tag_add("diffcell", *cached_cell_left)
                    if cached_cell_right:
                        self.right.tag_add("diffcell", *cached_cell_right)

                # paddingrow: grey slot for one-sided pairs (computed from row_pairs, not cached)
                _padding_left = []
                _padding_right = []
                for _i, _pidx in enumerate(self.display_rows):
                    if _pidx < len(self.row_pairs):
                        _ra, _rb = self.row_pairs[_pidx]
                        _ln = _i + 1
                        if _ra is None:
                            _padding_left.extend([f"{_ln}.0", f"{_ln}.end"])
                        elif _rb is None:
                            _padding_right.extend([f"{_ln}.0", f"{_ln}.end"])
                if _padding_left:
                    self.left.tag_add("paddingrow", *_padding_left)
                if _padding_right:
                    self.right.tag_add("paddingrow", *_padding_right)
                # rownum gutter tags are unused when row headers are separate

                mode = "只看差异" if self.only_diff_var.get() else "全量"
                total_rows = len(self.row_pairs) if self.row_pairs else self.max_row
                self.info.configure(text=f"{mode} | RowsShown: {len(self.display_rows)} / {total_rows}   Cols: {self.max_col}   DiffRows: {diff_row_count}")
                self._display_diff_row_count = diff_row_count
                self.app.set_sheet_has_diff(self.sheet, diff_row_count > 0, confirmed=True)
                self.app.refresh_sheet_nav()
                self._update_diff_nav_state()
                try:
                    self._update_diff_maps()
                except Exception:
                    pass
                return

        # Full render
        self.left.delete("1.0", "end")
        self.base.delete("1.0", "end")
        self.right.delete("1.0", "end")
        self.left.tag_remove("diffrow", "1.0", "end")
        self.base.tag_remove("diffrow", "1.0", "end")
        self.right.tag_remove("diffrow", "1.0", "end")
        self.left.tag_remove("diffcell", "1.0", "end")
        self.base.tag_remove("diffcell", "1.0", "end")
        self.right.tag_remove("diffcell", "1.0", "end")
        self.left.tag_remove("paddingrow", "1.0", "end")
        self.base.tag_remove("paddingrow", "1.0", "end")
        self.right.tag_remove("paddingrow", "1.0", "end")

        # Build full text in memory and insert once (faster)
        lines_a = []
        lines_b = []
        for pair_idx in self.display_rows:
            lines_a.append(self.pair_text_a.get(pair_idx, ""))
            lines_b.append(self.pair_text_b.get(pair_idx, ""))
        self.left.insert("1.0", "\n".join(lines_a) + ("\n" if lines_a else ""))
        self._render_base_full()
        self.right.insert("1.0", "\n".join(lines_b) + ("\n" if lines_b else ""))
        self._render_row_headers_full()

        # On some environments/large documents, forcing an idle layout pass improves tag correctness.
        try:
            self.left.update_idletasks()
            self.base.update_idletasks()
            self.right.update_idletasks()
        except Exception:
            pass

        # Restore scroll position if we just appended more rows
        if self._pending_yview is not None:
            try:
                self.left.yview_moveto(self._pending_yview)
                if self._is_three_way_enabled():
                    self.base.yview_moveto(self._pending_yview)
                self.right.yview_moveto(self._pending_yview)
            except Exception:
                pass
            self._pending_yview = None

        diff_row_count = 0
        tag_rows = []
        tag_cells = []
        # Collect all tag ranges first; apply in bulk (one Tcl call per tag instead of N).
        # tag_add(tagName, index1, *args) accepts multiple index pairs in a single call.
        diffrow_args = []
        diffcell_args_left = []
        diffcell_args_right = []
        for line_idx, pair_idx in enumerate(self.display_rows, start=1):
            cols = self.pair_diff_cols.get(pair_idx, set())
            if cols:
                diff_row_count += 1
                diffrow_args.extend([f"{line_idx}.0", f"{line_idx}.end"])
                tag_rows.append(line_idx)

                line_a = lines_a[line_idx - 1] if (line_idx - 1) < len(lines_a) else ""
                line_b = lines_b[line_idx - 1] if (line_idx - 1) < len(lines_b) else ""

                spans_a = spans_b = self._spans_for_line()  # result is argument-independent; compute once
                spans_a_ranges = []
                spans_b_ranges = []
                for c in cols:
                    if c in spans_a:
                        s, e = spans_a[c]
                        diffcell_args_left.extend([f"{line_idx}.{s}", f"{line_idx}.{e}"])
                        spans_a_ranges.append((s, e))
                    if c in spans_b:
                        s, e = spans_b[c]
                        diffcell_args_right.extend([f"{line_idx}.{s}", f"{line_idx}.{e}"])
                        spans_b_ranges.append((s, e))
                if spans_a_ranges or spans_b_ranges:
                    tag_cells.append((line_idx, spans_a_ranges, spans_b_ranges))

        # Apply all diffrow tags in one call per widget
        if diffrow_args:
            self.left.tag_add("diffrow", *diffrow_args)
            self.base.tag_add("diffrow", *diffrow_args)
            self.right.tag_add("diffrow", *diffrow_args)
            self.left_ln.tag_add("diffrow", *diffrow_args)
            self.base_ln.tag_add("diffrow", *diffrow_args)
            self.right_ln.tag_add("diffrow", *diffrow_args)
            for line_idx in tag_rows:
                try:
                    pidx = self.display_rows[line_idx - 1]
                    self._apply_rownum_diff_tag_line(line_idx, pidx)
                except Exception:
                    pass
        # Apply all diffcell tags in one call per widget
        if diffcell_args_left:
            self.left.tag_add("diffcell", *diffcell_args_left)
        if diffcell_args_right:
            self.right.tag_add("diffcell", *diffcell_args_right)
        # Apply paddingrow (grey) to empty slots of one-sided pairs
        _padding_left = []
        _padding_right = []
        for _i, _pidx in enumerate(self.display_rows):
            if _pidx < len(self.row_pairs):
                _ra, _rb = self.row_pairs[_pidx]
                _ln = _i + 1
                if _ra is None:
                    _padding_left.extend([f"{_ln}.0", f"{_ln}.end"])
                elif _rb is None:
                    _padding_right.extend([f"{_ln}.0", f"{_ln}.end"])
        if _padding_left:
            self.left.tag_add("paddingrow", *_padding_left)
        if _padding_right:
            self.right.tag_add("paddingrow", *_padding_right)
        # row-number styling handled by dedicated row-header widgets

        mode = "只看差异" if self.only_diff_var.get() else "全量"
        total_rows = len(self.row_pairs) if self.row_pairs else self.max_row
        self.info.configure(text=f"{mode} | RowsShown: {len(self.display_rows)} / {total_rows}   Cols: {self.max_col}   DiffRows: {diff_row_count}")
        self._display_diff_row_count = diff_row_count

        self.app.set_sheet_has_diff(self.sheet, diff_row_count > 0, confirmed=True)
        self.app.refresh_sheet_nav()
        self._update_diff_nav_state()
        try:
            self._update_diff_maps()
        except Exception:
            pass

        # Cache rendered result for fast toggle
        if row_only is None:
            text_a = "\n".join(lines_a) + ("\n" if lines_a else "")
            text_b = "\n".join(lines_b) + ("\n" if lines_b else "")
            self._render_cache[cache_key] = (text_a, text_b, tag_rows, tag_cells, diff_row_count)


class SowMergeApp:
    def __init__(self, file_a: str, file_b: str, merge_mode: bool = False, merged_path: str | None = None,
                 base_path: str | None = None,
                 merge_conflict_cells_by_sheet: dict | None = None, merge_conflict_mode: bool = False,
                 raw_base: str | None = None, raw_mine: str | None = None, raw_theirs: str | None = None):
        self.file_a = file_a
        self.file_b = file_b
        self.base_path = base_path
        self.has_base = bool(base_path and os.path.exists(base_path))
        self.raw_base = raw_base
        self.raw_mine = raw_mine
        self.raw_theirs = raw_theirs
        self.merge_mode = merge_mode
        self.diff_base_mine_mode = bool((not merge_mode) and raw_base and raw_mine)
        self.merged_path = merged_path
        self.merge_conflict_cells_by_sheet = merge_conflict_cells_by_sheet or {}
        self.merge_conflict_mode = merge_conflict_mode
        self.initial_conflict_cell_count = sum(
            len(cols)
            for rows in self.merge_conflict_cells_by_sheet.values()
            for cols in rows.values()
        )
        self.user_touched_conflicts = False
        # In 3-way manual merge mode, persist only explicitly operated A-side cells on save.
        # key: (sheet, row, col) -> edit value to write
        self.manual_a_cell_ops: dict[tuple[str, int, int], object] = {}
        self._merge_mine_snapshot = None
        self.undo_stack = []
        self._auto_recalc_started = False
        # append debug session marker each run (do not truncate old evidence)
        try:
            with open(_DEBUG_LOG_PATH, "a", encoding="utf-8") as f:
                f.write("\n" + "=" * 72 + "\n")
                f.write(f"SESSION {datetime.now().isoformat(sep=' ', timespec='seconds')}\n")
                f.write(f"{APP_NAME} {APP_VERSION} [{APP_BUILD_TAG}]\n")
                f.write(f"A={self.file_a}\nB={self.file_b}\n")
        except Exception:
            pass
        # load settings
        self.settings = {}
        self.only_diff_default = 0
        try:
            os.makedirs(os.path.dirname(_SETTINGS_PATH), exist_ok=True)
            if os.path.exists(_SETTINGS_PATH):
                with open(_SETTINGS_PATH, "r", encoding="utf-8") as f:
                    self.settings = json.load(f) or {}
            self.only_diff_default = int(self.settings.get("only_diff", 0))
        except Exception as e:
            _dlog(f"settings load failed: {e}")

        _dlog(f"SowMergeApp init only_diff_default={self.only_diff_default}")

        # Fast open: load value workbooks first; defer editable workbooks until first modification/save.
        self._wb_a_edit = None
        self._wb_b_edit = None
        self._wb_base_edit = None
        self._edit_loaded_event = threading.Event()
        self._edit_loading_started = False
        self._edit_fallback_lock = threading.Lock()

        try:
            t0 = datetime.now()
            self._file_a_val_path = _prepare_val_path(file_a)
            self._wb_a_val = load_workbook(self._file_a_val_path, data_only=True)
            _dlog(f"load wb_a_val: {(datetime.now()-t0).total_seconds():.3f}s")
            t0 = datetime.now()
            self._file_b_val_path = _prepare_val_path(file_b)
            self._wb_b_val = load_workbook(self._file_b_val_path, data_only=True)
            _dlog(f"load wb_b_val: {(datetime.now()-t0).total_seconds():.3f}s")
            self._wb_base_val = None
            if self.has_base:
                t0 = datetime.now()
                self._file_base_val_path = _prepare_val_path(self.base_path)
                self._wb_base_val = load_workbook(self._file_base_val_path, data_only=True)
                _dlog(f"load wb_base_val: {(datetime.now()-t0).total_seconds():.3f}s")
        except Exception:
            _wbs_close(
                getattr(self, "_wb_a_val", None),
                getattr(self, "_wb_b_val", None),
                getattr(self, "_wb_base_val", None),
            )
            raise
        if self.has_base:
            try:
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                snap = os.path.join(tempfile.gettempdir(), f"{APP_NAME}_mine_snapshot_{os.getpid()}_{ts}.xlsx")
                shutil.copy2(self.file_a, snap)
                self._merge_mine_snapshot = snap
                _dlog(f"mine snapshot created: {snap}")
            except Exception as e:
                _dlog(f"mine snapshot failed: {e}")

        # Preload editable workbooks in background to make the first overwrite fast.
        # Always run regardless of _FAST_OPEN_ENABLED: fast-open defers value loading
        # but edit workbooks must still be ready before the user's first row override.
        def _preload_edit():
            try:
                _dlog("preload edit workbooks (background) start")
                t1 = datetime.now()
                a_edit = load_workbook(self.file_a, data_only=False)
                _dlog(f"preload wb_a_edit: {(datetime.now()-t1).total_seconds():.3f}s")
                t2 = datetime.now()
                b_edit = load_workbook(self.file_b, data_only=False)
                _dlog(f"preload wb_b_edit: {(datetime.now()-t2).total_seconds():.3f}s")
                base_edit = None
                if self.has_base:
                    t3 = datetime.now()
                    base_edit = load_workbook(self.base_path, data_only=False)
                    _dlog(f"preload wb_base_edit: {(datetime.now()-t3).total_seconds():.3f}s")
                self._wb_a_edit = a_edit
                self._wb_b_edit = b_edit
                self._wb_base_edit = base_edit
            except Exception as e:
                _dlog(f"preload edit failed: {e}")
            finally:
                self._edit_loaded_event.set()
                _dlog("preload edit workbooks (background) done")

        self._edit_loading_started = True
        threading.Thread(target=_preload_edit, daemon=True).start()

        # Determine sheets from value workbooks (available immediately)
        set_a = set(self._wb_a_val.sheetnames)
        set_b = set(self._wb_b_val.sheetnames)
        self.common_sheets = sorted(set_a & set_b)
        if self.merge_conflict_mode and self.merge_conflict_cells_by_sheet:
            # Only keep sheets that actually have conflicts
            conflict_sheets = sorted(self.merge_conflict_cells_by_sheet.keys())
            self.common_sheets = [s for s in conflict_sheets if s in self.common_sheets]
        self.only_a = sorted(set_a - set_b)
        self.only_b = sorted(set_b - set_a)

        self.modified_a = False
        self.modified_b = False
        self.modified_sheets_a = set()
        self.modified_sheets_b = set()

        # sheet diff state: 0=none, 1=maybe (sampled), 2=confirmed
        self.sheet_diff_state = {s: 0 for s in self.common_sheets}

        self.root = tk.Tk()
        self._window_title_suffix = f"{APP_NAME} {APP_VERSION} [{APP_BUILD_TAG}]"
        self.root.title(self._window_title_suffix)
        ttk.Style().theme_use("clam")
        if self.merge_mode:
            self.root.title(f"{self._window_title_suffix} (SVN Merge)")
        else:
            self.root.title(f"{self._window_title_suffix} (TortoiseMerge-like)")
        self.root.geometry("1450x860")

        self._is_closing = False
        self._root_after_ids: set[str] = set()
        try:
            self.root.protocol("WM_DELETE_WINDOW", self._shutdown_root)
        except Exception:
            pass

        self._build_ui()
        self._schedule_auto_recalc()

    def _safe_root_after(self, delay_ms: int, callback):
        if getattr(self, "_is_closing", False):
            return None
        holder = {"id": None}

        def _wrapped():
            aid = holder.get("id")
            if aid is not None:
                try:
                    self._root_after_ids.discard(aid)
                except Exception:
                    pass
            if getattr(self, "_is_closing", False):
                return
            callback()

        try:
            aid = self.root.after(int(delay_ms), _wrapped)
            holder["id"] = aid
            self._root_after_ids.add(aid)
            return aid
        except Exception:
            return None

    def _cancel_root_afters(self):
        try:
            pending = list(getattr(self, "_root_after_ids", set()) or ())
        except Exception:
            pending = []
        try:
            self._root_after_ids.clear()
        except Exception:
            pass
        for aid in pending:
            try:
                self.root.after_cancel(aid)
            except Exception:
                pass

    def _shutdown_root(self):
        if getattr(self, "_is_closing", False):
            return
        self._is_closing = True
        self._cancel_root_afters()
        try:
            if self.root.winfo_exists():
                # Ensure Tk mainloop exits and the window is actually closed.
                try:
                    self.root.quit()
                except Exception:
                    pass
                self.root.destroy()
        except Exception:
            pass

    def _ensure_edit_loaded(self):
        if self._wb_a_edit is not None and self._wb_b_edit is not None and (not self.has_base or self._wb_base_edit is not None):
            return

        # If background preload is running, wait briefly.
        if getattr(self, "_edit_loading_started", False):
            _dlog("waiting for background edit preload")
            self._edit_loaded_event.wait(timeout=10)
            if self._wb_a_edit is not None and self._wb_b_edit is not None and (not self.has_base or self._wb_base_edit is not None):
                return

        _dlog("loading edit workbooks (fallback)")
        with self._edit_fallback_lock:
            # Re-check under lock: background thread may have just finished.
            if self._wb_a_edit is not None and self._wb_b_edit is not None and (not self.has_base or self._wb_base_edit is not None):
                return
            t0 = datetime.now()
            self._wb_a_edit = load_workbook(self.file_a, data_only=False)
            _dlog(f"load wb_a_edit: {(datetime.now()-t0).total_seconds():.3f}s")
            t0 = datetime.now()
            self._wb_b_edit = load_workbook(self.file_b, data_only=False)
            _dlog(f"load wb_b_edit: {(datetime.now()-t0).total_seconds():.3f}s")
            if self.has_base:
                t0 = datetime.now()
                self._wb_base_edit = load_workbook(self.base_path, data_only=False)
                _dlog(f"load wb_base_edit: {(datetime.now()-t0).total_seconds():.3f}s")

    def ws_a_edit(self, sheet: str):
        self._ensure_edit_loaded()
        if self._wb_a_edit is None:
            raise KeyError("file_a edit workbook not available")
        return self._wb_a_edit[sheet]

    def ws_b_edit(self, sheet: str):
        self._ensure_edit_loaded()
        if self._wb_b_edit is None:
            raise KeyError("file_b edit workbook not available")
        return self._wb_b_edit[sheet]

    def ws_base_edit(self, sheet: str):
        self._ensure_edit_loaded()
        if self._wb_base_edit is None:
            raise KeyError("base workbook not available")
        return self._wb_base_edit[sheet]

    def ws_a_val(self, sheet: str):
        if self._wb_a_val is None:
            raise KeyError("file_a val workbook not available")
        return self._wb_a_val[sheet]

    def ws_b_val(self, sheet: str):
        if self._wb_b_val is None:
            raise KeyError("file_b val workbook not available")
        return self._wb_b_val[sheet]

    def ws_base_val(self, sheet: str):
        if self._wb_base_val is None:
            raise KeyError("base workbook not available")
        return self._wb_base_val[sheet]

    def record_manual_a_cell(self, sheet: str, r: int, c: int, edit_value):
        try:
            self.manual_a_cell_ops[(sheet, int(r), int(c))] = edit_value
        except Exception:
            pass

    def build_manual_merge_output_file(self):
        """Build merge output by XML-level patching from pristine mine snapshot."""
        src = self._merge_mine_snapshot if (self._merge_mine_snapshot and os.path.exists(self._merge_mine_snapshot)) else self.file_a
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out = os.path.join(tempfile.gettempdir(), f"{APP_NAME}_merged_output_{os.getpid()}_{ts}.xlsx")
        if not self.manual_a_cell_ops:
            shutil.copy2(src, out)
            return out
        if _EXCEL_NATIVE_SAVE_ON_MERGE:
            ok = _build_manual_merge_output_with_excel(src, out, self.manual_a_cell_ops)
            if ok:
                return out
            _dlog("WARNING: excel native save failed, falling back to zip patch (may cause Excel repair on files with complex formulas)")
        _build_manual_merge_xlsx_via_zip(src, out, self.manual_a_cell_ops)
        if not os.path.exists(out):
            raise RuntimeError(f"zip patch produced no output: {out}")
        return out

    def set_sheet_has_diff(self, sheet: str, has: bool, confirmed: bool = True):
        # Keep API: mark sheet diff state
        if sheet not in self.sheet_diff_state:
            return
        if has:
            self.sheet_diff_state[sheet] = 2 if confirmed else max(self.sheet_diff_state[sheet], 1)
        else:
            # only downgrade when confirmed
            if confirmed:
                self.sheet_diff_state[sheet] = 0

    def _build_ui(self):
        top = ttk.Frame(self.root)
        top.pack(fill="x", padx=10, pady=8)

        # Keep top area minimal (summary + actions). File-source labels are shown inside each Sheet.
        summary = f"同名Sheet: {len(self.common_sheets)}   仅A: {len(self.only_a)}   仅B: {len(self.only_b)}"
        ttk.Label(top, text=summary).grid(row=0, column=0, columnspan=2, sticky="w", pady=(2, 0))
        detail_row = 1
        if self.merge_mode and (self.raw_mine or self.raw_base or self.raw_theirs):
            raw_line = (
                f"SVN原始传参: mine={os.path.basename(self.raw_mine or '-')}"
                f" | base={os.path.basename(self.raw_base or '-')}"
                f" | theirs={os.path.basename(self.raw_theirs or '-')}"
            )
            ttk.Label(top, text=raw_line, foreground="#555").grid(row=detail_row, column=0, columnspan=3, sticky="w", pady=(4, 0))
            detail_row += 1
            read_line = (
                f"当前实际读取: left(A)={os.path.basename(self.file_a or '-')}"
                f" | base={os.path.basename(self.base_path or '-')}"
                f" | right(B)={os.path.basename(self.file_b or '-')}"
            )
            ttk.Label(top, text=read_line, foreground="#555").grid(row=detail_row, column=0, columnspan=3, sticky="w", pady=(2, 0))
        ttk.Label(top, text=f"Build: {APP_BUILD_TAG}", foreground="#666").grid(row=0, column=3, sticky="ne", padx=(16, 0))

        ttk.Button(top, text="重算并刷新", command=self.recalc_and_refresh).grid(row=0, column=2, sticky="ne", padx=(10, 0))
        ttk.Button(top, text="导出诊断包", command=self.export_diagnostic_bundle).grid(row=0, column=4, sticky="ne", padx=(10, 0))
        ttk.Button(top, text="复制反馈信息", command=self.copy_feedback_info).grid(row=0, column=5, sticky="ne", padx=(10, 0))
        self.update_btn = ttk.Button(top, text="检查更新", command=self._do_svn_update)
        self.update_btn.grid(row=0, column=6, sticky="ne", padx=(10, 0))

        ttk.Separator(self.root, orient="horizontal").pack(fill="x", padx=10, pady=(0, 6))

        self.nb = ttk.Notebook(self.root)
        self.nb.pack(fill="both", expand=True, padx=10, pady=(8, 6))

        # Bottom bar: sheet nav (only)
        self.bottom = ttk.Frame(self.root)
        self.bottom.pack(fill="x", padx=10, pady=(0, 10))

        self.nav = ttk.Frame(self.bottom)
        self.nav.pack(side="left", fill="x", expand=True)
        ttk.Label(self.nav, text="Sheets:").pack(side="left")
        self.nav_canvas = tk.Canvas(self.nav, height=28, highlightthickness=0)
        self.nav_canvas.pack(side="left", fill="x", expand=True, padx=(8, 0))
        self.nav_scroll = ttk.Scrollbar(self.nav, orient="horizontal", command=self.nav_canvas.xview)
        self.nav_scroll.pack(side="bottom", fill="x")
        self.nav_canvas.configure(xscrollcommand=self.nav_scroll.set)
        self.nav_inner = ttk.Frame(self.nav_canvas)
        self.nav_canvas.create_window((0, 0), window=self.nav_inner, anchor="nw")
        self.nav_inner.bind("<Configure>", lambda e: self.nav_canvas.configure(scrollregion=self.nav_canvas.bbox("all")))

        if self.only_a:
            self._add_missing_tab("仅在A存在", self.only_a)
        if self.only_b:
            self._add_missing_tab("仅在B存在", self.only_b)

        # Tabs are created up-front, but heavy SheetView is created lazily on first activation.
        self.sheet_views = {}
        self._sheet_loaded = {}
        self._sheet_containers = {}
        for s in self.common_sheets:
            container = ttk.Frame(self.nb)
            self._sheet_containers[s] = container
            self.nb.add(container, text=s)
            self.sheet_views[s] = None
            self._sheet_loaded[s] = False

        # Background compute queue for sheet diffs
        self._compute_lock = threading.Lock()
        self._compute_queue = []  # list of sheet names
        self._compute_inflight = set()
        self._compute_thread = None
        self._ui_task_lock = threading.Lock()
        self._ui_tasks = []

        def _enqueue_sheet(sheet: str, front: bool = False):
            with self._compute_lock:
                if sheet in self._compute_inflight:
                    return
                if sheet in self._compute_queue:
                    # move to front if requested
                    if front:
                        self._compute_queue.remove(sheet)
                        self._compute_queue.insert(0, sheet)
                    return
                if front:
                    self._compute_queue.insert(0, sheet)
                else:
                    self._compute_queue.append(sheet)

        def _queue_ui_task(fn):
            with self._ui_task_lock:
                self._ui_tasks.append(fn)

        def _drain_ui_tasks():
            tasks = []
            try:
                with self._ui_task_lock:
                    if self._ui_tasks:
                        tasks = self._ui_tasks
                        self._ui_tasks = []
            except Exception:
                tasks = []
            ran = bool(tasks)
            for fn in tasks:
                try:
                    fn()
                except Exception as e:
                    _dlog(f"ui task failed: {e}")
            try:
                # Adaptive delay: poll frequently while work is flowing, back off when idle.
                delay = 50 if ran else 150
                self._safe_root_after(delay, _drain_ui_tasks)
            except Exception:
                pass

        def _compute_trim_bounds(ws):
            # Find last non-empty row/col; then +50 buffer.
            # Prefer ws._cells (only stored non-empty cells) to avoid missing data
            # when max_row/max_col are inflated by styles.
            max_r = ws.max_row or 1
            max_c = ws.max_column or 1
            last_r = 1
            last_c = 1
            found = False

            try:
                cells = getattr(ws, "_cells", None)
                if cells:
                    for cell in cells.values():
                        v = cell.value
                        if v not in (None, ""):
                            found = True
                            if cell.row > last_r:
                                last_r = cell.row
                            if cell.column > last_c:
                                last_c = cell.column
            except Exception:
                pass

            if not found:
                # Fallback: scan backwards up to 5000 rows.
                # If still not found, do NOT trim (avoid cutting real data).
                for r in range(max_r, max(1, max_r - 5000), -1):
                    row = next(ws.iter_rows(min_row=r, max_row=r, min_col=1, max_col=max_c, values_only=True), ())
                    if any(v not in (None, "") for v in row):
                        found = True
                        last_r = r
                        # determine last non-empty col from that row
                        for ci in range(len(row), 0, -1):
                            v = row[ci - 1]
                            if v not in (None, ""):
                                last_c = ci
                                break
                        break
                if not found:
                    return max_r, max_c

            use_r = min(max_r, last_r + 50)
            use_c = min(max_c, last_c + 50)
            return max(1, use_r), max(1, use_c)

        def _compute_row_pairs_bg(ws_a, ws_b, max_row_a: int, max_row_b: int, max_col: int):
            """Compute row alignment pairs using difflib.SequenceMatcher (background-safe)."""
            if max(max_row_a, max_row_b) >= _ROW_ALIGN_MAX_ROWS:
                max_row = max(max_row_a, max_row_b)
                pairs = []
                for r in range(1, max_row + 1):
                    ra = r if r <= max_row_a else None
                    rb = r if r <= max_row_b else None
                    pairs.append((ra, rb))
                return pairs

            def _bulk_sig_list(ws, max_row_local: int):
                sigs = []
                try:
                    for row in ws.iter_rows(
                        min_row=1,
                        max_row=max_row_local,
                        min_col=1,
                        max_col=max_col,
                        values_only=True,
                    ):
                        sigs.append("\x1f".join(_merge_cmp_value(v) for v in (row or ())))
                except Exception:
                    return []
                return sigs
            sig_a = _bulk_sig_list(ws_a, max_row_a)
            sig_b = _bulk_sig_list(ws_b, max_row_b)
            def _sim_score(sa: str, sb: str) -> float:
                if sa == sb:
                    return 2.0
                if (not sa) or (not sb):
                    return 0.0
                try:
                    return difflib.SequenceMatcher(a=sa, b=sb, autojunk=False).ratio()
                except Exception:
                    return 0.0
            sm = difflib.SequenceMatcher(a=sig_a, b=sig_b, autojunk=False)
            pairs: list[tuple[int | None, int | None]] = []
            for tag, i1, i2, j1, j2 in sm.get_opcodes():
                if tag == "equal":
                    for i, j in zip(range(i1, i2), range(j1, j2)):
                        pairs.append((i + 1, j + 1))
                elif tag == "replace":
                    len_a = i2 - i1
                    len_b = j2 - j1
                    common = min(len_a, len_b)
                    head_score = 0.0
                    tail_score = 0.0
                    for k in range(common):
                        head_score += _sim_score(sig_a[i1 + k], sig_b[j1 + k])
                        tail_score += _sim_score(sig_a[i2 - common + k], sig_b[j2 - common + k])
                    use_tail = tail_score >= head_score
                    if use_tail:
                        extra_a = len_a - common
                        extra_b = len_b - common
                        for k in range(extra_a):
                            pairs.append((i1 + k + 1, None))
                        for k in range(extra_b):
                            pairs.append((None, j1 + k + 1))
                        a_start = i2 - common
                        b_start = j2 - common
                        for k in range(common):
                            pairs.append((a_start + k + 1, b_start + k + 1))
                    else:
                        for k in range(common):
                            pairs.append((i1 + k + 1, j1 + k + 1))
                        for k in range(common, len_a):
                            pairs.append((i1 + k + 1, None))
                        for k in range(common, len_b):
                            pairs.append((None, j1 + k + 1))
                elif tag == "delete":
                    for i in range(i1, i2):
                        pairs.append((i + 1, None))
                elif tag == "insert":
                    for j in range(j1, j2):
                        pairs.append((None, j + 1))
            return pairs

        def _has_diff_by_blocks_bg(ws_a, ws_b, max_row_a: int, max_row_b: int, max_col: int):
            max_row = max(max_row_a, max_row_b)
            block = _LARGE_SHEET_BLOCK_ROWS
            for block_end in range(max_row, 0, -block):
                block_start = max(1, block_end - block + 1)
                rows_a = {}
                rows_b = {}
                if block_start <= max_row_a:
                    for idx, row in enumerate(
                        ws_a.iter_rows(
                            min_row=block_start,
                            max_row=min(block_end, max_row_a),
                            min_col=1,
                            max_col=max_col,
                            values_only=True,
                        ),
                        start=block_start,
                    ):
                        rows_a[idx] = row or ()
                if block_start <= max_row_b:
                    for idx, row in enumerate(
                        ws_b.iter_rows(
                            min_row=block_start,
                            max_row=min(block_end, max_row_b),
                            min_col=1,
                            max_col=max_col,
                            values_only=True,
                        ),
                        start=block_start,
                    ):
                        rows_b[idx] = row or ()
                for r in range(block_end, block_start - 1, -1):
                    row_a = rows_a.get(r, ())
                    row_b = rows_b.get(r, ())
                    for ci in range(max_col):
                        va = row_a[ci] if ci < len(row_a) else None
                        vb = row_b[ci] if ci < len(row_b) else None
                        if _merge_cmp_value(va) != _merge_cmp_value(vb):
                            return True
            return False

        def _compute_sheet_cache(wb_a_val, wb_b_val, wb_a_edit, wb_b_edit, sheet: str):
            ws_a = wb_a_val[sheet]
            ws_b = wb_b_val[sheet]
            max_r_a, max_c_a = _compute_trim_bounds(ws_a)
            max_r_b, max_c_b = _compute_trim_bounds(ws_b)
            max_row = max(max_r_a, max_r_b)
            max_col = max(max_c_a, max_c_b)

            # Compute row-aligned pairs (same algorithm as SheetView._build_row_pairs)
            row_pairs = _compute_row_pairs_bg(ws_a, ws_b, max_r_a, max_r_b, max_col)

            pair_diff_cols: dict[int, set] = {}
            # Keep raw per-cell display parts in cache; render with current grid mode on UI thread.
            pair_parts_a: dict[int, list[str]] = {}
            pair_parts_b: dict[int, list[str]] = {}
            col_char_widths: dict[int, int] = {}
            row_a_to_pair_idx: dict[int, int] = {}
            row_b_to_pair_idx: dict[int, int] = {}

            for idx, (ra, rb) in enumerate(row_pairs):
                if ra is not None:
                    row_a_to_pair_idx[ra] = idx
                if rb is not None:
                    row_b_to_pair_idx[rb] = idx

            # Large-sheet fast open: avoid full cell-by-cell precompute.
            # Still estimate display widths from head + tail samples to prevent 4-char collapse.
            if max_row >= _LARGE_SHEET_ROW_THRESHOLD:
                has_diff = _has_diff_by_blocks_bg(ws_a, ws_b, max_r_a, max_r_b, max_col)

                sample_head = min(_LARGE_SHEET_INITIAL_ROWS, len(row_pairs))
                sample_indices = list(range(sample_head))
                tail_n = 50
                tail_start = max(sample_head, len(row_pairs) - tail_n)
                if tail_start < len(row_pairs):
                    sample_indices.extend(range(tail_start, len(row_pairs)))

                def _row_values_bg(ws, row_idx: int | None):
                    if row_idx is None:
                        return ()
                    try:
                        row = next(
                            ws.iter_rows(
                                min_row=row_idx,
                                max_row=row_idx,
                                min_col=1,
                                max_col=max_col,
                                values_only=True,
                            ),
                            (),
                        )
                    except Exception:
                        row = ()
                    return row or ()

                for idx in sample_indices:
                    ra, rb = row_pairs[idx]
                    row_a_vals = _row_values_bg(ws_a, ra)
                    row_b_vals = _row_values_bg(ws_b, rb)
                    for ci in range(max_col):
                        va = row_a_vals[ci] if ci < len(row_a_vals) else None
                        vb = row_b_vals[ci] if ci < len(row_b_vals) else None
                        sa = _val_to_str(va)
                        sb = _val_to_str(vb)
                        w = min(max(len(sa), len(sb)), _COL_MAX_DISPLAY_WIDTH)
                        col_idx = ci + 1
                        if w > col_char_widths.get(col_idx, 0):
                            col_char_widths[col_idx] = w
            else:
                ws_a_e = wb_a_edit[sheet]
                ws_b_e = wb_b_edit[sheet]
                for idx, (ra, rb) in enumerate(row_pairs):
                    cols = set()
                    parts_a = []
                    parts_b = []
                    for c in range(1, max_col + 1):
                        da, db, eq = _cell_display_and_equal_by_row(ws_a, ws_b, ws_a_e, ws_b_e, ra, rb, c)
                        sa = _val_to_str(da)
                        sb = _val_to_str(db)
                        parts_a.append(sa)
                        parts_b.append(sb)
                        w = min(max(len(sa), len(sb)), _COL_MAX_DISPLAY_WIDTH)
                        if w > col_char_widths.get(c, 0):
                            col_char_widths[c] = w
                        if not eq:
                            cols.add(c)
                    pair_parts_a[idx] = parts_a
                    pair_parts_b[idx] = parts_b
                    pair_diff_cols[idx] = cols
                has_diff = any(bool(v) for v in pair_diff_cols.values())

            # Keep a readable lower bound and normalize missing columns.
            for c in range(1, max_col + 1):
                col_char_widths[c] = max(4, int(col_char_widths.get(c, 1)))

            return {
                "sheet": sheet,
                "max_row": max_row,
                "max_col": max_col,
                "row_pairs": row_pairs,
                "pair_diff_cols": pair_diff_cols,
                "pair_parts_a": pair_parts_a,
                "pair_parts_b": pair_parts_b,
                "col_char_widths": col_char_widths,
                "row_a_to_pair_idx": row_a_to_pair_idx,
                "row_b_to_pair_idx": row_b_to_pair_idx,
                "has_diff": has_diff,
            }

        def _apply_sheet_cache(cache: dict):
            sheet = cache["sheet"]
            view = self.sheet_views.get(sheet)
            if view is None:
                # No visible view yet: safe to update tab mark directly.
                self.set_sheet_has_diff(sheet, cache.get("has_diff", False), confirmed=True)
                self.refresh_sheet_nav()
                return
            if getattr(view, "_suppress_bg_apply", False):
                _dlog(f"skip bg cache apply by user action: sheet={sheet}")
                self.refresh_sheet_nav()
                return
            # Skip if the user has made edits in this view; background data (from read-only copies)
            # would be stale relative to the user's in-memory changes.
            if getattr(view, "_data_ready", False) and view.touched_rows:
                self.refresh_sheet_nav()
                return
            # Guard against late background cache downgrading an already rendered sheet to no-diff.
            # This has been observed as a delayed "DiffRows -> 0 / rows disappear" regression.
            try:
                old_diff_count = sum(1 for _k, _v in (view.pair_diff_cols or {}).items() if _v)
            except Exception:
                old_diff_count = 0
            try:
                new_diff_count = sum(1 for _k, _v in (cache.get("pair_diff_cols", {}) or {}).items() if _v)
            except Exception:
                new_diff_count = 0
            if getattr(view, "_data_ready", False) and old_diff_count > 0 and new_diff_count == 0:
                _dlog(f"skip stale cache downgrade: sheet={sheet} old_diff={old_diff_count} new_diff={new_diff_count}")
                self.refresh_sheet_nav()
                return
            # From this point we will apply this cache to the visible view.
            self.set_sheet_has_diff(sheet, cache.get("has_diff", False), confirmed=True)
            view.max_row = cache["max_row"]
            view.max_col = cache["max_col"]
            view._is_large_sheet = view.max_row >= _LARGE_SHEET_ROW_THRESHOLD
            view._bounds_checked = True

            # Apply row-aligned pair data (computed in background with row alignment)
            view.row_pairs = cache["row_pairs"]
            view.pair_diff_cols = cache["pair_diff_cols"]
            view.col_char_widths = cache.get("col_char_widths", {}) or {c: 4 for c in range(1, view.max_col + 1)}

            pair_parts_a = cache.get("pair_parts_a", {}) or {}
            pair_parts_b = cache.get("pair_parts_b", {}) or {}
            grid_on = view._is_grid_overlay_enabled()
            sep = _COL_SEP if grid_on else "   "
            trail = " │" if grid_on else ""
            if pair_parts_a or pair_parts_b:
                view.pair_text_a = {}
                view.pair_text_b = {}
                for idx, parts in pair_parts_a.items():
                    view.pair_text_a[idx] = sep.join(
                        _format_cell(parts[i], view.col_char_widths.get(i + 1, 1))
                        for i in range(len(parts))
                    ) + (trail if parts else "")
                for idx, parts in pair_parts_b.items():
                    view.pair_text_b[idx] = sep.join(
                        _format_cell(parts[i], view.col_char_widths.get(i + 1, 1))
                        for i in range(len(parts))
                    ) + (trail if parts else "")
            else:
                # Backward-compatible fallback for older cache shape.
                view.pair_text_a = cache.get("pair_text_a", {})
                view.pair_text_b = cache.get("pair_text_b", {})

            view.row_a_to_pair_idx = cache["row_a_to_pair_idx"]
            view.row_b_to_pair_idx = cache["row_b_to_pair_idx"]
            view._align_rows_enabled = True
            view._diff_partial = False
            # Mark data as ready so refresh(rescan=False) uses it without rescanning
            view._data_ready = True
            view._invalidate_render_cache()

            if view._prefer_only_diff_when_ready:
                # For large sheets, keep full-mode initial render (first 200 rows) for responsiveness.
                if view.max_row >= _LARGE_SHEET_ROW_THRESHOLD:
                    view.only_diff_var.set(0)
                    view._full_render = False
                    view._render_limit = min(_LARGE_SHEET_INITIAL_ROWS, view.max_row)
                elif cache.get("has_diff", False):
                    view.only_diff_var.set(1)
                else:
                    view.only_diff_var.set(0)
                    view._full_render = False
                    view._render_limit = min(_LARGE_SHEET_INITIAL_ROWS, view.max_row)
                view._prefer_only_diff_when_ready = False
            # Preserve viewport/cursor when background cache is applied; otherwise
            # user operations (overwrite/resolve) appear to "jump to first row/first column".
            prev_first = 0.0
            prev_x = 0.0
            prev_insert = "1.0"
            try:
                prev_first = float((view.left.yview() or (0.0, 1.0))[0])
                prev_x = float((view.left.xview() or (0.0, 1.0))[0])
                prev_insert = view.left.index("insert")
            except Exception:
                pass
            view.refresh(row_only=None, rescan=False)
            try:
                view.left.yview_moveto(prev_first)
                if view._is_three_way_enabled():
                    view.base.yview_moveto(prev_first)
                view.right.yview_moveto(prev_first)
            except Exception:
                pass
            try:
                view._sync_main_x_to_frac(prev_x)
                view._sync_c_x_to_frac(prev_x)
            except Exception:
                pass
            try:
                line = int(str(prev_insert).split(".")[0])
                col = int(str(prev_insert).split(".")[1])
            except Exception:
                line = 1
                col = 0
            try:
                max_line = max(1, len(view.display_rows))
            except Exception:
                max_line = 1
            if line < 1:
                line = 1
            if line > max_line:
                line = max_line
            try:
                idx = f"{line}.{max(0, col)}"
                view.left.mark_set("insert", idx)
                if view._is_three_way_enabled():
                    view.base.mark_set("insert", idx)
                view.right.mark_set("insert", idx)
            except Exception:
                pass
            view._update_cursor_lines()
            self.refresh_sheet_nav()

        def _compute_worker():
            wb_a_ro = None
            wb_b_ro = None
            wb_a_e = None
            wb_b_e = None
            try:
                try:
                    # Use separate read-only workbooks to avoid threading issues
                    wb_a_ro = load_workbook(self._file_a_val_path, data_only=True, read_only=True)
                    wb_b_ro = load_workbook(self._file_b_val_path, data_only=True, read_only=True)
                    wb_a_e = load_workbook(self.file_a, data_only=False, read_only=True)
                    wb_b_e = load_workbook(self.file_b, data_only=False, read_only=True)
                except Exception as e:
                    _dlog(f"bg compute open read-only failed: {e}")
                    return
                if wb_a_ro is None or wb_b_ro is None or wb_a_e is None or wb_b_e is None:
                    _dlog("bg compute read-only workbooks not available; skip background compute")
                    return

                while True:
                    with self._compute_lock:
                        if not self._compute_queue:
                            break
                        sheet = self._compute_queue.pop(0)
                        self._compute_inflight.add(sheet)
                    try:
                        _dlog(f"bg compute sheet: {sheet}")
                        cache = _compute_sheet_cache(wb_a_ro, wb_b_ro, wb_a_e, wb_b_e, sheet)
                        # Never call tkinter APIs from background threads.
                        _queue_ui_task(lambda c=cache: _apply_sheet_cache(c))
                    except Exception as e:
                        _dlog(f"bg compute failed {sheet}: {e}")
                    finally:
                        with self._compute_lock:
                            self._compute_inflight.discard(sheet)
            finally:
                _wbs_close(wb_a_ro, wb_b_ro, wb_a_e, wb_b_e)
                with self._compute_lock:
                    if self._compute_thread is threading.current_thread():
                        self._compute_thread = None

        def _kick_worker():
            # start a worker if not running
            with self._compute_lock:
                th = self._compute_thread
                if th is not None and th.is_alive():
                    return
                if not self._compute_queue:
                    return
                th = threading.Thread(target=_compute_worker, daemon=True)
                self._compute_thread = th
                th.start()
        self._kick_worker = _kick_worker

        # Lazy-create SheetView UI immediately; compute diff in background.
        def _on_tab_changed(_evt=None):
            try:
                tab_id = self.nb.select()
                tab_text = self.nb.tab(tab_id, "text")
                self.selected_sheet = tab_text
                self.refresh_sheet_nav()
                if tab_text in self._sheet_containers and not self._sheet_loaded.get(tab_text, False):
                    _dlog(f"lazy create SheetView (ui only): {tab_text}")
                    view = SheetView(self._sheet_containers[tab_text], self, tab_text)
                    self.sheet_views[tab_text] = view
                    self._sheet_loaded[tab_text] = True
                    # Show loading placeholder immediately (non-blocking).
                    # The background worker will compute diffs and call _apply_sheet_cache
                    # which sets _data_ready=True and calls refresh(rescan=False).
                    view._show_loading()
                if tab_text in self._sheet_containers:
                    # Skip background recompute if data is already ready (no edits pending).
                    # Reopening workbooks on every tab switch is the main perf regression.
                    _view = self.sheet_views.get(tab_text)
                    if not (_view and getattr(_view, "_data_ready", False)):
                        _enqueue_sheet(tab_text, front=True)
                        _kick_worker()
                        # Fallback: if background path stalls, force sync refresh on UI thread.
                        # This guarantees that switching tabs eventually shows compare results.
                        def _force_refresh_if_still_loading(sheet_name=tab_text):
                            try:
                                cur_id = self.nb.select()
                                cur_sheet = self.nb.tab(cur_id, "text")
                                if cur_sheet != sheet_name:
                                    return
                                v = self.sheet_views.get(sheet_name)
                                if not v:
                                    return
                                if getattr(v, "_data_ready", False):
                                    return
                                v.refresh(row_only=None, rescan=True)
                                v._update_cursor_lines()
                            except Exception as e:
                                _dlog(f"force refresh fallback failed {sheet_name}: {e}")
                        try:
                            self._safe_root_after(700, _force_refresh_if_still_loading)
                        except Exception:
                            pass
            except Exception as e:
                _dlog(f"tab changed handler failed: {e}")

        try:
            self.nb.bind("<<NotebookTabChanged>>", _on_tab_changed)
        except Exception:
            pass

        # Main-thread UI task pump (for background compute/sample updates).
        try:
            self._safe_root_after(50, _drain_ui_tasks)
        except Exception:
            pass

        # Load the initially selected tab immediately so first-open state is ready.
        _on_tab_changed()

        self.refresh_sheet_nav()

        # Background fast pre-mark for sheet tabs:
        # exact by cached values (tail-first block scan), no random sampling.
        def _apply_fast_mark_result(sheet: str, has: bool):
            self.set_sheet_has_diff(sheet, has, confirmed=True)
            view = self.sheet_views.get(sheet)
            if has and view and view._prefer_only_diff_when_ready and view.only_diff_var.get() == 0:
                view.only_diff_var.set(1)
                view.refresh(row_only=None, rescan=False)
                view._update_cursor_lines()
            self.refresh_sheet_nav()

        def _sheet_has_diff_fast_tail(ws_a, ws_b, max_row: int, max_col: int, min_row: int = 1):
            none_sig = tuple("" for _ in range(max_col))
            block = _LARGE_SHEET_BLOCK_ROWS
            max_row_a = ws_a.max_row or 1
            max_row_b = ws_b.max_row or 1

            for block_end in range(max_row, 0, -block):
                block_start = max(1, block_end - block + 1)
                if block_end < min_row:
                    break
                if block_start < min_row:
                    block_start = min_row
                end_a = min(block_end, max_row_a)
                end_b = min(block_end, max_row_b)

                rows_a = []
                rows_b = []
                if block_start <= end_a:
                    rows_a = list(ws_a.iter_rows(
                        min_row=block_start,
                        max_row=end_a,
                        min_col=1,
                        max_col=max_col,
                        values_only=True,
                    ))
                if block_start <= end_b:
                    rows_b = list(ws_b.iter_rows(
                        min_row=block_start,
                        max_row=end_b,
                        min_col=1,
                        max_col=max_col,
                        values_only=True,
                    ))

                sig_a = [tuple(_merge_cmp_value(v) for v in (row or ())) for row in rows_a]
                sig_b = [tuple(_merge_cmp_value(v) for v in (row or ())) for row in rows_b]

                for r in range(block_end, block_start - 1, -1):
                    if r <= max_row_a:
                        ia = r - block_start
                        sa = sig_a[ia] if 0 <= ia < len(sig_a) else none_sig
                    else:
                        sa = none_sig
                    if r <= max_row_b:
                        ib = r - block_start
                        sb = sig_b[ib] if 0 <= ib < len(sig_b) else none_sig
                    else:
                        sb = none_sig
                    if sa != sb:
                        return True
            return False

        def _sheet_has_diff_quick_tail(ws_a, ws_b, max_row: int, max_col: int):
            # Phase-1 quick check: scan only the tail window.
            # True means "confirmed diff"; False means "unknown yet".
            quick_rows = min(max_row, _TABMARK_QUICK_TAIL_ROWS)
            if quick_rows <= 0:
                return False
            start = max(1, max_row - quick_rows + 1)
            return _sheet_has_diff_fast_tail(ws_a, ws_b, max_row, max_col, min_row=start)

        def _scan_sheet_has_diff_fast():
            wb_a_ro = None
            wb_b_ro = None
            try:
                try:
                    a_sz = os.path.getsize(self._file_a_val_path)
                    b_sz = os.path.getsize(self._file_b_val_path)
                    if max(a_sz, b_sz) >= (_FAST_TABMARK_SCAN_SKIP_MB * 1024 * 1024):
                        _dlog(
                            f"skip fast diff mark scan on large files: "
                            f"a={a_sz} b={b_sz} threshold_mb={_FAST_TABMARK_SCAN_SKIP_MB}"
                        )
                        return
                except Exception:
                    pass

                wb_a_ro = load_workbook(self._file_a_val_path, data_only=True, read_only=True)
                wb_b_ro = load_workbook(self._file_b_val_path, data_only=True, read_only=True)
                ordered = list(self.common_sheets)
                if ordered:
                    # Prefer currently selected sheet first, then newer tabs first.
                    cur = getattr(self, "selected_sheet", None)
                    if cur in ordered:
                        ordered.remove(cur)
                        ordered = [cur] + list(reversed(ordered))
                    else:
                        ordered = list(reversed(ordered))

                unknown_sheets = []

                # Phase-1: quick tail scan to surface diff tabs early.
                for s in ordered:
                    ws_a = wb_a_ro[s]
                    ws_b = wb_b_ro[s]
                    max_row = max(ws_a.max_row or 1, ws_b.max_row or 1)
                    max_col = max(ws_a.max_column or 1, ws_b.max_column or 1)
                    has_quick = _sheet_has_diff_quick_tail(ws_a, ws_b, max_row, max_col)
                    if has_quick:
                        _queue_ui_task(lambda s=s: _apply_fast_mark_result(s, True))
                    else:
                        unknown_sheets.append((s, ws_a, ws_b, max_row, max_col))

                # Phase-2 exact scan is optional: background compute will always confirm.
                if _FAST_TABMARK_PHASE2_ENABLED:
                    # Re-fetch worksheet objects: read_only iterators from Phase-1 are consumed.
                    for s, _stale_a, _stale_b, max_row, max_col in unknown_sheets:
                        ws_a = wb_a_ro[s]
                        ws_b = wb_b_ro[s]
                        has = _sheet_has_diff_fast_tail(ws_a, ws_b, max_row, max_col)
                        _queue_ui_task(lambda s=s, has=has: _apply_fast_mark_result(s, has))
            except Exception as e:
                _dlog(f"fast diff mark scan failed: {e}")
            finally:
                try:
                    if wb_a_ro:
                        wb_a_ro.close()
                    if wb_b_ro:
                        wb_b_ro.close()
                except Exception:
                    pass

        try:
            threading.Thread(target=_scan_sheet_has_diff_fast, daemon=True).start()
        except Exception:
            pass

        # Enqueue all sheets for background confirmation (slow compute)
        try:
            for s in self.common_sheets:
                _enqueue_sheet(s, front=False)
            _kick_worker()
        except Exception as e:
            _dlog(f"enqueue all sheets failed: {e}")

    def push_undo(self, action: dict):
        try:
            self.undo_stack.append(action)
            if len(self.undo_stack) > 20:
                self.undo_stack.pop(0)
        except Exception:
            pass

    def pop_undo(self) -> dict | None:
        try:
            if not self.undo_stack:
                return None
            return self.undo_stack.pop()
        except Exception:
            return None

    def _add_missing_tab(self, title: str, items):
        frame = ttk.Frame(self.nb)
        self.nb.add(frame, text=title)
        ttk.Label(frame, text=title, font=("Segoe UI", 11, "bold")).pack(anchor="w", padx=8, pady=(8, 4))
        txt = tk.Text(frame, wrap="none", height=10)
        txt.pack(fill="both", expand=True, padx=8, pady=8)
        txt.insert("1.0", "\n".join(items))
        txt.configure(state="disabled")

    def _select_tab(self, tab_text: str):
        for tab_id in self.nb.tabs():
            if self.nb.tab(tab_id, "text") == tab_text:
                self.nb.select(tab_id)
                return

    def refresh_sheet_nav(self):
        for child in list(self.nav_inner.winfo_children()):
            child.destroy()

        try:
            from tkinter import font as tkfont
            if not hasattr(self, "_nav_font"):
                self._nav_font = tkfont.nametofont("TkDefaultFont")
                self._nav_font_bold = self._nav_font.copy()
                self._nav_font_bold.configure(weight="bold")
        except Exception:
            self._nav_font = None
            self._nav_font_bold = None

        def add_btn(label: str, tab_text: str, kind: str, state: int = 0):
            if kind in ("onlyA", "onlyB"):
                bg = "#FFE5E5"
            else:
                # 0=none, 1=maybe (pale), 2=confirmed (bright)
                if state >= 2:
                    bg = "#FFD400"  # bright yellow
                elif state == 1:
                    bg = "#FFF3B0"  # pale yellow
                else:
                    bg = "#F2F2F2"
            is_selected = (tab_text == getattr(self, "selected_sheet", None))
            if is_selected:
                bg = "#D9D9D9"
            b = tk.Button(self.nav_inner, text=label,
                          relief="sunken" if is_selected else "groove",
                          bd=2 if is_selected else 1,
                          padx=8, pady=2, bg=bg,
                          command=lambda: self._select_tab(tab_text))
            try:
                if is_selected and self._nav_font_bold:
                    b.configure(font=self._nav_font_bold)
                elif self._nav_font:
                    b.configure(font=self._nav_font)
            except Exception:
                pass
            b.pack(side="left", padx=4)

        if self.only_a:
            add_btn(f"仅A({len(self.only_a)})", "仅在A存在", "onlyA")
        if self.only_b:
            add_btn(f"仅B({len(self.only_b)})", "仅在B存在", "onlyB")

        for s in self.common_sheets:
            add_btn(s, s, "common", state=int(self.sheet_diff_state.get(s, 0)))

        self.nav_canvas.update_idletasks()
        self.nav_canvas.configure(scrollregion=self.nav_canvas.bbox("all"))

    def open_textdiff(self):
        try:
            temp_root = os.path.join(os.environ.get("LOCALAPPDATA", tempfile.gettempdir()), "Temp", "TortoiseXlsTemp")
            os.makedirs(temp_root, exist_ok=True)
        except Exception:
            temp_root = tempfile.gettempdir()

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        left_txt = os.path.join(temp_root, f"{APP_NAME}_left_{ts}.txt")
        right_txt = os.path.join(temp_root, f"{APP_NAME}_right_{ts}.txt")
        excel_to_text(self.file_a, left_txt, thick_sep_char="=")
        excel_to_text(self.file_b, right_txt, thick_sep_char="=")
        open_tortoise_merge(left_txt, right_txt, title=f"{APP_NAME}: {os.path.basename(self.file_a)}")

    def export_diagnostic_bundle(self):
        try:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_name = f"{APP_NAME}_diag_{APP_BUILD_TAG}_{ts}.zip"
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            initial_dir = desktop if os.path.isdir(desktop) else tempfile.gettempdir()
            save_path = filedialog.asksaveasfilename(
                title="导出诊断包",
                defaultextension=".zip",
                initialdir=initial_dir,
                initialfile=default_name,
                filetypes=[("Zip Archive", "*.zip")],
            )
            if not save_path:
                return

            notes = []
            notes.append(f"app={APP_NAME}")
            notes.append(f"version={APP_VERSION}")
            notes.append(f"build={APP_BUILD_TAG}")
            notes.append(f"time={datetime.now().isoformat(timespec='seconds')}")
            notes.append(f"python={sys.version.splitlines()[0]}")
            notes.append(f"platform={platform.platform()}")
            notes.append(f"merge_mode={self.merge_mode}")
            notes.append(f"merge_conflict_mode={self.merge_conflict_mode}")
            notes.append(f"file_a={self.file_a}")
            notes.append(f"file_b={self.file_b}")
            notes.append(f"base_path={self.base_path}")
            notes.append(f"raw_mine={self.raw_mine}")
            notes.append(f"raw_base={self.raw_base}")
            notes.append(f"raw_theirs={self.raw_theirs}")

            with zipfile.ZipFile(save_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                zf.writestr("diagnostic_summary.txt", "\n".join(notes) + "\n")
                for p in (_DEBUG_LOG_PATH, _LAUNCH_TRACE_PATH, _SETTINGS_PATH):
                    try:
                        if p and os.path.exists(p):
                            zf.write(p, arcname=os.path.basename(p))
                    except Exception:
                        pass

            messagebox.showinfo("导出完成", f"诊断包已导出：\n{save_path}")
        except Exception as e:
            messagebox.showerror("导出失败", f"导出诊断包失败：\n{e}")

    def copy_feedback_info(self):
        try:
            selected_sheet = "-"
            try:
                tab_id = self.nb.select()
                if tab_id:
                    selected_sheet = self.nb.tab(tab_id, "text")
            except Exception:
                selected_sheet = "-"

            lines = [
                f"app={APP_NAME}",
                f"version={APP_VERSION}",
                f"build={APP_BUILD_TAG}",
                f"time={datetime.now().isoformat(timespec='seconds')}",
                f"merge_mode={self.merge_mode}",
                f"merge_conflict_mode={self.merge_conflict_mode}",
                f"selected_sheet={selected_sheet}",
                f"file_a={self.file_a}",
                f"file_b={self.file_b}",
                f"base_path={self.base_path}",
                f"raw_mine={self.raw_mine}",
                f"raw_base={self.raw_base}",
                f"raw_theirs={self.raw_theirs}",
                f"debug_log={_DEBUG_LOG_PATH}",
                f"launch_trace={_LAUNCH_TRACE_PATH}",
            ]
            text = "\n".join(lines)
            self.root.clipboard_clear()
            self.root.clipboard_append(text)
            self.root.update()
            messagebox.showinfo("已复制", "反馈信息已复制到剪贴板。")
        except Exception as e:
            messagebox.showerror("复制失败", f"复制反馈信息失败：\n{e}")

    def _find_svn_wc_root(self, start_dir: str) -> str | None:
        try:
            cur = os.path.abspath(start_dir)
            while True:
                if os.path.isdir(os.path.join(cur, ".svn")):
                    return cur
                parent = os.path.dirname(cur)
                if parent == cur:
                    return None
                cur = parent
        except Exception:
            return None

    def _ask_update_confirm_dialog(self, wc_root: str) -> bool:
        dlg = tk.Toplevel(self.root)
        dlg.title("检查更新")
        dlg.transient(self.root)
        dlg.resizable(False, False)

        result = {"ok": False}

        frm = ttk.Frame(dlg, padding=12)
        frm.pack(fill="both", expand=True)

        ttk.Label(frm, text=f"当前版本：{APP_VERSION} [{APP_BUILD_TAG}]").pack(anchor="w")
        ttk.Label(
            frm,
            text=(
                "将从 SVN 获取最新版本。\n"
                "确认后工具将自动关闭，更新完成后请重新启动。"
            ),
            justify="left",
            wraplength=620,
        ).pack(anchor="w", pady=(6, 0))

        tk.Label(
            frm,
            text="注意：更新流程不会保存当前工具中的未保存修改。",
            fg="#C62828",
            justify="left",
            anchor="w",
        ).pack(anchor="w", pady=(10, 0))
        tk.Label(
            frm,
            text="请先保存再更新，避免数据丢失。",
            fg="#C62828",
            justify="left",
            anchor="w",
        ).pack(anchor="w")

        ttk.Label(
            frm,
            text=f"工作副本目录：\n{wc_root}",
            justify="left",
            wraplength=620,
        ).pack(anchor="w", pady=(10, 0))

        btns = ttk.Frame(frm)
        btns.pack(fill="x", pady=(12, 0))

        def _cancel():
            result["ok"] = False
            try:
                dlg.destroy()
            except Exception:
                pass

        def _confirm():
            result["ok"] = True
            try:
                dlg.destroy()
            except Exception:
                pass

        ttk.Button(btns, text="取消", command=_cancel).pack(side="right")
        ttk.Button(btns, text="确认更新", command=_confirm).pack(side="right", padx=(0, 8))

        dlg.protocol("WM_DELETE_WINDOW", _cancel)
        try:
            dlg.update_idletasks()
            rx = self.root.winfo_rootx()
            ry = self.root.winfo_rooty()
            rw = self.root.winfo_width()
            rh = self.root.winfo_height()
            dw = dlg.winfo_width()
            dh = dlg.winfo_height()
            x = max(0, rx + (rw - dw) // 2)
            y = max(0, ry + (rh - dh) // 2)
            dlg.geometry(f"+{x}+{y}")
        except Exception:
            pass

        dlg.grab_set()
        try:
            dlg.focus_force()
        except Exception:
            pass
        dlg.wait_window()
        return bool(result.get("ok", False))

    def _do_svn_update(self):
        if getattr(self, "_update_launching", False):
            return

        def _reenable_update_btn():
            try:
                if getattr(self, "update_btn", None) is not None:
                    self.update_btn.configure(state="normal")
            except Exception:
                pass
            self._update_launching = False

        # Protect users from silently losing in-memory changes.
        if bool(getattr(self, "modified_a", False) or getattr(self, "modified_b", False)):
            if not messagebox.askyesno(
                "未保存改动",
                "检测到当前有未保存改动。\n"
                "如果继续更新并关闭工具，这些改动将丢失。\n\n"
                "仍然继续更新吗？",
            ):
                return

        # Locate tool directory: PyInstaller bundle uses sys.executable, plain script uses __file__.
        if getattr(sys, "frozen", False):
            tool_dir = os.path.dirname(sys.executable)
        else:
            tool_dir = os.path.dirname(os.path.abspath(__file__))
        tool_dir = os.path.abspath(tool_dir)

        wc_root = self._find_svn_wc_root(tool_dir)
        if not wc_root:
            messagebox.showerror(
                "更新失败",
                "当前工具目录不在 SVN 工作副本中。\n"
                "请确认工具是从 SVN checkout 目录启动。",
            )
            return
        # Keep update scope limited to the tool directory.
        update_target = tool_dir
        proc_exe = _find_tortoise_proc_exe()
        proc_exists = bool(proc_exe and os.path.exists(proc_exe))
        svn_exe = _find_svn_cli_exe()
        if not svn_exe and proc_exists:
            if not self._ask_update_confirm_dialog(update_target):
                return
            if not messagebox.askyesno(
                "更新方式",
                "未找到 svn 命令行，将改用 TortoiseSVN 图形更新。\n"
                "确认后工具将关闭，并弹出 TortoiseSVN Update 窗口。\n\n"
                "是否继续？",
            ):
                return
            self._update_launching = True
            try:
                if getattr(self, "update_btn", None) is not None:
                    self.update_btn.configure(state="disabled")
            except Exception:
                pass
            try:
                subprocess.Popen([
                    proc_exe,
                    "/command:update",
                    f"/path:{update_target}",
                    "/closeonend:0",
                ], close_fds=True)
            except Exception as e:
                _reenable_update_btn()
                messagebox.showerror("更新失败", f"无法启动 TortoiseSVN 更新：\n{e}")
                return
            try:
                self._shutdown_root()
            except Exception:
                pass
            sys.exit(0)

        if not svn_exe:
            if not proc_exists:
                messagebox.showerror(
                    "更新失败",
                    "未找到 svn 命令，且未检测到 TortoiseSVN 安装目录。\n"
                    "请安装 TortoiseSVN（含命令行工具）或把 svn.exe 加入 PATH。",
                )
            else:
                messagebox.showerror(
                    "更新失败",
                    "未找到 svn 命令。\n"
                    f"已检测到 TortoiseSVN：{proc_exe}\n"
                    "请确认存在 svn.exe（通常在 TortoiseSVN\\bin）或将其加入 PATH。",
                )
            return

        svn_ver = _query_svn_version(svn_exe)
        if not svn_ver:
            messagebox.showerror(
                "更新失败",
                "检测到 svn 可执行文件，但无法获取版本信息。\n"
                f"路径：{svn_exe}\n"
                "请检查 SVN 安装是否完整。",
            )
            return
        _dlog(f"svn detected: exe={svn_exe} version={svn_ver}")

        ps_exe = shutil.which("powershell") or shutil.which("pwsh")
        if not ps_exe:
            messagebox.showerror("更新失败", "未找到 PowerShell，无法启动更新脚本。")
            return

        no_window = getattr(subprocess, "CREATE_NO_WINDOW", 0)
        try:
            info = subprocess.run(
                [svn_exe, "info", update_target],
                capture_output=True,
                text=True,
                timeout=12,
                creationflags=no_window,
            )
        except Exception as e:
            messagebox.showerror("更新失败", f"执行 svn info 失败：\n{e}")
            return
        if info.returncode != 0:
            err = (info.stderr or info.stdout or "").strip()
            messagebox.showerror("更新失败", f"当前目录不是有效 SVN 工作副本或无权限：\n{err}")
            return

        # Ask SVN for canonical wc-root (more robust than .svn upward walk).
        try:
            wc_info = subprocess.run(
                [svn_exe, "info", "--show-item", "wc-root", wc_root],
                capture_output=True,
                text=True,
                timeout=10,
                creationflags=no_window,
            )
            if wc_info.returncode == 0:
                wc_root_svn = (wc_info.stdout or "").strip()
                if wc_root_svn:
                    wc_root = os.path.abspath(wc_root_svn)
        except Exception as e:
            _dlog(f"resolve wc-root from svn info failed: {e}")

        # Safety check: tool path must still reside under detected working copy.
        try:
            if os.path.normcase(os.path.commonpath([tool_dir, wc_root])) != os.path.normcase(wc_root):
                messagebox.showerror(
                    "更新失败",
                    "检测到工作副本路径异常。\n"
                    "为避免误更新，已中止本次操作。\n\n"
                    f"工具目录：{tool_dir}\n"
                    f"工作副本：{wc_root}",
                )
                return
        except Exception as e:
            messagebox.showerror("更新失败", f"无法校验工作副本路径关系：\n{e}")
            return

        # Preflight status: block known-bad states and warn risky states.
        # Performance optimization:
        # - First check the tool directory only (most relevant for self-update).
        # - Fallback to wc_root only if the narrow-scope status call fails.
        local_changes = []
        conflict_changes = []
        status_error = None
        status_scope_label = "工具目录"
        status_target = update_target
        try:
            status = subprocess.run(
                [svn_exe, "status", "-q", status_target],
                capture_output=True,
                text=True,
                timeout=20,
                creationflags=no_window,
            )
            if status.returncode != 0:
                # Fallback: if status on tool dir failed unexpectedly, retry on wc-root.
                status_scope_label = "工作副本"
                status_target = wc_root
                status = subprocess.run(
                    [svn_exe, "status", "-q", status_target],
                    capture_output=True,
                    text=True,
                    timeout=20,
                    creationflags=no_window,
                )
            if status.returncode == 0:
                for ln in (status.stdout or "").splitlines():
                    if not ln.strip():
                        continue
                    flags = (ln[:7] + "       ")[:7]
                    path_part = ln[8:].strip() if len(ln) > 8 else ln.strip()
                    if "C" in flags:
                        conflict_changes.append(path_part)
                        continue
                    # Ignore unversioned-only lines; these do not block update.
                    if flags[0] == "?":
                        continue
                    # Versioned local mods/add/delete/replace/tree-conflict-ish states.
                    if any(ch in "MADR~!" for ch in flags[:2]):
                        local_changes.append(path_part)
            else:
                status_error = (status.stderr or status.stdout or "").strip() or f"returncode={status.returncode}"
        except Exception as e:
            status_error = str(e)

        if status_error:
            messagebox.showerror(
                "更新前检查失败",
                f"无法完成{status_scope_label}状态检查（svn status）。\n"
                "请先在工作副本中手工执行 svn status 并处理异常后重试。\n\n"
                f"{status_error}",
            )
            return

        if conflict_changes:
            preview = "\n".join(conflict_changes[:6])
            if len(conflict_changes) > 6:
                preview += f"\n... 另有 {len(conflict_changes) - 6} 项"
            messagebox.showerror(
                "更新前检查失败",
                f"检测到{status_scope_label}存在冲突项，请先解决冲突后再更新。\n\n" + preview,
            )
            return

        if local_changes:
            preview = "\n".join(local_changes[:6])
            if len(local_changes) > 6:
                preview += f"\n... 另有 {len(local_changes) - 6} 项"
            if not messagebox.askyesno(
                "检测到本地改动",
                f"{status_scope_label}包含本地版本化改动，继续 update 可能产生冲突。\n\n"
                f"{preview}\n\n仍然继续更新吗？",
            ):
                return

        if not self._ask_update_confirm_dialog(update_target):
            return

        self._update_launching = True
        try:
            if getattr(self, "update_btn", None) is not None:
                self.update_btn.configure(state="disabled")
        except Exception:
            pass

        log_path = os.path.join(update_target, "sow_update.log")
        ps1_path = os.path.join(tempfile.gettempdir(), f"sow_update_{os.getpid()}.ps1")

        update_target_q = update_target.replace("'", "''")
        svn_exe_q = svn_exe.replace("'", "''")
        log_path_q = log_path.replace("'", "''")

        ps_script = (
            "$ErrorActionPreference = 'Continue'\n"
            f"$updateTarget = '{update_target_q}'\n"
            f"$svnExe = '{svn_exe_q}'\n"
            f"$logPath = '{log_path_q}'\n"
            "$maxTry = 3\n"
            "Start-Sleep -Seconds 2\n"
            "if (-not (Test-Path -LiteralPath $updateTarget)) { Write-Host '更新目录不存在，更新终止。'; exit 2 }\n"
            "Set-Location -LiteralPath $updateTarget\n"
            "Add-Content -LiteralPath $logPath -Value (\"===== {0} update start =====\" -f (Get-Date -Format 'yyyy-MM-dd HH:mm:ss'))\n"
            "$ok = $false\n"
            "$lastCode = 0\n"
            "for ($i = 1; $i -le $maxTry; $i++) {\n"
            "  Add-Content -LiteralPath $logPath -Value (\"[try {0}] svn cleanup\" -f $i)\n"
            "  & $svnExe cleanup --non-interactive \"$updateTarget\" >> $logPath 2>&1\n"
            "  Add-Content -LiteralPath $logPath -Value (\"[try {0}] svn update\" -f $i)\n"
            "  & $svnExe update --non-interactive \"$updateTarget\" >> $logPath 2>&1\n"
            "  $lastCode = $LASTEXITCODE\n"
            "  if ($lastCode -eq 0) { $ok = $true; break }\n"
            "  Start-Sleep -Seconds 2\n"
            "}\n"
            "if ($ok) {\n"
            "  Add-Content -LiteralPath $logPath -Value 'update succeeded'\n"
            "  Write-Host ''\n"
            "  Write-Host '更新成功，请重新启动工具。'\n"
            "} else {\n"
            "  Add-Content -LiteralPath $logPath -Value (\"update failed rc={0}\" -f $lastCode)\n"
            "  Write-Host ''\n"
            "  Write-Host '更新失败，请查看日志：'\n"
            "  Write-Host $logPath\n"
            "}\n"
            "Write-Host ''\n"
            "Write-Host '按回车关闭窗口...'\n"
            "[void][System.Console]::ReadLine()\n"
            "Remove-Item -LiteralPath $PSCommandPath -Force -ErrorAction SilentlyContinue\n"
        )

        try:
            with open(ps1_path, "w", encoding="utf-8-sig") as f:
                f.write(ps_script)
            new_console = getattr(subprocess, "CREATE_NEW_CONSOLE", 0)
            subprocess.Popen(
                [ps_exe, "-NoProfile", "-ExecutionPolicy", "Bypass", "-File", ps1_path],
                creationflags=new_console,
                close_fds=True,
            )
        except Exception as e:
            try:
                os.remove(ps1_path)
            except Exception:
                pass
            _reenable_update_btn()
            messagebox.showerror("更新失败", f"无法启动更新脚本：\n{e}")
            return

        try:
            self._shutdown_root()
        except Exception:
            pass
        sys.exit(0)

    def recalc_and_refresh(self):
        # Manual: force Excel recalc to refresh cached values, then reload view.
        def _do_recalc():
            new_a = _recalc_and_prepare_val_path(self.file_a)
            new_b = _recalc_and_prepare_val_path(self.file_b)
            new_base = _recalc_and_prepare_val_path(self.base_path) if getattr(self, "has_base", False) else None
            if new_a:
                _loaded = load_workbook(new_a, data_only=True)
                _wbs_close(getattr(self, "_wb_a_val", None))
                self._file_a_val_path = new_a
                self._wb_a_val = _loaded
            if new_b:
                _loaded = load_workbook(new_b, data_only=True)
                _wbs_close(getattr(self, "_wb_b_val", None))
                self._file_b_val_path = new_b
                self._wb_b_val = _loaded
            if new_base and getattr(self, "has_base", False):
                _loaded = load_workbook(new_base, data_only=True)
                _wbs_close(getattr(self, "_wb_base_val", None))
                self._file_base_val_path = new_base
                self._wb_base_val = _loaded

            # Refresh current sheet immediately
            try:
                tab_id = self.nb.select()
                tab_text = self.nb.tab(tab_id, "text")
                view = self.sheet_views.get(tab_text)
                if view:
                    view.refresh(row_only=None, rescan=True)
            except Exception:
                pass
            # Recompute diff states in background
            try:
                for s in self.common_sheets:
                    self.set_sheet_has_diff(s, False, confirmed=False)
                with self._compute_lock:
                    self._compute_queue = [s for s in self.common_sheets if s not in self._compute_inflight]
                self._kick_worker()
            except Exception:
                pass

        try:
            self._with_progress("重算中", "正在重算并刷新，请稍候...", _do_recalc)
        except Exception as e:
            messagebox.showerror("重算失败", f"重算失败：\n{e}")

    def _schedule_auto_recalc(self):
        if not (_AUTO_RECALC_ON_OPEN and _USE_CACHED_VALUES_ONLY):
            return
        if not getattr(self, "merge_mode", False):
            return
        if self._auto_recalc_started:
            return
        self._auto_recalc_started = True

        def _worker():
            try:
                new_a = _recalc_and_prepare_val_path(self.file_a)
                new_b = _recalc_and_prepare_val_path(self.file_b)
                new_base = _recalc_and_prepare_val_path(self.base_path) if getattr(self, "has_base", False) else None
            except Exception:
                new_a = None
                new_b = None
                new_base = None

            if not (new_a or new_b or new_base):
                return

            def _apply():
                try:
                    if new_a:
                        _loaded = load_workbook(new_a, data_only=True)
                        _wbs_close(getattr(self, "_wb_a_val", None))
                        self._file_a_val_path = new_a
                        self._wb_a_val = _loaded
                    if new_b:
                        _loaded = load_workbook(new_b, data_only=True)
                        _wbs_close(getattr(self, "_wb_b_val", None))
                        self._file_b_val_path = new_b
                        self._wb_b_val = _loaded
                    if new_base and getattr(self, "has_base", False):
                        _loaded = load_workbook(new_base, data_only=True)
                        _wbs_close(getattr(self, "_wb_base_val", None))
                        self._file_base_val_path = new_base
                        self._wb_base_val = _loaded

                    try:
                        tab_id = self.nb.select()
                        tab_text = self.nb.tab(tab_id, "text")
                        view = self.sheet_views.get(tab_text)
                        if view:
                            view.refresh(row_only=None, rescan=True)
                    except Exception:
                        pass

                    try:
                        for s in self.common_sheets:
                            self.set_sheet_has_diff(s, False, confirmed=False)
                        with self._compute_lock:
                            self._compute_queue = [s for s in self.common_sheets if s not in self._compute_inflight]
                        self._kick_worker()
                    except Exception:
                        pass
                except Exception as e:
                    _dlog(f"auto recalc apply failed: {e}")

            try:
                self._safe_root_after(0, _apply)
            except Exception:
                pass

        threading.Thread(target=_worker, daemon=True).start()

    def _with_progress(self, title: str, message: str, fn):
        dlg = tk.Toplevel(self.root)
        dlg.title(title)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.resizable(False, False)
        dlg.geometry("+{}+{}".format(self.root.winfo_rootx() + 200, self.root.winfo_rooty() + 150))
        ttk.Label(dlg, text=message, padding=10).pack(fill="x")
        pb = ttk.Progressbar(dlg, mode="indeterminate")
        pb.pack(fill="x", padx=10, pady=(0, 10))
        pb.start(10)
        self.root.update_idletasks()
        try:
            fn()
        finally:
            try:
                pb.stop()
                dlg.destroy()
            except Exception:
                pass

    def _atomic_save(self, wb, target_path: str):
        """Safely overwrite a workbook.

        Writes to a temp file in the same directory, then os.replace.
        This avoids corrupting the target if the process is interrupted.
        """
        ext_parts = _capture_external_link_parts(target_path)
        # If we need to preserve external-link parts, force tmp-path save to avoid
        # a second replace on the final target (often locked by SVN shell after save).
        use_fast_direct = _FAST_SAVE_ENABLED and not ext_parts
        if use_fast_direct:
            # Fast path: write directly (faster, but not atomic)
            try:
                if _FAST_SAVE_VALUES_ONLY and _USE_CACHED_VALUES_ONLY:
                    _save_values_only_from_wb(wb, target_path)
                else:
                    wb.save(target_path)
                return
            except PermissionError:
                # fallback to safe path below
                pass
        folder = os.path.dirname(target_path)
        base = os.path.basename(target_path)
        tmp_path = os.path.join(folder, f"~{base}.{os.getpid()}.tmp")
        if _FAST_SAVE_VALUES_ONLY and _USE_CACHED_VALUES_ONLY:
            _save_values_only_from_wb(wb, tmp_path)
        else:
            wb.save(tmp_path)
        if ext_parts:
            _apply_external_link_parts_on_file(tmp_path, ext_parts)
        try:
            os.replace(tmp_path, target_path)
            return
        except PermissionError:
            # Try clearing readonly flag then retry a few times (file may be locked briefly)
            try:
                if os.path.exists(target_path):
                    os.chmod(target_path, stat.S_IWRITE | stat.S_IREAD)
            except Exception:
                pass
            # If we can delete the target, try that once (replace may fail on readonly)
            try:
                if os.path.exists(target_path):
                    os.remove(target_path)
            except Exception:
                pass
            for _ in range(8):
                try:
                    os.replace(tmp_path, target_path)
                    return
                except PermissionError:
                    time.sleep(0.5)
            # If replace keeps failing, try overwrite-in-place (requires write but not delete)
            try:
                with open(tmp_path, "rb") as src, open(target_path, "wb") as dst:
                    shutil.copyfileobj(src, dst, length=1024 * 1024)
                return
            except Exception:
                raise
        except Exception:
            # Last-resort fallback to shutil.move
            try:
                shutil.move(tmp_path, target_path)
                return
            except Exception:
                raise
        finally:
            try:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
            except Exception:
                pass

    def _atomic_save_with_retry(self, wb, target_path: str, retries: int = 6, delay_sec: float = 0.5):
        """Retry save when target is temporarily locked."""
        last_err = None
        for _ in range(max(1, retries)):
            try:
                self._atomic_save(wb, target_path)
                return
            except Exception as e:
                if getattr(e, "winerror", None) in (5, 32, 33) or isinstance(e, PermissionError):
                    last_err = e
                    time.sleep(delay_sec)
                    continue
                raise
        if last_err:
            raise last_err

    def _atomic_replace_file(self, src_path: str, target_path: str):
        """Safely replace target with an existing source file (no openpyxl roundtrip)."""
        folder = os.path.dirname(target_path)
        base = os.path.basename(target_path)
        tmp_path = os.path.join(folder, f"~{base}.{os.getpid()}.tmp")
        shutil.copy2(src_path, tmp_path)
        try:
            os.replace(tmp_path, target_path)
            return
        except PermissionError:
            try:
                if os.path.exists(target_path):
                    os.chmod(target_path, stat.S_IWRITE | stat.S_IREAD)
            except Exception:
                pass
            try:
                if os.path.exists(target_path):
                    os.remove(target_path)
            except Exception:
                pass
            for _ in range(8):
                try:
                    os.replace(tmp_path, target_path)
                    return
                except PermissionError:
                    time.sleep(0.5)
            with open(tmp_path, "rb") as src, open(target_path, "wb") as dst:
                shutil.copyfileobj(src, dst, length=1024 * 1024)
        finally:
            try:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
            except Exception:
                pass

    def _atomic_replace_file_with_retry(self, src_path: str, target_path: str, retries: int = 6, delay_sec: float = 0.5):
        last_err = None
        for _ in range(max(1, retries)):
            try:
                self._atomic_replace_file(src_path, target_path)
                return
            except Exception as e:
                if getattr(e, "winerror", None) in (5, 32, 33) or isinstance(e, PermissionError):
                    last_err = e
                    time.sleep(delay_sec)
                    continue
                raise
        if last_err:
            raise last_err

    def _alt_save_path(self, path: str, which: str):
        folder = os.path.dirname(path)
        base, ext = os.path.splitext(os.path.basename(path))
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        return os.path.join(folder, f"{base}_{which}_saved_{ts}{ext or '.xlsx'}")

    def _try_alt_save(self, wb, path: str, which: str) -> bool:
        alt = self._alt_save_path(path, which)
        try:
            self._atomic_save(wb, alt)
            messagebox.showinfo("另存为成功", f"无法覆盖原文件，已另存为：\n{alt}")
            return True
        except Exception as e:
            messagebox.showerror("另存为失败", f"另存为失败：\n{e}")
            return False

    def _path_diagnostics(self, path: str) -> str:
        try:
            folder = os.path.dirname(path)
            exists = os.path.exists(path)
            readonly = False
            if exists:
                try:
                    readonly = not os.access(path, os.W_OK)
                except Exception:
                    readonly = False
            dir_writable = False
            try:
                test_file = os.path.join(folder, f"~perm_test_{os.getpid()}.tmp")
                with open(test_file, "w", encoding="utf-8") as f:
                    f.write("x")
                os.remove(test_file)
                dir_writable = True
            except Exception:
                dir_writable = False
            return f"exists={exists}, readonly={readonly}, dir_writable={dir_writable}"
        except Exception:
            return "diagnostics_failed"

    def _confirm_overwrite(self, which: str, path: str) -> bool:
        if which == "A":
            modified = self.modified_a
        else:
            modified = self.modified_b

        if not modified:
            return messagebox.askyesno("提示", f"{which} 没有检测到改动。仍然要覆盖保存原文件吗？\n\n{path}")

        return messagebox.askyesno(
            "确认保存",
            f"将直接覆盖保存 {which} 文件（原路径、原文件名）：\n\n{path}\n\n建议确保该 Excel 未在 WPS/Excel 中打开。继续吗？",
        )

    def save_b_inplace(self):
        self._ensure_edit_loaded()
        path = self.file_b
        if not self._confirm_overwrite("B", path):
            return
        try:
            self._with_progress("保存中", f"正在保存：\n{path}", lambda: self._atomic_save(self._wb_b_edit, path))
            self.modified_b = False
            messagebox.showinfo("Saved", f"已保存并覆盖：\n{path}")
        except Exception as e:
            # If the file is locked or denied, offer save-as fallback
            if getattr(e, "winerror", None) in (5, 32, 33) or isinstance(e, PermissionError):
                diag = self._path_diagnostics(path)
                if messagebox.askyesno("保存失败", f"保存 B 失败（可能文件被占用/无权限）：\n{e}\n\n诊断：{diag}\n\n是否另存为？"):
                    if self._try_alt_save(self._wb_b_edit, path, "B"):
                        self.modified_b = False
                        return
            messagebox.showerror("保存失败", f"保存 B 失败：\n{e}")

    def save_a_inplace(self):
        self._ensure_edit_loaded()
        path = self.file_a
        if not self._confirm_overwrite("A", path):
            return
        try:
            self._with_progress("保存中", f"正在保存：\n{path}", lambda: self._atomic_save(self._wb_a_edit, path))
            self.modified_a = False
            messagebox.showinfo("Saved", f"已保存并覆盖：\n{path}")
        except Exception as e:
            if getattr(e, "winerror", None) in (5, 32, 33) or isinstance(e, PermissionError):
                diag = self._path_diagnostics(path)
                if messagebox.askyesno("保存失败", f"保存 A 失败（可能文件被占用/无权限）：\n{e}\n\n诊断：{diag}\n\n是否另存为？"):
                    if self._try_alt_save(self._wb_a_edit, path, "A"):
                        self.modified_a = False
                        return
            messagebox.showerror("保存失败", f"保存 A 失败：\n{e}")

    def save_merged_and_exit(self, auto: bool = False):
        if not self.merged_path:
            return
        self._ensure_edit_loaded()
        if not auto:
            if self.merge_mode and self.initial_conflict_cell_count > 0:
                unresolved = sum(
                    len(cols)
                    for rows in self.merge_conflict_cells_by_sheet.values()
                    for cols in rows.values()
                )
                if not messagebox.askyesno(
                    "确认冲突处理",
                    f"三方扫描检测到 {self.initial_conflict_cell_count} 个冲突单元格。"
                    f"\n当前仍标记 {unresolved} 个（手动模式下不会自动清零）。"
                    "\n\n请确认你已完成需要处理的冲突数据。是否继续保存？",
                ):
                    return
            if not messagebox.askyesno("确认保存", f"将保存合并结果到：\n\n{self.merged_path}\n\n继续吗？"):
                return
        wb_to_save = self._wb_a_edit
        merged_source_path = None
        try:
            # Small delay to allow SVN/Tortoise to release file locks
            try:
                time.sleep(1.2)
            except Exception:
                pass
            if self.merge_mode and self.has_base:
                # Safety guard for manual 3-way mode:
                # save from pristine mine + explicit operations only.
                merged_source_path = self.build_manual_merge_output_file()
            try:
                by_sheet = {}
                for (s, _r, _c) in getattr(self, "manual_a_cell_ops", {}).keys():
                    by_sheet[s] = by_sheet.get(s, 0) + 1
                _dlog(
                    f"SAVE_MERGED path={self.merged_path} "
                    f"merge_mode={self.merge_mode} has_base={self.has_base} "
                    f"manual_a_ops={len(getattr(self, 'manual_a_cell_ops', {}))} "
                    f"manual_ops_by_sheet={by_sheet} "
                    f"snapshot={getattr(self, '_merge_mine_snapshot', None)}"
                )
            except Exception:
                pass
            if merged_source_path:
                self._with_progress("保存中", f"正在保存合并结果：\n{self.merged_path}",
                                    lambda: self._atomic_replace_file_with_retry(merged_source_path, self.merged_path))
            else:
                self._with_progress("保存中", f"正在保存合并结果：\n{self.merged_path}",
                                    lambda: self._atomic_save_with_retry(wb_to_save, self.merged_path))
            self.modified_a = False
            try:
                messagebox.showinfo("Saved", f"已保存合并结果：\n{self.merged_path}")
            except Exception:
                pass
        except Exception as e:
            if getattr(e, "winerror", None) in (5, 32, 33) or isinstance(e, PermissionError):
                excel_locked = _log_lock_holders(self.merged_path)
                if excel_locked:
                    try:
                        messagebox.showwarning("文件被占用", "检测到 Excel 正在占用目标文件。\n请关闭 Excel 后再保存。")
                    except Exception:
                        pass
                # In conflict UI, target file might still be locked by SVN/Tortoise.
                # Save to a temp file and schedule a deferred replace.
                try:
                    folder = os.path.dirname(self.merged_path)
                    base = os.path.basename(self.merged_path)
                    tmp_path = os.path.join(folder, f"~{base}.deferred.{os.getpid()}.tmp")
                    if merged_source_path:
                        shutil.copy2(merged_source_path, tmp_path)
                    elif _FAST_SAVE_VALUES_ONLY and _USE_CACHED_VALUES_ONLY:
                        _save_values_only_from_wb(wb_to_save, tmp_path)
                    else:
                        wb_to_save.save(tmp_path)
                    if not os.path.exists(tmp_path) or os.path.getsize(tmp_path) == 0:
                        raise RuntimeError(f"临时文件写入失败或为空：{tmp_path}")
                    _launch_deferred_copy(tmp_path, self.merged_path)
                    messagebox.showinfo("保存中", f"目标文件被占用，已写入临时文件并将在关闭后自动覆盖：\n{self.merged_path}")
                    try:
                        self._shutdown_root()
                    except Exception:
                        pass
                    sys.exit(0)
                except Exception:
                    diag = self._path_diagnostics(self.merged_path)
                    if messagebox.askyesno("保存失败", f"保存合并结果失败（可能文件被占用/无权限）：\n{e}\n\n诊断：{diag}\n\n是否另存为？"):
                        alt_ok = False
                        if merged_source_path and os.path.exists(merged_source_path):
                            alt = self._alt_save_path(self.merged_path, "MERGED")
                            try:
                                shutil.copy2(merged_source_path, alt)
                                messagebox.showinfo("另存为成功", f"无法覆盖原文件，已另存为：\n{alt}")
                                alt_ok = True
                            except Exception:
                                pass
                        if not alt_ok:
                            alt_ok = self._try_alt_save(wb_to_save, self.merged_path, "MERGED")
                        if alt_ok:
                            try:
                                self._shutdown_root()
                            except Exception:
                                pass
                            sys.exit(0)
                        return
            messagebox.showerror("保存失败", f"保存合并结果失败：\n{e}")
            return
        finally:
            if merged_source_path:
                try:
                    os.remove(merged_source_path)
                except Exception:
                    pass
        # Try auto-resolve in SVN if conflict artifacts exist
        try:
            if _has_svn_conflict_artifacts(self.merged_path):
                _try_svn_resolve(self.merged_path)
        except Exception:
            pass
        try:
            self._shutdown_root()
        except Exception:
            pass
        sys.exit(0)

    def resolve_conflict_cell(self, sheet: str, r: int, c: int) -> bool:
        rows = self.merge_conflict_cells_by_sheet.get(sheet)
        if not rows:
            return False
        cols = rows.get(r)
        if not cols or c not in cols:
            return False
        cols.discard(c)
        if not cols:
            rows.pop(r, None)
        if not rows:
            self.merge_conflict_cells_by_sheet.pop(sheet, None)
        self._auto_save_if_no_conflicts()
        return True

    def resolve_conflict_row(self, sheet: str, r: int, cols) -> bool:
        rows = self.merge_conflict_cells_by_sheet.get(sheet)
        if not rows or r not in rows:
            return False
        for c in list(cols):
            rows[r].discard(c)
        if not rows[r]:
            rows.pop(r, None)
        if not rows:
            self.merge_conflict_cells_by_sheet.pop(sheet, None)
        self._auto_save_if_no_conflicts()
        return True

    def _auto_save_if_no_conflicts(self):
        if not self.merge_conflict_cells_by_sheet:
            # If user has manually touched conflicts, require explicit save.
            if getattr(self, "user_touched_conflicts", False):
                return
            # all conflicts resolved
            self.save_merged_and_exit(auto=True)

    def run(self):
        self.root.mainloop()


def main():
    try:
        try:
            _trace_launch("=" * 72)
            _trace_launch(f"cwd={os.getcwd()}")
            _trace_launch(f"argv={repr(sys.argv)}")
        except Exception:
            pass
        # Log raw args early for troubleshooting (even if argparse fails)
        try:
            _dlog(f"argv: {' '.join(sys.argv[1:])}")
        except Exception:
            pass

        def _parse_slash_args(argv):
            out = {}
            keys = ("path", "path2", "base", "mine", "theirs", "merged")
            i = 0
            n = len(argv)
            while i < n:
                a = argv[i]
                la = a.lower()
                matched = False
                for k in keys:
                    p1 = f"/{k}:"
                    p2 = f"/{k}="
                    p3 = f"-{k}:"
                    p4 = f"-{k}="
                    p5 = f"/{k}"
                    p6 = f"-{k}"
                    if la.startswith(p1) or la.startswith(p3):
                        # Colon-delimited: /key:value or -key:value
                        out[k] = a.split(":", 1)[1]
                        matched = True
                        break
                    elif la.startswith(p2) or la.startswith(p4):
                        # Equals-delimited: /key=value or -key=value
                        # Must split on "=" only; ":" in the value is a drive-letter separator on Windows.
                        out[k] = a.split("=", 1)[1]
                        matched = True
                        break
                    if la == p5 or la == p6:
                        if i + 1 < n:
                            out[k] = argv[i + 1]
                            i += 1
                        matched = True
                        break
                i += 1
            return out

        slash_args = _parse_slash_args(sys.argv[1:])
        try:
            _trace_launch(f"slash_args={repr(slash_args)}")
        except Exception:
            pass

        parser = argparse.ArgumentParser(add_help=True)
        parser.add_argument("file_a", nargs="?")
        parser.add_argument("file_b", nargs="?")
        # SVN/TortoiseSVN style args
        parser.add_argument("--base")
        parser.add_argument("--mine")
        parser.add_argument("--theirs")
        parser.add_argument("--merged")
        parser.add_argument("--title")
        parser.add_argument("--textdiff", action="store_true", help="Only generate text files and open TortoiseMerge")
        args, unknown = parser.parse_known_args()
        try:
            _trace_launch(f"argparse={repr(vars(args))} unknown={repr(unknown)}")
        except Exception:
            pass
        if unknown:
            try:
                _dlog(f"unknown args: {unknown}")
            except Exception:
                pass

        # Map /path:/path2:/base: style args (TortoiseProc)
        if not args.base and "base" in slash_args:
            args.base = slash_args.get("base")
        if not args.mine and "mine" in slash_args:
            args.mine = slash_args.get("mine")
        if not args.theirs and "theirs" in slash_args:
            args.theirs = slash_args.get("theirs")
        if not args.merged and "merged" in slash_args:
            args.merged = slash_args.get("merged")
        if not args.file_a and "path" in slash_args:
            args.file_a = slash_args.get("path")
        if not args.file_b and "path2" in slash_args:
            args.file_b = slash_args.get("path2")
        # Fallback: some launchers pass paths as plain unknown tokens.
        # Try extracting existing filesystem paths from unknown args.
        if (not args.file_a) and unknown:
            path_tokens = []
            for u in unknown:
                if not u:
                    continue
                su = str(u).strip().strip('"')
                if not su or su.startswith("-") or su.startswith("/"):
                    continue
                try:
                    if os.path.exists(su):
                        path_tokens.append(su)
                except Exception:
                    pass
            if path_tokens:
                args.file_a = path_tokens[0]
                if len(path_tokens) >= 2:
                    args.file_b = path_tokens[1]
        try:
            _trace_launch(
                "resolved args: "
                + f"file_a={repr(args.file_a)} file_b={repr(args.file_b)} "
                + f"base={repr(args.base)} mine={repr(args.mine)} "
                + f"theirs={repr(args.theirs)} merged={repr(args.merged)}"
            )
        except Exception:
            pass

        # Map SVN-style args to our 2-pane viewer (diff mode) / merge mode.
        if args.base and args.mine and args.theirs and args.merged:
            # Full 3-way merge args are already provided; do not fall back to file picker.
            a, b = None, None
        elif args.base and args.mine and not args.theirs:
            # Diff mode: keep base(revision/older) on left, mine(working-copy) on right.
            a, b = args.base, args.mine
        elif args.file_a and args.file_b:
            a, b = args.file_a, args.file_b
        elif args.file_a and (not args.file_b) and (not args.base) and (not args.mine) and (not args.theirs):
            # Single file provided (e.g., from Explorer/TortoiseSVN). If it's a conflicted file, auto-merge it.
            conflict = _detect_svn_conflict_files(args.file_a)
            if (not conflict) and args.file_a:
                try:
                    auto_target = _find_conflict_in_dir(os.path.dirname(os.path.abspath(args.file_a)))
                    if auto_target:
                        conflict = _detect_svn_conflict_files(auto_target)
                except Exception:
                    conflict = conflict
            if conflict:
                args.base, args.mine, args.theirs, args.merged = conflict
                args.force_ui = True
            else:
                a, b = args.file_a, None
        else:
            sel = pick_files_or_conflict()
            if not sel:
                return
            if sel[0] == "merge":
                _mode, base_p, mine_p, theirs_p, merged_p, force_ui = sel
                args.base = _ensure_xlsx_copy(base_p)
                args.mine = _ensure_xlsx_copy(mine_p)
                args.theirs = _ensure_xlsx_copy(theirs_p)
                args.merged = merged_p
                args.force_ui = bool(force_ui)
            else:
                _mode, a, b = sel

        if args.file_a and (args.file_b is None) and (not args.base) and b is None:
            # Need second file for diff mode
            root = tk.Tk()
            root.withdraw()
            b = filedialog.askopenfilename(title="Select second .xlsx file (same filename)", filetypes=[("Excel Workbook", "*.xlsx")])
            if not b:
                root.destroy()
                return
            if os.path.basename(args.file_a).lower() != os.path.basename(b).lower():
                messagebox.showerror(
                    "Filename mismatch",
                    f"The two files must have the same filename.\n\nA: {os.path.basename(args.file_a)}\nB: {os.path.basename(b)}",
                )
                root.destroy()
                return
            root.destroy()
            a = args.file_a

        raw_base_arg = args.base
        raw_mine_arg = args.mine
        raw_theirs_arg = args.theirs

        # Normalize SVN merge temp files (merge-left/right.r####) by exporting true revision.
        # IMPORTANT:
        # - base/theirs may legitimately be revision snapshots.
        # - mine must stay as the working-copy side; do NOT rewrite mine to a revision export,
        #   otherwise local edits can be replaced by an old revision file.
        if args.base:
            base_from_wc = None
            try:
                # Prefer WC BASE from the conflicted working-copy file.
                if args.merged:
                    base_from_wc = _try_export_svn_base_from_working_copy(args.merged)
            except Exception:
                base_from_wc = None
            if base_from_wc:
                try:
                    _dlog(f"merge base selected from WC BASE: {base_from_wc}")
                except Exception:
                    pass
                args.base = base_from_wc
                try:
                    mine_for_note = raw_mine_arg or args.mine or args.merged or "-"
                    raw_base_arg = f"{mine_for_note}@BASE(.svn)"
                except Exception:
                    raw_base_arg = "BASE(.svn)"
            else:
                args.base = _try_export_svn_revision_from_merge_temp(args.base)
        if args.theirs:
            # In merge mode, keep "theirs" exactly as passed by SVN/Tortoise wrapper.
            # This avoids accidental re-export to another revision snapshot and ensures
            # content matches the user-visible *.merge-right.r#### sidecar file.
            if not (args.base and args.mine and args.merged):
                try:
                    args.theirs = _try_export_svn_revision_from_merge_temp(args.theirs)
                except Exception:
                    args.theirs = args.theirs
        if args.file_a:
            args.file_a = _try_export_svn_revision_from_merge_temp(args.file_a)
        if args.file_b:
            args.file_b = _try_export_svn_revision_from_merge_temp(args.file_b)
        try:
            _trace_launch(
                "normalized args: "
                + f"base={repr(args.base)} mine={repr(args.mine)} "
                + f"theirs={repr(args.theirs)} merged={repr(args.merged)} "
                + f"raw_base={repr(raw_base_arg)} raw_theirs={repr(raw_theirs_arg)}"
            )
        except Exception:
            pass

        # Merge mode (manual 3-way): detect conflicts only; do NOT pre-merge before UI.
        if args.base and args.mine and args.theirs and args.merged:
            conflicts = []
            conflict_map = {}
            try:
                _dlog(f"merge args: base={args.base} mine={args.mine} theirs={args.theirs} merged={args.merged}")
                _dlog(f"merge manual mode unknown={unknown}")
            except Exception:
                pass
            try:
                # Ensure .r#### / .mine files are converted to temp .xlsx for openpyxl
                args.base = _ensure_xlsx_copy(args.base)
                args.mine = _ensure_xlsx_copy(args.mine)
                args.theirs = _ensure_xlsx_copy(args.theirs)
                try:
                    _dlog("merge start: calling _scan_three_way_conflicts (no pre-merge)")
                except Exception:
                    pass
                conflicts, conflict_map = _scan_three_way_conflicts(args.base, args.mine, args.theirs)
                try:
                    _dlog(f"merge scan result: conflicts={len(conflicts)} conflict_sheets={len(conflict_map)}")
                except Exception:
                    pass
            except Exception as e:
                try:
                    _dlog(f"merge exception: {e}")
                except Exception:
                    pass
                try:
                    messagebox.showerror("Merge failed", f"合并失败：\n{e}")
                except Exception:
                    print(f"Merge failed: {e}", file=sys.stderr)
                sys.exit(1)

            if conflicts:
                _show_conflict_popup(conflicts)

                try:
                    messagebox.showinfo(
                        "进入手动处理",
                        f"检测到 {len(conflicts)} 个冲突单元格。\n将进入手动 3 视图处理界面。",
                    )
                except Exception:
                    pass
            else:
                try:
                    messagebox.showinfo(
                        "进入手动处理",
                        "未检测到直接冲突。\n仍将进入手动 3 视图，所有差异由你确认后保存。"
                    )
                except Exception:
                    pass

            app = SowMergeApp(
                args.mine,
                args.theirs,
                merge_mode=True,
                merged_path=args.merged,
                base_path=args.base,
                merge_conflict_cells_by_sheet=conflict_map,
                merge_conflict_mode=False,
                raw_base=raw_base_arg,
                raw_mine=raw_mine_arg,
                raw_theirs=raw_theirs_arg,
            )
            try:
                _dlog("open UI: manual 3-way mode")
            except Exception:
                pass
            app.run()
            sys.exit(0)

        if args.textdiff:
            try:
                temp_root = os.path.join(os.environ.get("LOCALAPPDATA", tempfile.gettempdir()), "Temp", "TortoiseXlsTemp")
                os.makedirs(temp_root, exist_ok=True)
            except Exception:
                temp_root = tempfile.gettempdir()

            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            left_txt = os.path.join(temp_root, f"{APP_NAME}_left_{ts}.txt")
            right_txt = os.path.join(temp_root, f"{APP_NAME}_right_{ts}.txt")
            excel_to_text(a, left_txt, thick_sep_char="=")
            excel_to_text(b, right_txt, thick_sep_char="=")
            open_tortoise_merge(left_txt, right_txt, title=f"{APP_NAME}: {os.path.basename(a)}")
            return

        app = SowMergeApp(
            a,
            b,
            raw_base=raw_base_arg,
            raw_mine=raw_mine_arg,
            raw_theirs=raw_theirs_arg,
        )
        app.run()

    except Exception:
        err = traceback.format_exc()
        try:
            messagebox.showerror("Error", err)
        except Exception:
            print(err, file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
